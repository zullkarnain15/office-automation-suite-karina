"""
=========================================================
Office Automation Suite - Karina (OAS-K)
---------------------------------------------------------
File        : gui.py
Module      : Attendance
Version     : 1.0.0
Author      : OpenAI & Zulkarnain Shiddiq
Python      : 3.14+
=========================================================
Attendance GUI

This module contains user interface, light input validation,
date picker, and integration to Attendance Process Engine.

No attendance business logic is allowed here.

=========================================================
"""

from __future__ import annotations

import calendar
import tkinter as tk
from datetime import date
from datetime import datetime
from pathlib import Path
from tkinter import filedialog
from tkinter import ttk
from tkinter.scrolledtext import ScrolledText
from typing import Any

from openpyxl import load_workbook

from attendance.engine import AttendanceProcessEngine
from config.app_config import ATTENDANCE_ICON
from config.app_config import DATE_FORMAT
from config.ui_config import BACKGROUND_COLOR
from config.ui_config import BORDER_COLOR
from config.ui_config import BUTTON_FONT
from config.ui_config import CARD_COLOR
from config.ui_config import DEFAULT_FONT
from config.ui_config import HEADER_FONT
from config.ui_config import PRIMARY_COLOR
from config.ui_config import SECONDARY_COLOR
from config.ui_config import SUCCESS_COLOR
from config.ui_config import TEXT_PRIMARY
from config.ui_config import TEXT_SECONDARY
from shared.config_manager import AttendanceConfigurationReader
from shared.dialogs import Dialog
from shared.logger import get_logger
from shared.validators import validate_extension
from shared.validators import validate_file
from shared.validators import validate_folder
from shared.validators import validate_required

logger = get_logger(__name__)

APP_BACKGROUND = "#EAF2FA"
APP_PANEL = "#F8FBFF"
APP_SURFACE = "#FFFFFF"
APP_SURFACE_ACTIVE = "#E2F2F5"
APP_INPUT = "#F4F7FB"
APP_BORDER = "#C7D5E6"
APP_TEXT = "#102A43"
APP_MUTED_TEXT = "#60758A"
APP_ACCENT = "#198FA3"
APP_ACCENT_HOVER = "#123B63"
APP_SUCCESS = "#43A58F"
APP_SUCCESS_ACTIVE = "#2E8775"
APP_SUCCESS_BORDER = "#247363"
WORKFLOW_ACCENT = "#B45309"
WORKFLOW_ACCENT_HOVER = "#92400E"
APP_LOG_BG = "#24384C"
APP_LOG_FG = "#F4F7FB"
APP_SOFT_ACCENT = "#DDF3F3"
APP_TITLE_FONT = ("Segoe UI", 24, "bold")
APP_SECTION_FONT = ("Segoe UI", 12, "bold")
DEFAULT_FONT = ("Segoe UI", 9)
BUTTON_FONT = ("Segoe UI", 9, "bold")


class AttendanceGUI:
    """Attendance Module GUI."""

    def __init__(self, master: tk.Toplevel) -> None:
        self.master = master

        self.master.title("OAS-K | Attendance Module - by ZSH")
        self.master.geometry("1180x700")
        self.master.minsize(1000, 650)
        self.master.configure(bg=APP_BACKGROUND)

        try:
            self.master.iconbitmap(ATTENDANCE_ICON)
        except Exception:
            logger.warning("Attendance icon could not be loaded.")

        self.workflow_var = tk.StringVar(value="HO")
        self.use_config_output_var = tk.BooleanVar(value=True)
        self.use_config_date_var = tk.BooleanVar(value=False)
        self.generate_txt_var = tk.BooleanVar(value=True)
        self.generate_report_var = tk.BooleanVar(value=True)

        self._configure_ttk_style()
        self._create_widgets()
        self._apply_widget_theme(self.master)
        self._style_primary_action()
        self._update_workflow_selector()

    # =====================================================
    # GUI
    # =====================================================

    def _create_widgets(self) -> None:
        self.header_label = tk.Label(
            self.master,
            text="Attendance Module",
            font=APP_TITLE_FONT,
            bg=APP_BACKGROUND,
            fg=APP_ACCENT_HOVER,
        )

        self.header_label.pack(pady=(8, 0))

        self.subtitle_label = tk.Label(
            self.master,
            text="Prepare HRIS-ready attendance files and reports",
            font=("Segoe UI", 10),
            bg=APP_BACKGROUND,
            fg=APP_MUTED_TEXT,
        )
        self.subtitle_label.pack(pady=(0, 5))

        frame = tk.LabelFrame(
            self.master,
            text="Attendance Configuration",
            font=APP_SECTION_FONT,
            padx=10,
            pady=6,
        )

        frame.pack(
            fill="x",
            padx=18,
            pady=(0, 8),
        )

        # CONFIGURATION FILE

        tk.Label(
            frame,
            text="Attendance Configuration (.xlsx)",
            font=DEFAULT_FONT,
        ).grid(
            row=0,
            column=0,
            columnspan=2,
            sticky="w",
            pady=(0, 4),
        )

        self.config_entry = tk.Entry(
            frame,
            width=34,
        )

        self.config_entry.grid(
            row=1,
            column=0,
            sticky="we",
            ipady=1,
            pady=(0, 10),
        )

        tk.Button(
            frame,
            text="Browse",
            font=BUTTON_FONT,
            command=self.browse_configuration,
        ).grid(
            row=1,
            column=1,
            sticky="e",
            padx=(8, 18),
            pady=(0, 10),
        )

        # OUTPUT FOLDER

        tk.Label(
            frame,
            text="Output Folder",
            font=DEFAULT_FONT,
        ).grid(
            row=0,
            column=2,
            columnspan=2,
            sticky="w",
            pady=(0, 4),
        )

        self.output_entry = tk.Entry(
            frame,
            width=34,
        )

        self.output_entry.grid(
            row=1,
            column=2,
            sticky="we",
            ipady=1,
            pady=(0, 10),
        )

        self.output_browse_button = tk.Button(
            frame,
            text="Browse",
            font=BUTTON_FONT,
            command=self.browse_output,
        )

        self.output_browse_button.grid(
            row=1,
            column=3,
            sticky="e",
            padx=(8, 0),
            pady=(0, 10),
        )

        tk.Checkbutton(
            frame,
            text="Same as Configuration File",
            variable=self.use_config_output_var,
            font=DEFAULT_FONT,
            command=self.toggle_output_source,
        ).grid(
            row=2,
            column=0,
            columnspan=4,
            sticky="w",
            pady=(0, 10),
        )

        # DATE RANGE

        tk.Label(
            frame,
            text="Date Range",
            font=DEFAULT_FONT,
        ).grid(
            row=3,
            column=0,
            columnspan=4,
            sticky="w",
            pady=(0, 4),
        )

        date_frame = tk.Frame(
            frame,
        )

        date_frame.grid(
            row=4,
            column=0,
            columnspan=4,
            sticky="w",
            pady=(0, 2),
        )

        tk.Label(
            date_frame,
            text="From",
            font=DEFAULT_FONT,
        ).pack(side="left")

        self.date_from_entry = tk.Entry(
            date_frame,
            width=12,
        )

        self.date_from_entry.pack(
            side="left",
            padx=(5, 3),
            ipady=1,
        )

        self.date_from_button = tk.Button(
            date_frame,
            text="Cal",
            width=3,
            command=lambda: self.open_date_picker("from"),
        )

        self.date_from_button.pack(
            side="left",
            padx=(0, 15),
        )

        tk.Label(
            date_frame,
            text="To",
            font=DEFAULT_FONT,
        ).pack(side="left")

        self.date_to_entry = tk.Entry(
            date_frame,
            width=12,
        )

        self.date_to_entry.pack(
            side="left",
            padx=(5, 3),
            ipady=1,
        )

        self.date_to_button = tk.Button(
            date_frame,
            text="Cal",
            width=3,
            command=lambda: self.open_date_picker("to"),
        )

        self.date_to_button.pack(
            side="left",
            padx=(0, 15),
        )

        tk.Label(
            date_frame,
            text="Format: MM/DD/YYYY",
            font=DEFAULT_FONT,
        ).pack(side="left")

        tk.Checkbutton(
            frame,
            text="Same as Configuration Date (General B8:B9)",
            variable=self.use_config_date_var,
            font=DEFAULT_FONT,
            command=self.toggle_date_source,
        ).grid(
            row=5,
            column=0,
            columnspan=4,
            sticky="w",
            pady=(8, 0),
        )

        frame.columnconfigure(0, weight=1)
        frame.columnconfigure(2, weight=1)
        self.toggle_output_source()
        self.toggle_date_source()

        # =================================================
        # OPTIONS WRAPPER
        # =================================================

        option_wrapper = tk.Frame(
            self.master,
            bg=APP_BACKGROUND,
        )

        option_wrapper.pack(
            fill="x",
            padx=18,
            pady=(8, 0),
        )

        # WORKFLOW FRAME

        workflow_frame = tk.LabelFrame(
            option_wrapper,
            text="Workflow",
            font=APP_SECTION_FONT,
            padx=12,
            pady=6,
        )

        workflow_frame.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(0, 8),
        )

        workflow_selector = tk.Frame(workflow_frame)
        workflow_selector.grid(
            row=0,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=5,
        )
        workflow_selector.columnconfigure(0, weight=1, uniform="workflow")
        workflow_selector.columnconfigure(1, weight=1, uniform="workflow")

        self.workflow_buttons: dict[str, tk.Button] = {}

        for column, (value, label) in enumerate(
            (("HO", "Head Office (HO)"), ("Branch", "Branch"))
        ):
            button = tk.Button(
                workflow_selector,
                text=label,
                font=BUTTON_FONT,
                command=lambda selected=value: self._select_workflow(selected),
                padx=8,
                pady=2,
                takefocus=True,
            )
            button.grid(
                row=0,
                column=column,
                sticky="ew",
                padx=(0, 4) if column == 0 else (4, 0),
            )
            self.workflow_buttons[value] = button

        self.workflow_status_label = tk.Label(
            workflow_frame,
            text="Selected Workflow : Head Office (HO)",
            font=DEFAULT_FONT,
        )

        self.workflow_status_label.grid(
            row=1,
            column=0,
            columnspan=2,
            sticky="w",
            padx=5,
            pady=(6, 0),
        )

        # OUTPUT OPTIONS FRAME

        output_option_frame = tk.LabelFrame(
            option_wrapper,
            text="Output Options",
            font=APP_SECTION_FONT,
            padx=12,
            pady=6,
        )

        output_option_frame.pack(
            side="left",
            fill="both",
            expand=True,
            padx=(8, 0),
        )

        tk.Checkbutton(
            output_option_frame,
            text="Generate HRIS TXT",
            variable=self.generate_txt_var,
            font=DEFAULT_FONT,
        ).grid(
            row=0,
            column=0,
            sticky="w",
            padx=5,
        )

        tk.Checkbutton(
            output_option_frame,
            text="Generate Excel Report",
            variable=self.generate_report_var,
            font=DEFAULT_FONT,
        ).grid(
            row=1,
            column=0,
            sticky="w",
            padx=5,
            pady=(5, 0),
        )

        # =================================================
        # ACTION + PROGRESS
        # =================================================

        action_frame = tk.Frame(
            self.master,
            bg=APP_BACKGROUND,
        )

        action_frame.pack(
            fill="x",
            padx=18,
            pady=(8, 0),
        )

        self.generate_button = tk.Button(
            action_frame,
            text="Generate",
            width=24,
            font=("Segoe UI", 13, "bold"),
            command=self.generate,
        )

        self.generate_button.pack(
            side="left",
            padx=(0, 18),
            ipady=5,
        )

        progress_frame = tk.LabelFrame(
            action_frame,
            text="Progress",
            font=APP_SECTION_FONT,
            padx=10,
            pady=6,
        )

        progress_frame.pack(
            side="left",
            fill="x",
            expand=True,
        )

        self.progress = ttk.Progressbar(
            progress_frame,
            length=500,
            mode="determinate",
            style="Attendance.Horizontal.TProgressbar",
        )

        self.progress.pack(
            fill="x",
        )

        # =================================================
        # PROCESS LOG
        # =================================================

        self.log_frame = tk.LabelFrame(
            self.master,
            text="Process Log",
            font=APP_SECTION_FONT,
            padx=10,
            pady=8,
        )

        self.log_frame.pack(
            fill="both",
            expand=True,
            padx=18,
            pady=(8, 6),
        )

        self.log_text = ScrolledText(
            self.log_frame,
            height=7,
            wrap="word",
            font=("Consolas", 10),
        )

        self.log_text.pack(
            fill="both",
            expand=True,
        )

        self.append_log("Application Ready.")
        self.append_log("Waiting for user input...")

        # STATUS BAR

        self.status_label = tk.Label(
            self.master,
            text="Status : Ready",
            anchor="w",
            relief="sunken",
        )

        self.status_label.pack(
            fill="x",
            side="bottom",
        )

    # =====================================================
    # EVENT
    # =====================================================

    def browse_configuration(self) -> None:
        """Browse Attendance Configuration workbook."""

        filename = filedialog.askopenfilename(
            parent=self.master,
            title="Select Attendance Configuration",
            filetypes=[
                ("Excel Workbook", "*.xlsx"),
            ],
        )
        self._bring_to_front()

        if filename:
            self.config_entry.delete(0, tk.END)
            self.config_entry.insert(0, filename)

            self.append_log("Attendance Configuration selected.")
            self.update_status("Configuration selected")

            if self.use_config_output_var.get():
                self.load_output_folder_from_configuration(
                    show_warning=True
                )

            if self.use_config_date_var.get():
                self.load_date_range_from_configuration(
                    show_warning=True
                )

    def browse_output(self) -> None:
        """Browse output folder."""

        folder = filedialog.askdirectory(
            parent=self.master,
            title="Select Output Folder"
        )
        self._bring_to_front()

        if folder:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, folder)

            self.append_log("Output folder selected.")
            self.update_status("Output folder selected")

    def toggle_output_source(self) -> None:
        """Toggle output folder source between config and manual."""

        if self.use_config_output_var.get():
            self.output_browse_button.config(state="disabled")
            self.output_entry.config(state="normal")
            self.load_output_folder_from_configuration(
                show_warning=False
            )
            self.output_entry.config(state="disabled")
            if hasattr(self, "log_text"):
                self.append_log(
                    "Output folder source: Attendance Configuration."
                )
        else:
            self.output_entry.config(state="normal")
            self.output_browse_button.config(state="normal")
            if hasattr(self, "log_text"):
                self.append_log("Output folder source: Manual selection.")

    def toggle_date_source(self) -> None:
        """Toggle date range source between config and manual."""

        if self.use_config_date_var.get():
            self.date_from_entry.config(state="normal")
            self.date_to_entry.config(state="normal")
            self.load_date_range_from_configuration(
                show_warning=False
            )
            self.date_from_entry.config(state="disabled")
            self.date_to_entry.config(state="disabled")
            self.date_from_button.config(state="disabled")
            self.date_to_button.config(state="disabled")

            if hasattr(self, "log_text"):
                self.append_log(
                    "Date range source: Attendance Configuration."
                )
        else:
            self.date_from_entry.config(state="normal")
            self.date_to_entry.config(state="normal")
            self.date_from_button.config(state="normal")
            self.date_to_button.config(state="normal")

            if hasattr(self, "log_text"):
                self.append_log("Date range source: Manual selection.")

    def load_date_range_from_configuration(
        self,
        show_warning: bool,
    ) -> bool:
        """Load Date From and Date To from General!B8 and General!B9."""

        config_file = self.config_entry.get().strip()

        if not config_file:
            return False

        try:
            workbook = load_workbook(
                filename=config_file,
                data_only=True,
                read_only=True,
            )

            if "General" not in workbook.sheetnames:
                raise ValueError("Sheet General was not found.")

            sheet = workbook["General"]
            date_from = self._format_config_date(sheet["B8"].value)
            date_to = self._format_config_date(sheet["B9"].value)
            workbook.close()

            if not date_from or not date_to:
                raise ValueError(
                    "General!B8 and General!B9 must contain dates."
                )

        except Exception as exc:
            if "workbook" in locals():
                workbook.close()

            self.append_log(
                f"Failed to read date range from configuration: {exc}"
            )

            if show_warning:
                Dialog.warning(
                    "Date range could not be read from "
                    "Attendance Configuration.\n\n"
                    "Expected dates in General!B8 and General!B9.\n\n"
                    f"{exc}"
                )
                self._bring_to_front()

            return False

        self.date_from_entry.config(state="normal")
        self.date_to_entry.config(state="normal")
        self.date_from_entry.delete(0, tk.END)
        self.date_from_entry.insert(0, date_from)
        self.date_to_entry.delete(0, tk.END)
        self.date_to_entry.insert(0, date_to)

        if self.use_config_date_var.get():
            self.date_from_entry.config(state="disabled")
            self.date_to_entry.config(state="disabled")

        self.append_log(
            f"Date range loaded from configuration: {date_from} - {date_to}"
        )
        self.update_status("Date range loaded from configuration")

        return True

    def load_output_folder_from_configuration(
        self,
        show_warning: bool,
    ) -> bool:
        """Load output folder from selected Attendance Configuration."""

        config_file = self.config_entry.get().strip()

        if not config_file:
            return False

        try:
            configuration_reader = AttendanceConfigurationReader(
                config_file
            )
            configuration = configuration_reader.read()
            output_folder = configuration.get_output_folder()
        except Exception as exc:
            self.append_log(
                f"Failed to read output folder from configuration: {exc}"
            )

            if show_warning:
                Dialog.warning(
                    "Output Folder could not be read from "
                    "Attendance Configuration.\n\n"
                    f"{exc}"
                )
                self._bring_to_front()

            return False

        if output_folder is None:
            self.append_log(
                "OutputFolder is not set in Attendance Configuration."
            )

            if show_warning:
                Dialog.warning(
                    "OutputFolder is not set in Attendance Configuration."
                )
                self._bring_to_front()

            return False

        self.output_entry.config(state="normal")
        self.output_entry.delete(0, tk.END)
        self.output_entry.insert(0, str(output_folder))

        if self.use_config_output_var.get():
            self.output_entry.config(state="disabled")

        self.append_log(
            f"Output folder loaded from configuration: {output_folder}"
        )
        self.update_status("Output folder loaded from configuration")

        return True

    def update_workflow_status(self) -> None:
        """Update selected workflow information."""

        workflow = self.workflow_var.get()

        if workflow == "HO":
            text = "Selected Workflow : Head Office (HO)"
        else:
            text = "Selected Workflow : Branch"

        self.workflow_status_label.config(text=text)
        self._update_workflow_selector()

        self.append_log(text)

    def _select_workflow(self, workflow: str) -> None:
        """Select a workflow from the compact card selector."""

        self.workflow_var.set(workflow)
        self.update_workflow_status()

    def _update_workflow_selector(self) -> None:
        """Refresh workflow button colors and selected marker."""

        selected_workflow = self.workflow_var.get()
        labels = {
            "HO": "Head Office (HO)",
            "Branch": "Branch",
        }

        for workflow, button in self.workflow_buttons.items():
            selected = workflow == selected_workflow
            button.config(
                text=("\u2713 " if selected else "") + labels[workflow],
                bg=WORKFLOW_ACCENT if selected else APP_INPUT,
                fg="#FFFFFF" if selected else APP_TEXT,
                activebackground=(
                    WORKFLOW_ACCENT_HOVER if selected else APP_SOFT_ACCENT
                ),
                activeforeground="#FFFFFF" if selected else APP_TEXT,
                relief="solid",
                bd=1,
                highlightthickness=1,
                highlightbackground=(
                    WORKFLOW_ACCENT if selected else APP_BORDER
                ),
                highlightcolor=WORKFLOW_ACCENT,
                cursor="hand2",
            )

    def generate(self) -> None:
        """Validate input and run Attendance Process Engine."""

        self.append_log("Generate button clicked.")
        self.update_status("Validating input")

        if not self.validate_input():
            return

        process_input = self.collect_input()

        self.append_log("Input validation success.")
        self.log_process_input(process_input)

        self.run_attendance_process(process_input)

    # =====================================================
    # DATE PICKER
    # =====================================================

    def open_date_picker(self, target: str) -> None:
        """Open simple calendar date picker."""

        initial_date = self._get_initial_picker_date(target)

        picker = tk.Toplevel(self.master)
        picker.title("Select Date")
        picker.geometry("320x300")
        picker.resizable(False, False)
        picker.transient(self.master)
        picker.grab_set()

        selected_year = tk.IntVar(value=initial_date.year)
        selected_month = tk.IntVar(value=initial_date.month)

        header_frame = tk.Frame(picker)
        header_frame.pack(fill="x", pady=8)

        calendar_frame = tk.Frame(picker)
        calendar_frame.pack(pady=5)

        def previous_month() -> None:
            month = selected_month.get()
            year = selected_year.get()

            if month == 1:
                selected_month.set(12)
                selected_year.set(year - 1)
            else:
                selected_month.set(month - 1)

            refresh_calendar()

        def next_month() -> None:
            month = selected_month.get()
            year = selected_year.get()

            if month == 12:
                selected_month.set(1)
                selected_year.set(year + 1)
            else:
                selected_month.set(month + 1)

            refresh_calendar()

        def select_date(day_number: int) -> None:
            selected_date = date(
                selected_year.get(),
                selected_month.get(),
                day_number,
            )

            formatted_date = selected_date.strftime(DATE_FORMAT)

            if target == "from":
                self.date_from_entry.delete(0, tk.END)
                self.date_from_entry.insert(0, formatted_date)
                self.append_log(f"Date From selected: {formatted_date}")
            else:
                self.date_to_entry.delete(0, tk.END)
                self.date_to_entry.insert(0, formatted_date)
                self.append_log(f"Date To selected: {formatted_date}")

            picker.destroy()

        def refresh_calendar() -> None:
            for widget in calendar_frame.winfo_children():
                widget.destroy()

            month_name = calendar.month_name[selected_month.get()]

            title_label.config(
                text=f"{month_name} {selected_year.get()}"
            )

            day_names = [
                "Mon",
                "Tue",
                "Wed",
                "Thu",
                "Fri",
                "Sat",
                "Sun",
            ]

            for column_index, day_name in enumerate(day_names):
                tk.Label(
                    calendar_frame,
                    text=day_name,
                    width=4,
                    font=DEFAULT_FONT,
                ).grid(
                    row=0,
                    column=column_index,
                    padx=2,
                    pady=2,
                )

            month_calendar = calendar.monthcalendar(
                selected_year.get(),
                selected_month.get(),
            )

            for row_index, week in enumerate(
                month_calendar,
                start=1,
            ):
                for column_index, day_number in enumerate(week):
                    if day_number == 0:
                        tk.Label(
                            calendar_frame,
                            text="",
                            width=4,
                        ).grid(
                            row=row_index,
                            column=column_index,
                            padx=2,
                            pady=2,
                        )
                    else:
                        tk.Button(
                            calendar_frame,
                            text=str(day_number),
                            width=4,
                            command=lambda day=day_number: select_date(
                                day
                            ),
                        ).grid(
                            row=row_index,
                            column=column_index,
                            padx=2,
                            pady=2,
                        )

            self._apply_widget_theme(picker)

        tk.Button(
            header_frame,
            text="<",
            width=4,
            command=previous_month,
        ).pack(
            side="left",
            padx=10,
        )

        title_label = tk.Label(
            header_frame,
            text="",
            font=HEADER_FONT,
        )

        title_label.pack(
            side="left",
            expand=True,
        )

        tk.Button(
            header_frame,
            text=">",
            width=4,
            command=next_month,
        ).pack(
            side="right",
            padx=10,
        )

        refresh_calendar()

    def _get_initial_picker_date(self, target: str) -> date:
        """Return initial date for date picker."""

        if target == "from":
            value = self.date_from_entry.get().strip()
        else:
            value = self.date_to_entry.get().strip()

        try:
            parsed_date = datetime.strptime(
                value,
                DATE_FORMAT,
            )
            return parsed_date.date()
        except ValueError:
            return date.today()

    # =====================================================
    # ENGINE INTEGRATION
    # =====================================================

    def collect_input(self) -> dict[str, Any]:
        """
        Collect GUI input for Attendance Engine.

        This method is the integration boundary between
        GUI and future Attendance Engine.
        """

        return {
            "configuration_file": Path(
                self.config_entry.get().strip()
            ),
            "output_folder": Path(
                self.output_entry.get().strip()
            ),
            "date_from": self.date_from_entry.get().strip(),
            "date_to": self.date_to_entry.get().strip(),
            "workflow": self.workflow_var.get(),
            "generate_txt": self.generate_txt_var.get(),
            "generate_report": self.generate_report_var.get(),
        }

    def log_process_input(
        self,
        process_input: dict[str, Any],
    ) -> None:
        """Write process input summary to process log."""

        self.append_log(
            f"Configuration file : "
            f"{process_input['configuration_file']}"
        )

        self.append_log(
            f"Output folder      : "
            f"{process_input['output_folder']}"
        )

        self.append_log(
            f"Date From          : "
            f"{process_input['date_from']}"
        )

        self.append_log(
            f"Date To            : "
            f"{process_input['date_to']}"
        )

        self.append_log(
            f"Workflow           : "
            f"{process_input['workflow']}"
        )

        self.append_log(
            f"Generate HRIS TXT  : "
            f"{process_input['generate_txt']}"
        )

        self.append_log(
            f"Generate Report    : "
            f"{process_input['generate_report']}"
        )

    def run_attendance_process(
        self,
        process_input: dict[str, Any],
    ) -> None:
        """Run real Attendance Process Engine."""

        self.generate_button.config(state="disabled")
        self.progress["value"] = 0

        try:
            self.update_status("Reading configuration")
            self.append_log("Reading Attendance Configuration...")
            self.master.update_idletasks()

            configuration_reader = AttendanceConfigurationReader(
                process_input["configuration_file"]
            )

            configuration = configuration_reader.read()

            self.progress["value"] = 15
            self.append_log("Attendance Configuration loaded.")
            self.master.update_idletasks()

            date_from = datetime.strptime(
                process_input["date_from"],
                DATE_FORMAT,
            )

            date_to = datetime.strptime(
                process_input["date_to"],
                DATE_FORMAT,
            )

            self.progress["value"] = 30
            self.update_status("Running Attendance Engine")
            self.append_log("Running Attendance Process Engine...")
            self.master.update_idletasks()

            process_engine = AttendanceProcessEngine()

            result = process_engine.run(
                configuration=configuration,
                output_root=process_input["output_folder"],
                workflow=process_input["workflow"],
                date_from=date_from,
                date_to=date_to,
                generate_txt=process_input["generate_txt"],
                generate_report=process_input["generate_report"],
            )

            self.progress["value"] = 100
            self.update_status("Process finished")

            self._log_process_result(result)

            Dialog.info(
                "Attendance process completed successfully.\n\n"
                f"Job ID: {result['job_id']}\n"
                f"Workflow: {result['workflow']}\n"
                f"Valid Records: {result['valid_record_count']}\n"
                f"Anomaly Records: {result['anomaly_record_count']}"
            )
            self._bring_to_front()

        except Exception as exc:
            self.progress["value"] = 0
            self.update_status("Process failed")
            self.append_log(f"Process failed: {exc}")

            logger.exception(
                "Attendance process failed."
            )

            Dialog.error(
                "Attendance process failed.\n\n"
                f"{exc}"
            )
            self._bring_to_front()

        finally:
            self.generate_button.config(state="normal")

    def _log_process_result(
        self,
        result: dict[str, Any],
    ) -> None:
        """Write engine result summary to process log."""

        self.append_log("Attendance process finished.")
        self.append_log(f"Job ID              : {result['job_id']}")
        self.append_log(f"Workflow            : {result['workflow']}")
        self.append_log(f"Raw log count       : {result['raw_log_count']}")
        self.append_log(
            f"Paired record count : {result['paired_record_count']}"
        )
        self.append_log(
            f"Valid record count  : {result['valid_record_count']}"
        )
        self.append_log(
            f"Duplicate removed   : "
            f"{result.get('duplicate_removed_count', 0)}"
        )
        self.append_log(
            f"Anomaly count       : {result['anomaly_record_count']}"
        )

        self.append_log("MDB Summary:")

        for item in result["mdb_summary"]:
            self.append_log(
                f"- {item['code']} | "
                f"{item['status']} | "
                f"Raw Logs: {item['raw_log_count']} | "
                f"{item['mdb_path']}"
            )

            if item.get("error"):
                self.append_log(
                    f"  Error: {item['error']}"
                )

        txt_result = result.get("txt_result")

        if txt_result:
            self.append_log(
                f"TXT output folder   : {txt_result['output_folder']}"
            )
            self.append_log(
                f"TXT files generated : {txt_result['total_files']}"
            )

            for file_info in txt_result["generated_files"]:
                self.append_log(
                    f"- TXT: {file_info['file_path']} "
                    f"({file_info['record_count']} rows)"
                )
        else:
            self.append_log("TXT generation skipped.")

        report_result = result.get("report_result")

        if report_result:
            self.append_log(
                f"Report file         : {report_result['report_file']}"
            )
        else:
            self.append_log("Report generation skipped.")

        artifact_result = result.get("artifact_result")

        if artifact_result:
            self.append_log(
                f"Process log         : {artifact_result['process_log']}"
            )
            self.append_log(
                f"Summary JSON        : {artifact_result['summary_json']}"
            )

    # =====================================================
    # VALIDATION
    # =====================================================

    def validate_input(self) -> bool:
        """Validate Attendance GUI input."""

        config_file = self.config_entry.get().strip()
        output_folder = self.output_entry.get().strip()
        date_from_text = self.date_from_entry.get().strip()
        date_to_text = self.date_to_entry.get().strip()
        workflow = self.workflow_var.get()

        result = validate_required(config_file)

        if not result.valid:
            return self._validation_failed(
                "Attendance Configuration is required."
            )

        config_path = Path(config_file)

        result = validate_file(config_path)

        if not result.valid:
            return self._validation_failed(
                f"Attendance Configuration is invalid.\n\n"
                f"{result.message}"
            )

        result = validate_extension(
            config_path,
            (".xlsx",),
        )

        if not result.valid:
            return self._validation_failed(
                "Attendance Configuration must be an Excel file (.xlsx)."
            )

        if self.use_config_output_var.get():
            self.load_output_folder_from_configuration(
                show_warning=False
            )
            output_folder = self.output_entry.get().strip()

        if self.use_config_date_var.get():
            if not self.load_date_range_from_configuration(
                show_warning=False
            ):
                return self._validation_failed(
                    "Date range could not be read from "
                    "Attendance Configuration.\n\n"
                    "Expected dates in General!B8 and General!B9."
                )

            date_from_text = self.date_from_entry.get().strip()
            date_to_text = self.date_to_entry.get().strip()

        result = validate_required(output_folder)

        if not result.valid:
            return self._validation_failed(
                "Output Folder is required.\n\n"
                "Set OutputFolder in Attendance Configuration or "
                "turn off Same as Configuration File and select it manually."
            )

        output_path = Path(output_folder)

        result = validate_folder(output_path)

        if not result.valid:
            return self._validation_failed(
                f"Output Folder is invalid.\n\n{result.message}"
            )

        result = validate_required(date_from_text)

        if not result.valid:
            return self._validation_failed(
                "Date From is required."
            )

        result = validate_required(date_to_text)

        if not result.valid:
            return self._validation_failed(
                "Date To is required."
            )

        date_from = self._parse_date(date_from_text)

        if date_from is None:
            return self._validation_failed(
                "Date From format is invalid.\n\n"
                "Please use MM/DD/YYYY format.\n"
                "Example: 07/01/2026"
            )

        date_to = self._parse_date(date_to_text)

        if date_to is None:
            return self._validation_failed(
                "Date To format is invalid.\n\n"
                "Please use MM/DD/YYYY format.\n"
                "Example: 07/31/2026"
            )

        if date_from > date_to:
            return self._validation_failed(
                "Date From cannot be greater than Date To."
            )

        if workflow not in ("HO", "Branch"):
            return self._validation_failed(
                "Workflow must be Head Office (HO) or Branch."
            )

        if (
            not self.generate_txt_var.get()
            and not self.generate_report_var.get()
        ):
            return self._validation_failed(
                "Please select at least one output option:\n\n"
                "- Generate HRIS TXT\n"
                "- Generate Excel Report"
            )

        return True

    def _parse_date(self, value: str) -> datetime | None:
        """Parse date using application date format."""

        try:
            return datetime.strptime(
                value,
                DATE_FORMAT,
            )
        except ValueError:
            return None

    def _format_config_date(self, value: Any) -> str:
        """Format Excel date value using application date format."""

        if value is None:
            return ""

        if isinstance(value, datetime):
            return value.strftime(DATE_FORMAT)

        if isinstance(value, date):
            return value.strftime(DATE_FORMAT)

        value_text = str(value).strip()

        if not value_text:
            return ""

        date_formats = (
            DATE_FORMAT,
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M:%S",
            "%d/%m/%Y",
            "%d-%m-%Y",
        )

        for date_format in date_formats:
            try:
                return datetime.strptime(
                    value_text,
                    date_format,
                ).strftime(DATE_FORMAT)
            except ValueError:
                continue

        return value_text

    def _validation_failed(self, message: str) -> bool:
        """Handle validation failure."""

        self.append_log(f"Validation failed: {message}")
        self.update_status("Validation failed")

        Dialog.warning(message)
        self._bring_to_front()

        return False

    # =====================================================
    # HELPER
    # =====================================================

    def _configure_ttk_style(self) -> None:
        """Configure themed widget colors."""

        style = ttk.Style(self.master)

        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(
            "Attendance.Horizontal.TProgressbar",
            background=APP_SUCCESS,
            troughcolor="#E2E9F2",
            bordercolor=APP_BORDER,
            lightcolor=APP_SUCCESS,
            darkcolor=APP_SUCCESS_ACTIVE,
            thickness=22,
        )

    def _apply_widget_theme(self, widget: tk.Misc) -> None:
        """Apply launcher color theme to existing widgets."""

        widget_class = widget.winfo_class()

        if widget_class in {"Toplevel", "Tk"}:
            self._safe_configure(
                widget,
                bg=APP_BACKGROUND,
            )
        elif widget_class == "Frame":
            parent_class = widget.master.winfo_class()
            frame_bg = (
                APP_PANEL
                if parent_class == "Labelframe"
                else APP_BACKGROUND
            )

            self._safe_configure(
                widget,
                bg=frame_bg,
            )
        elif widget is getattr(self, "log_frame", None):
            self._safe_configure(
                widget,
                bg=APP_LOG_BG,
                fg=APP_LOG_FG,
                highlightbackground=APP_LOG_BG,
                highlightcolor=APP_LOG_BG,
            )
        elif widget_class == "Labelframe":
            self._safe_configure(
                widget,
                bg=APP_PANEL,
                fg=APP_ACCENT_HOVER,
                relief="solid",
                bd=1,
                highlightthickness=1,
                highlightbackground=APP_BORDER,
                highlightcolor=APP_ACCENT,
            )
        elif widget_class == "Label":
            parent_class = widget.master.winfo_class()
            parent_bg = None

            try:
                parent_bg = widget.master.cget("bg")
            except tk.TclError:
                parent_bg = None

            if parent_class in {"Labelframe", "Frame"} and parent_bg:
                label_bg = parent_bg
            else:
                label_bg = APP_BACKGROUND

            self._safe_configure(
                widget,
                bg=label_bg,
                fg=APP_TEXT,
            )
        elif widget_class == "Button":
            self._safe_configure(
                widget,
                bg=APP_ACCENT,
                fg="#FFFFFF",
                activebackground=APP_ACCENT_HOVER,
                activeforeground="#FFFFFF",
                relief="flat",
                bd=0,
                cursor="hand2",
                highlightthickness=1,
                highlightbackground=APP_BORDER,
            )
        elif widget_class in {"Radiobutton", "Checkbutton"}:
            parent_bg = APP_PANEL

            try:
                parent_bg = widget.master.cget("bg")
            except tk.TclError:
                pass

            self._safe_configure(
                widget,
                bg=parent_bg,
                fg=APP_TEXT,
                activebackground=parent_bg,
                activeforeground=APP_MUTED_TEXT,
                selectcolor=APP_SOFT_ACCENT,
            )
        elif widget_class == "Entry":
            self._safe_configure(
                widget,
                bg=APP_INPUT,
                fg=APP_TEXT,
                insertbackground=APP_TEXT,
                relief="solid",
                bd=1,
                highlightthickness=1,
                highlightbackground=APP_BORDER,
                highlightcolor=APP_ACCENT,
                disabledbackground="#E8EEF5",
                disabledforeground=APP_MUTED_TEXT,
            )
        elif widget_class == "Text":
            text_bg = APP_INPUT
            text_fg = APP_TEXT
            border_color = APP_BORDER

            if widget.master is getattr(self, "log_frame", None):
                text_bg = APP_LOG_BG
                text_fg = APP_LOG_FG
                border_color = APP_LOG_BG

            self._safe_configure(
                widget,
                bg=text_bg,
                fg=text_fg,
                insertbackground=text_fg,
                relief="flat",
                bd=0,
                highlightthickness=1,
                highlightbackground=border_color,
            )

        for child in widget.winfo_children():
            self._apply_widget_theme(child)

        if widget is self.status_label:
            self._safe_configure(
                widget,
                bg=APP_ACCENT_HOVER,
                fg="#FFFFFF",
                relief="flat",
                padx=12,
                pady=4,
                font=("Segoe UI", 9, "bold"),
            )
        elif widget is getattr(self, "header_label", None):
            self._safe_configure(
                widget,
                bg=APP_BACKGROUND,
                fg=APP_ACCENT_HOVER,
            )
        elif widget is getattr(self, "subtitle_label", None):
            self._safe_configure(
                widget,
                bg=APP_BACKGROUND,
                fg=APP_MUTED_TEXT,
            )

    def _style_primary_action(self) -> None:
        """Apply generate action color to the Generate button."""

        self._safe_configure(
            self.generate_button,
            bg=APP_SUCCESS,
            fg="#FFFFFF",
            activebackground=APP_SUCCESS_ACTIVE,
            activeforeground="#FFFFFF",
            relief="flat",
            bd=0,
            overrelief="flat",
            highlightthickness=1,
            highlightbackground=APP_SUCCESS_BORDER,
            highlightcolor=APP_SUCCESS_BORDER,
            cursor="hand2",
        )

    @staticmethod
    def _safe_configure(widget: tk.Misc, **options: Any) -> None:
        """Configure supported widget options only."""

        try:
            widget.configure(**options)
        except tk.TclError:
            pass

    def _bring_to_front(self) -> None:
        """Bring Attendance window back after native dialogs."""

        try:
            self.master.lift()
            self.master.focus_force()
        except tk.TclError:
            pass

    def append_log(self, message: str) -> None:
        """Append message to process log."""

        self.log_text.insert(
            tk.END,
            message + "\n",
        )

        self.log_text.see(tk.END)

        logger.info(message)

    def update_status(self, message: str) -> None:
        """Update status bar."""

        self.status_label.config(
            text=f"Status : {message}"
        )
