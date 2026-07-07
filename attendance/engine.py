"""
=========================================================
Office Automation Suite - Karina (OAS-K)
---------------------------------------------------------
File        : engine.py
Module      : Attendance
Version     : 1.0.0
Author      : OpenAI & Zulkarnain Shiddiq
Python      : 3.14+
=========================================================
Attendance Engine

Sprint 5.4:
Attendance Pairing Engine.

Sprint 5.5:
Attendance Validation Engine.

This module converts raw attendance logs into paired
attendance records and validates which records are allowed
to continue to HRIS TXT generation.

Business rule:
For each NIK + tap date + source MDB:
- Earliest CHECKTIME = Jam Masuk
- Latest CHECKTIME   = Jam Pulang

TXT rule:
Only complete and valid attendance records are allowed
to be written to HRIS TXT.

=========================================================
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from config.app_config import DATE_FORMAT
from config.app_config import TIME_FORMAT
from collections import defaultdict
from datetime import date
from datetime import datetime
from typing import Any

from shared.logger import get_logger

logger = get_logger(__name__)


PAIR_STATUS_PAIRED: str = "PAIRED"
PAIR_STATUS_SINGLE_TAP: str = "SINGLE_TAP"
PAIR_STATUS_MISSING_NIK: str = "MISSING_NIK"
PAIR_STATUS_INVALID_CHECKTIME: str = "INVALID_CHECKTIME"

VALIDATION_STATUS_VALID: str = "VALID_FOR_TXT"
VALIDATION_STATUS_VALID_WARNING: str = "VALID_WITH_WARNING"
VALIDATION_STATUS_INVALID: str = "INVALID"


class AttendancePairingEngine:
    """
    Attendance pairing engine.

    This class receives raw attendance logs from
    AttendanceMDBExtractor and groups them into paired
    attendance records.
    """

    def pair_raw_logs(
        self,
        raw_logs: list[dict[str, Any]],
    ) -> list[dict[str, Any]]:
        """
        Pair raw attendance logs.
        """
        logger.info(
            "Starting attendance pairing for %s raw log(s).",
            len(raw_logs),
        )

        grouped_logs: dict[
            tuple[str, date, str],
            list[dict[str, Any]],
        ] = defaultdict(list)

        paired_records: list[dict[str, Any]] = []

        for raw_log in raw_logs:
            nik = str(raw_log.get("nik") or "").strip()
            checktime = raw_log.get("checktime")
            source_mdb_path = str(
                raw_log.get("source_mdb_path") or ""
            )

            if not nik:
                paired_records.append(
                    self._create_invalid_record(
                        raw_log=raw_log,
                        pair_status=PAIR_STATUS_MISSING_NIK,
                        remarks="NIK is empty.",
                    )
                )
                continue

            if not isinstance(checktime, datetime):
                paired_records.append(
                    self._create_invalid_record(
                        raw_log=raw_log,
                        pair_status=PAIR_STATUS_INVALID_CHECKTIME,
                        remarks="CHECKTIME is empty or invalid.",
                    )
                )
                continue

            group_key = (
                nik,
                checktime.date(),
                source_mdb_path,
            )

            grouped_logs[group_key].append(raw_log)

        for logs in grouped_logs.values():
            paired_records.append(
                self._pair_group(logs)
            )

        paired_records.sort(
            key=lambda record: (
                str(record.get("attendance_date") or ""),
                str(record.get("nik") or ""),
                str(record.get("source_mdb") or ""),
            )
        )

        summary = self.summarize_pairing(paired_records)

        logger.info(
            "Attendance pairing finished. "
            "Total pair record(s): %s, Paired: %s, Single tap: %s",
            summary["total_records"],
            summary["paired_records"],
            summary["single_tap_records"],
        )

        return paired_records

    def summarize_pairing(
        self,
        paired_records: list[dict[str, Any]],
    ) -> dict[str, int]:
        """
        Summarize pairing result.
        """
        summary = {
            "total_records": len(paired_records),
            "paired_records": 0,
            "single_tap_records": 0,
            "missing_nik_records": 0,
            "invalid_checktime_records": 0,
        }

        for record in paired_records:
            pair_status = record.get("pair_status")

            if pair_status == PAIR_STATUS_PAIRED:
                summary["paired_records"] += 1
            elif pair_status == PAIR_STATUS_SINGLE_TAP:
                summary["single_tap_records"] += 1
            elif pair_status == PAIR_STATUS_MISSING_NIK:
                summary["missing_nik_records"] += 1
            elif pair_status == PAIR_STATUS_INVALID_CHECKTIME:
                summary["invalid_checktime_records"] += 1

        return summary

    def _pair_group(
        self,
        logs: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """
        Pair one group of raw logs.

        One group represents:
        NIK + tap date + source MDB.
        """
        sorted_logs = sorted(
            logs,
            key=lambda item: item["checktime"],
        )

        first_log = sorted_logs[0]
        last_log = sorted_logs[-1]

        nik = str(first_log.get("nik") or "").strip()
        name = str(first_log.get("name") or "").strip()
        attendance_date = first_log["checktime"].date()

        tap_times = [
            raw_log["checktime"]
            for raw_log in sorted_logs
        ]

        tap_count = len(sorted_logs)

        if tap_count == 1:
            return {
                "nik": nik,
                "name": name,
                "attendance_date": attendance_date,
                "check_in": first_log["checktime"],
                "check_out": None,
                "tap_count": tap_count,
                "tap_times": tap_times,
                "source_mdb": first_log.get("source_mdb"),
                "source_mdb_path": first_log.get("source_mdb_path"),
                "pair_status": PAIR_STATUS_SINGLE_TAP,
                "remarks": "Only one tap record found.",
            }

        return {
            "nik": nik,
            "name": name,
            "attendance_date": attendance_date,
            "check_in": first_log["checktime"],
            "check_out": last_log["checktime"],
            "tap_count": tap_count,
            "tap_times": tap_times,
            "source_mdb": first_log.get("source_mdb"),
            "source_mdb_path": first_log.get("source_mdb_path"),
            "pair_status": PAIR_STATUS_PAIRED,
            "remarks": "",
        }

    @staticmethod
    def _create_invalid_record(
        raw_log: dict[str, Any],
        pair_status: str,
        remarks: str,
    ) -> dict[str, Any]:
        """
        Create invalid pairing record from problematic raw log.
        """
        checktime = raw_log.get("checktime")

        attendance_date = None

        if isinstance(checktime, datetime):
            attendance_date = checktime.date()

        return {
            "nik": str(raw_log.get("nik") or "").strip(),
            "name": str(raw_log.get("name") or "").strip(),
            "attendance_date": attendance_date,
            "check_in": checktime,
            "check_out": None,
            "tap_count": 1,
            "tap_times": [checktime] if checktime else [],
            "source_mdb": raw_log.get("source_mdb"),
            "source_mdb_path": raw_log.get("source_mdb_path"),
            "pair_status": pair_status,
            "remarks": remarks,
        }


class AttendanceValidationEngine:
    """
    Attendance validation engine.

    This class validates paired attendance records before
    they are allowed to continue to TXT/report generation.
    """

    def validate_paired_records(
        self,
        paired_records: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """
        Validate paired attendance records.

        Returns
        -------
        dict[str, Any]
            {
                "valid_records": [...],
                "anomaly_records": [...],
                "all_records": [...],
                "summary": {...},
            }
        """
        logger.info(
            "Starting attendance validation for %s paired record(s).",
            len(paired_records),
        )

        valid_records: list[dict[str, Any]] = []
        anomaly_records: list[dict[str, Any]] = []
        all_records: list[dict[str, Any]] = []

        for record in paired_records:
            validated_record = self._validate_record(record)

            all_records.append(validated_record)

            if validated_record["is_valid_for_txt"]:
                valid_records.append(validated_record)
            else:
                anomaly_records.append(validated_record)

        summary = self.summarize_validation(all_records)

        logger.info(
            "Attendance validation finished. "
            "Total: %s, Valid for TXT: %s, Anomaly: %s",
            summary["total_records"],
            summary["valid_for_txt_records"],
            summary["anomaly_records"],
        )

        return {
            "valid_records": valid_records,
            "anomaly_records": anomaly_records,
            "all_records": all_records,
            "summary": summary,
        }

    def summarize_validation(
        self,
        records: list[dict[str, Any]],
    ) -> dict[str, int]:
        """
        Summarize validation result.
        """
        summary = {
            "total_records": len(records),
            "valid_for_txt_records": 0,
            "valid_with_warning_records": 0,
            "anomaly_records": 0,
            "single_tap_records": 0,
            "missing_nik_records": 0,
            "invalid_checktime_records": 0,
            "missing_check_in_records": 0,
            "missing_check_out_records": 0,
        }

        for record in records:
            validation_status = record.get("validation_status")
            pair_status = record.get("pair_status")
            is_valid_for_txt = record.get("is_valid_for_txt")

            if is_valid_for_txt:
                summary["valid_for_txt_records"] += 1

            if validation_status == VALIDATION_STATUS_VALID_WARNING:
                summary["valid_with_warning_records"] += 1

            if validation_status == VALIDATION_STATUS_INVALID:
                summary["anomaly_records"] += 1

            if pair_status == PAIR_STATUS_SINGLE_TAP:
                summary["single_tap_records"] += 1
            elif pair_status == PAIR_STATUS_MISSING_NIK:
                summary["missing_nik_records"] += 1
            elif pair_status == PAIR_STATUS_INVALID_CHECKTIME:
                summary["invalid_checktime_records"] += 1

            if not isinstance(record.get("check_in"), datetime):
                summary["missing_check_in_records"] += 1

            if not isinstance(record.get("check_out"), datetime):
                summary["missing_check_out_records"] += 1

        return summary

    def _validate_record(
        self,
        record: dict[str, Any],
    ) -> dict[str, Any]:
        """
        Validate one paired attendance record.
        """
        validated_record = record.copy()

        nik = str(validated_record.get("nik") or "").strip()
        check_in = validated_record.get("check_in")
        check_out = validated_record.get("check_out")
        pair_status = validated_record.get("pair_status")
        tap_count = int(validated_record.get("tap_count") or 0)

        validation_messages: list[str] = []

        if not nik:
            validation_messages.append("NIK is empty.")

        if pair_status == PAIR_STATUS_MISSING_NIK:
            validation_messages.append("Pair status is MISSING_NIK.")

        if pair_status == PAIR_STATUS_INVALID_CHECKTIME:
            validation_messages.append(
                "Pair status is INVALID_CHECKTIME."
            )

        if pair_status == PAIR_STATUS_SINGLE_TAP:
            validation_messages.append(
                "Only one tap record found."
            )

        if not isinstance(check_in, datetime):
            validation_messages.append("Check in is missing.")

        if not isinstance(check_out, datetime):
            validation_messages.append("Check out is missing.")

        if validation_messages:
            validated_record["is_valid_for_txt"] = False
            validated_record[
                "validation_status"
            ] = VALIDATION_STATUS_INVALID
            validated_record[
                "validation_remarks"
            ] = " | ".join(validation_messages)

            return validated_record

        if (
            pair_status == PAIR_STATUS_PAIRED
            and isinstance(check_in, datetime)
            and isinstance(check_out, datetime)
        ):
            if tap_count > 2:
                validated_record["is_valid_for_txt"] = True
                validated_record[
                    "validation_status"
                ] = VALIDATION_STATUS_VALID_WARNING
                validated_record[
                    "validation_remarks"
                ] = (
                    "More than two tap records found. "
                    "Earliest tap is used as check in and latest "
                    "tap is used as check out."
                )

                return validated_record

            validated_record["is_valid_for_txt"] = True
            validated_record[
                "validation_status"
            ] = VALIDATION_STATUS_VALID
            validated_record["validation_remarks"] = ""

            return validated_record

        validated_record["is_valid_for_txt"] = False
        validated_record[
            "validation_status"
        ] = VALIDATION_STATUS_INVALID
        validated_record[
            "validation_remarks"
        ] = "Record does not meet valid TXT criteria."

        return validated_record
class AttendanceHRISTXTWriter:
    """
    HRIS TXT writer.

    This class writes valid attendance records into HRIS TXT
    format.

    Only records that passed validation are allowed here.
    """

    DEFAULT_MAX_ROWS_PER_FILE: int = 10000

    def write_txt_files(
        self,
        valid_records: list[dict[str, Any]],
        output_root: str | Path,
        workflow: str,
        max_rows_per_file: int = DEFAULT_MAX_ROWS_PER_FILE,
        job_id: str | None = None,
    ) -> dict[str, Any]:
        """
        Write valid attendance records into HRIS TXT files.
        """
        if max_rows_per_file <= 0:
            raise ValueError(
                "max_rows_per_file must be greater than zero."
            )

        workflow_label = self._normalize_workflow(workflow)

        if job_id is None:
            job_id = self._create_job_id()

        output_folder = (
            Path(output_root)
            / workflow_label
            / "TXT"
            / job_id
        )

        output_folder.mkdir(
            parents=True,
            exist_ok=True,
        )

        sorted_records = sorted(
            valid_records,
            key=lambda record: (
                str(record.get("attendance_date") or ""),
                str(record.get("nik") or ""),
                str(record.get("source_mdb") or ""),
            ),
        )

        generated_files: list[dict[str, Any]] = []

        if not sorted_records:
            logger.warning(
                "No valid attendance record found. TXT file not generated."
            )

            return {
                "job_id": job_id,
                "workflow": workflow_label,
                "output_folder": str(output_folder),
                "total_records": 0,
                "total_files": 0,
                "generated_files": generated_files,
            }

        chunks = self._chunk_records(
            sorted_records,
            max_rows_per_file,
        )

        for file_index, records_chunk in enumerate(
            chunks,
            start=1,
        ):
            file_name = self._create_txt_filename(
                workflow_label,
                file_index,
            )

            file_path = output_folder / file_name

            self._write_single_file(
                file_path=file_path,
                records=records_chunk,
            )

            generated_files.append(
                {
                    "file_name": file_name,
                    "file_path": str(file_path),
                    "record_count": len(records_chunk),
                }
            )

            logger.info(
                "Generated TXT file: %s (%s record(s))",
                file_path,
                len(records_chunk),
            )

        return {
            "job_id": job_id,
            "workflow": workflow_label,
            "output_folder": str(output_folder),
            "total_records": len(sorted_records),
            "total_files": len(generated_files),
            "generated_files": generated_files,
        }

    def _write_single_file(
        self,
        file_path: Path,
        records: list[dict[str, Any]],
    ) -> None:
        """
        Write one TXT file.
        """
        with file_path.open(
            mode="w",
            encoding="utf-8",
        ) as file:
            for record in records:
                file.write(
                    self._format_hris_line(record) + "\n"
                )

    def _format_hris_line(
        self,
        record: dict[str, Any],
    ) -> str:
        """
        Format one valid record into HRIS TXT line.

        Format:
        "MM/DD/YYYY","NIK","MM/DD/YYYY","HH:MM","MM/DD/YYYY","HH:MM"
        """
        nik = str(record.get("nik") or "").strip()
        check_in = record.get("check_in")
        check_out = record.get("check_out")

        if not nik:
            raise ValueError("NIK is empty.")

        if not isinstance(check_in, datetime):
            raise ValueError("check_in must be datetime.")

        if not isinstance(check_out, datetime):
            raise ValueError("check_out must be datetime.")

        hris_date = check_out.strftime(DATE_FORMAT)
        date_in = check_in.strftime(DATE_FORMAT)
        time_in = check_in.strftime(TIME_FORMAT)
        date_out = check_out.strftime(DATE_FORMAT)
        time_out = check_out.strftime(TIME_FORMAT)

        fields = [
            hris_date,
            nik,
            date_in,
            time_in,
            date_out,
            time_out,
        ]

        return ",".join(
            self._quote_field(field)
            for field in fields
        )

    @staticmethod
    def _quote_field(value: str) -> str:
        """
        Quote one TXT field.
        """
        safe_value = str(value).replace('"', '""')
        return f'"{safe_value}"'

    @staticmethod
    def _chunk_records(
        records: list[dict[str, Any]],
        chunk_size: int,
    ) -> list[list[dict[str, Any]]]:
        """
        Split records into chunks.
        """
        return [
            records[index:index + chunk_size]
            for index in range(0, len(records), chunk_size)
        ]

    @staticmethod
    def _create_job_id() -> str:
        """
        Create timestamp job ID.
        """
        return datetime.now().strftime("%Y%m%d_%H%M%S")

    @staticmethod
    def _normalize_workflow(workflow: str) -> str:
        """
        Normalize workflow label.
        """
        normalized = workflow.strip().lower()

        if normalized == "ho":
            return "HO"

        if normalized == "branch":
            return "Branch"

        raise ValueError(
            "workflow must be 'HO' or 'Branch'."
        )

    @staticmethod
    def _create_txt_filename(
        workflow_label: str,
        file_index: int,
    ) -> str:
        """
        Create TXT filename.

        Example:
        Attendance_HO_001.txt
        Attendance_Branch_001.txt
        """
        return f"Attendance_{workflow_label}_{file_index:03d}.txt"

class AttendanceExcelReportWriter:
    """
    Attendance Excel report writer.

    Report sheets:
    1. Summary
    2. Attendance_Detail
    3. Summary_Per_Karyawan
    4. Summary_Per_Hari
    5. Anomaly
    """

    def write_report(
        self,
        all_records: list[dict[str, Any]],
        valid_records: list[dict[str, Any]],
        anomaly_records: list[dict[str, Any]],
        validation_summary: dict[str, int],
        output_root: str | Path,
        workflow: str,
        job_id: str,
        report_name: str | None = None,
    ) -> dict[str, Any]:
        """
        Write Attendance Excel report.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
        from openpyxl.styles import Font
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter

        workflow_label = self._normalize_workflow(workflow)
        output_folder = (
            Path(output_root)
            / workflow_label
            / "Report"
            / job_id
        )

        output_folder.mkdir(
            parents=True,
            exist_ok=True,
        )

        if report_name is None:
            report_name = f"Report_{job_id}.xlsx"

        report_path = output_folder / report_name

        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        detail_sheet = workbook.create_sheet("Attendance_Detail")
        employee_sheet = workbook.create_sheet("Summary_Per_Karyawan")
        daily_sheet = workbook.create_sheet("Summary_Per_Hari")
        anomaly_sheet = workbook.create_sheet("Anomaly")

        self._write_summary_sheet(
            sheet=summary_sheet,
            workflow_label=workflow_label,
            job_id=job_id,
            validation_summary=validation_summary,
            valid_records=valid_records,
            anomaly_records=anomaly_records,
        )

        self._write_detail_sheet(
            sheet=detail_sheet,
            records=valid_records,
        )

        self._write_employee_summary_sheet(
            sheet=employee_sheet,
            records=all_records,
        )

        self._write_daily_summary_sheet(
            sheet=daily_sheet,
            records=all_records,
        )

        self._write_anomaly_sheet(
            sheet=anomaly_sheet,
            records=anomaly_records,
        )

        for sheet in workbook.worksheets:
            self._style_sheet(
                sheet=sheet,
                font_class=Font,
                fill_class=PatternFill,
                alignment_class=Alignment,
                get_column_letter=get_column_letter,
            )

        workbook.save(report_path)
        workbook.close()

        logger.info(
            "Generated Attendance Excel report: %s",
            report_path,
        )

        return {
            "job_id": job_id,
            "workflow": workflow_label,
            "report_folder": str(output_folder),
            "report_file": str(report_path),
            "sheet_count": len(workbook.worksheets),
            "valid_records": len(valid_records),
            "anomaly_records": len(anomaly_records),
        }

    def _write_summary_sheet(
        self,
        sheet: Any,
        workflow_label: str,
        job_id: str,
        validation_summary: dict[str, int],
        valid_records: list[dict[str, Any]],
        anomaly_records: list[dict[str, Any]],
    ) -> None:
        """
        Write Summary sheet.
        """
        rows = [
            ["Item", "Value"],
            ["Application", "Office Automation Suite - Karina"],
            ["Module", "Attendance"],
            ["Workflow", workflow_label],
            ["Job ID", job_id],
            ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Total Valid TXT Records", len(valid_records)],
            ["Total Anomaly Records", len(anomaly_records)],
            ["Total Validated Records", validation_summary.get("total_records", 0)],
            ["Valid For TXT", validation_summary.get("valid_for_txt_records", 0)],
            ["Valid With Warning", validation_summary.get("valid_with_warning_records", 0)],
            ["Single Tap Records", validation_summary.get("single_tap_records", 0)],
            ["Missing NIK Records", validation_summary.get("missing_nik_records", 0)],
            ["Invalid CHECKTIME Records", validation_summary.get("invalid_checktime_records", 0)],
            ["Missing Check Out Records", validation_summary.get("missing_check_out_records", 0)],
        ]

        for row in rows:
            sheet.append(row)

    def _write_detail_sheet(
        self,
        sheet: Any,
        records: list[dict[str, Any]],
    ) -> None:
        """
        Write valid attendance detail sheet.
        """
        sheet.append(
            [
                "NIK",
                "Nama",
                "Tanggal",
                "Jam Masuk",
                "Jam Keluar",
                "Tap Count",
                "Validation Status",
                "Validation Remarks",
                "Source MDB",
            ]
        )

        for record in records:
            sheet.append(
                [
                    record.get("nik", ""),
                    record.get("name", ""),
                    self._format_date(record.get("attendance_date")),
                    self._format_time(record.get("check_in")),
                    self._format_time(record.get("check_out")),
                    record.get("tap_count", 0),
                    record.get("validation_status", ""),
                    record.get("validation_remarks", ""),
                    record.get("source_mdb", ""),
                ]
            )

    def _write_employee_summary_sheet(
        self,
        sheet: Any,
        records: list[dict[str, Any]],
    ) -> None:
        """
        Write summary per employee.
        """
        sheet.append(
            [
                "NIK",
                "Nama",
                "Total Hari",
                "Valid For TXT",
                "Anomaly",
                "Valid With Warning",
            ]
        )

        summary: dict[str, dict[str, Any]] = {}

        for record in records:
            nik = str(record.get("nik") or "").strip()
            name = str(record.get("name") or "").strip()

            if not nik:
                nik = "(EMPTY)"

            if nik not in summary:
                summary[nik] = {
                    "name": name,
                    "total_days": 0,
                    "valid": 0,
                    "anomaly": 0,
                    "warning": 0,
                }

            summary[nik]["total_days"] += 1

            if record.get("is_valid_for_txt"):
                summary[nik]["valid"] += 1
            else:
                summary[nik]["anomaly"] += 1

            if record.get("validation_status") == VALIDATION_STATUS_VALID_WARNING:
                summary[nik]["warning"] += 1

        for nik, item in sorted(summary.items()):
            sheet.append(
                [
                    nik,
                    item["name"],
                    item["total_days"],
                    item["valid"],
                    item["anomaly"],
                    item["warning"],
                ]
            )

    def _write_daily_summary_sheet(
        self,
        sheet: Any,
        records: list[dict[str, Any]],
    ) -> None:
        """
        Write summary per day.
        """
        sheet.append(
            [
                "Tanggal",
                "Total Record",
                "Valid For TXT",
                "Anomaly",
                "Valid With Warning",
            ]
        )

        summary: dict[str, dict[str, int]] = {}

        for record in records:
            attendance_date = self._format_date(
                record.get("attendance_date")
            )

            if not attendance_date:
                attendance_date = "(EMPTY)"

            if attendance_date not in summary:
                summary[attendance_date] = {
                    "total": 0,
                    "valid": 0,
                    "anomaly": 0,
                    "warning": 0,
                }

            summary[attendance_date]["total"] += 1

            if record.get("is_valid_for_txt"):
                summary[attendance_date]["valid"] += 1
            else:
                summary[attendance_date]["anomaly"] += 1

            if record.get("validation_status") == VALIDATION_STATUS_VALID_WARNING:
                summary[attendance_date]["warning"] += 1

        for attendance_date, item in sorted(summary.items()):
            sheet.append(
                [
                    attendance_date,
                    item["total"],
                    item["valid"],
                    item["anomaly"],
                    item["warning"],
                ]
            )

    def _write_anomaly_sheet(
        self,
        sheet: Any,
        records: list[dict[str, Any]],
    ) -> None:
        """
        Write anomaly sheet.
        """
        sheet.append(
            [
                "NIK",
                "Nama",
                "Tanggal",
                "Jam Masuk",
                "Jam Keluar",
                "Tap Count",
                "Pair Status",
                "Validation Status",
                "Validation Remarks",
                "Source MDB",
            ]
        )

        for record in records:
            sheet.append(
                [
                    record.get("nik", ""),
                    record.get("name", ""),
                    self._format_date(record.get("attendance_date")),
                    self._format_time(record.get("check_in")),
                    self._format_time(record.get("check_out")),
                    record.get("tap_count", 0),
                    record.get("pair_status", ""),
                    record.get("validation_status", ""),
                    record.get("validation_remarks", ""),
                    record.get("source_mdb", ""),
                ]
            )

    def _style_sheet(
        self,
        sheet: Any,
        font_class: Any,
        fill_class: Any,
        alignment_class: Any,
        get_column_letter: Any,
    ) -> None:
        """
        Apply basic Excel formatting.
        """
        header_fill = fill_class(
            fill_type="solid",
            fgColor="1F4E78",
        )

        header_font = font_class(
            bold=True,
            color="FFFFFF",
        )

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment_class(
                horizontal="center",
                vertical="center",
            )

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = alignment_class(
                    vertical="top",
                    wrap_text=True,
                )

        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                value = cell.value

                if value is None:
                    continue

                max_length = max(
                    max_length,
                    len(str(value)),
                )

            adjusted_width = min(
                max(max_length + 2, 12),
                60,
            )

            sheet.column_dimensions[column_letter].width = adjusted_width

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    @staticmethod
    def _format_date(value: Any) -> str:
        """
        Format date value.
        """
        if value is None:
            return ""

        if hasattr(value, "strftime"):
            return value.strftime(DATE_FORMAT)

        return str(value)

    @staticmethod
    def _format_time(value: Any) -> str:
        """
        Format time value.
        """
        if value is None:
            return ""

        if hasattr(value, "strftime"):
            return value.strftime(TIME_FORMAT)

        return str(value)

    @staticmethod
    def _normalize_workflow(workflow: str) -> str:
        """
        Normalize workflow label.
        """
        normalized = workflow.strip().lower()

        if normalized == "ho":
            return "HO"

        if normalized == "branch":
            return "Branch"

        raise ValueError("workflow must be 'HO' or 'Branch'.")

class AttendanceProcessEngine:
    """
    End-to-end Attendance process engine.

    Sprint 5.9:
    - Read selected MDB list from Attendance Configuration
    - Extract raw logs
    - Pair attendance
    - Validate records
    - Generate HRIS TXT
    - Generate Excel report
    """

    def run(
        self,
        configuration: Any,
        output_root: str | Path,
        workflow: str,
        date_from: datetime,
        date_to: datetime,
        generate_txt: bool = True,
        generate_report: bool = True,
    ) -> dict[str, Any]:
        """
        Run end-to-end Attendance process.
        """
        if not generate_txt and not generate_report:
            raise ValueError(
                "At least one output must be selected."
            )

        workflow_label = self._normalize_workflow(workflow)

        mdb_list = self._get_mdb_list(
            configuration=configuration,
            workflow_label=workflow_label,
        )

        if not mdb_list:
            raise ValueError(
                f"No active MDB found for workflow: {workflow_label}"
            )

        job_id = self._create_job_id()

        split_rows = self._get_split_rows(configuration)

        raw_logs: list[dict[str, Any]] = []
        mdb_summary: list[dict[str, Any]] = []

        from attendance.extractor import AttendanceMDBExtractor

        for mdb_item in mdb_list:
            mdb_path = Path(mdb_item.mdb_path)

            try:
                with AttendanceMDBExtractor(mdb_path) as extractor:
                    extracted_logs = extractor.fetch_raw_logs(
                        date_from=date_from,
                        date_to=date_to,
                    )

                for raw_log in extracted_logs:
                    raw_log["source_code"] = mdb_item.code
                    raw_log["source_description"] = (
                        mdb_item.description
                    )

                raw_logs.extend(extracted_logs)

                mdb_summary.append(
                    {
                        "code": mdb_item.code,
                        "description": mdb_item.description,
                        "mdb_path": str(mdb_path),
                        "status": "SUCCESS",
                        "raw_log_count": len(extracted_logs),
                        "error": "",
                    }
                )

            except Exception as exc:
                mdb_summary.append(
                    {
                        "code": mdb_item.code,
                        "description": mdb_item.description,
                        "mdb_path": str(mdb_path),
                        "status": "FAILED",
                        "raw_log_count": 0,
                        "error": str(exc),
                    }
                )

        successful_mdb_count = sum(
            1
            for item in mdb_summary
            if item.get("status") == "SUCCESS"
        )

        if successful_mdb_count == 0:
            failed_sources = ", ".join(
                str(item.get("code") or item.get("mdb_path"))
                for item in mdb_summary
            )
            raise RuntimeError(
                "All MDB extraction failed. "
                f"Failed source(s): {failed_sources}"
            )

        pairing_engine = AttendancePairingEngine()
        paired_records = pairing_engine.pair_raw_logs(raw_logs)

        validation_engine = AttendanceValidationEngine()
        validation_result = validation_engine.validate_paired_records(
            paired_records
        )

        valid_records = validation_result["valid_records"]
        anomaly_records = validation_result["anomaly_records"]
        all_records = validation_result["all_records"]
        validation_summary = validation_result["summary"]

        txt_result = None
        report_result = None

        if generate_txt:
            txt_writer = AttendanceHRISTXTWriter()

            txt_result = txt_writer.write_txt_files(
                valid_records=valid_records,
                output_root=output_root,
                workflow=workflow_label,
                max_rows_per_file=split_rows,
                job_id=job_id,
            )

        if generate_report:
            report_writer = AttendanceExcelReportWriter()

            report_result = report_writer.write_report(
                all_records=all_records,
                valid_records=valid_records,
                anomaly_records=anomaly_records,
                validation_summary=validation_summary,
                output_root=output_root,
                workflow=workflow_label,
                job_id=job_id,
            )

        result = {
            "job_id": job_id,
            "workflow": workflow_label,
            "date_from": date_from,
            "date_to": date_to,
            "mdb_summary": mdb_summary,
            "raw_log_count": len(raw_logs),
            "paired_record_count": len(paired_records),
            "valid_record_count": len(valid_records),
            "anomaly_record_count": len(anomaly_records),
            "validation_summary": validation_summary,
            "txt_result": txt_result,
            "report_result": report_result,
        }

        artifact_writer = AttendanceRunArtifactWriter()

        artifact_result = artifact_writer.write_artifacts(
            result=result,
            output_root=output_root,
            workflow=workflow_label,
            job_id=job_id,
        )

        result["artifact_result"] = artifact_result

        return result

    @staticmethod
    def _get_mdb_list(
        configuration: Any,
        workflow_label: str,
    ) -> list[Any]:
        """
        Return selected workflow MDB list.
        """
        if workflow_label == "HO":
            return configuration.ho_mdb_list

        if workflow_label == "Branch":
            return configuration.branch_mdb_list

        raise ValueError(
            "workflow must be 'HO' or 'Branch'."
        )

    @staticmethod
    def _get_split_rows(configuration: Any) -> int:
        """
        Get split row setting from Attendance Configuration.
        """
        value = configuration.general.get(
            "Split_TXT_Rows",
            10000,
        )

        try:
            split_rows = int(value)
        except (TypeError, ValueError):
            split_rows = 10000

        if split_rows <= 0:
            return 10000

        return split_rows

    @staticmethod
    def _create_job_id() -> str:
        """
        Create timestamp job ID.
        """
        return datetime.now().strftime("%Y%m%d_%H%M%S")

    @staticmethod
    def _normalize_workflow(workflow: str) -> str:
        """
        Normalize workflow label.
        """
        normalized = workflow.strip().lower()

        if normalized == "ho":
            return "HO"

        if normalized == "branch":
            return "Branch"

        raise ValueError(
            "workflow must be 'HO' or 'Branch'."
        )
    
class AttendanceRunArtifactWriter:
    """
    Attendance run artifact writer.

    Sprint 5.11:
    Write Process_<JOB_ID>.txt and summary_<JOB_ID>.json
    for each Attendance run.

    Output location:
    output_root / workflow / Report / job_id
    """

    def write_artifacts(
        self,
        result: dict[str, Any],
        output_root: str | Path,
        workflow: str,
        job_id: str,
    ) -> dict[str, Any]:
        """
        Write Process.log and summary.json.
        """
        workflow_label = self._normalize_workflow(workflow)

        artifact_folder = (
            Path(output_root)
            / workflow_label
            / "Report"
            / job_id
        )

        artifact_folder.mkdir(
            parents=True,
            exist_ok=True,
        )

        process_log_path = artifact_folder / f"Process_{job_id}.txt"
        summary_json_path = artifact_folder / f"summary_{job_id}.json"

        self._write_process_log(
            file_path=process_log_path,
            result=result,
        )

        self._write_summary_json(
            file_path=summary_json_path,
            result=result,
        )

        logger.info(
            "Generated run artifacts: %s and %s",
            process_log_path,
            summary_json_path,
        )

        return {
            "artifact_folder": str(artifact_folder),
            "process_log": str(process_log_path),
            "summary_json": str(summary_json_path),
        }

    def _write_process_log(
        self,
        file_path: Path,
        result: dict[str, Any],
    ) -> None:
        """
        Write human-readable process log.
        """
        lines: list[str] = []

        lines.append("=" * 80)
        lines.append("Office Automation Suite - Karina")
        lines.append("Attendance Process Log")
        lines.append("=" * 80)
        lines.append(f"Job ID              : {result.get('job_id', '')}")
        lines.append(f"Workflow            : {result.get('workflow', '')}")
        lines.append(
            f"Date From           : "
            f"{self._format_datetime(result.get('date_from'))}"
        )
        lines.append(
            f"Date To             : "
            f"{self._format_datetime(result.get('date_to'))}"
        )
        lines.append(f"Generated At        : {datetime.now():%Y-%m-%d %H:%M:%S}")
        lines.append("")
        lines.append("PROCESS SUMMARY")
        lines.append("-" * 80)
        lines.append(f"Raw log count       : {result.get('raw_log_count', 0)}")
        lines.append(
            f"Paired record count : "
            f"{result.get('paired_record_count', 0)}"
        )
        lines.append(
            f"Valid record count  : "
            f"{result.get('valid_record_count', 0)}"
        )
        lines.append(
            f"Anomaly count       : "
            f"{result.get('anomaly_record_count', 0)}"
        )
        lines.append("")
        lines.append("MDB SUMMARY")
        lines.append("-" * 80)

        for item in result.get("mdb_summary", []):
            lines.append(
                f"{item.get('code', '')} | "
                f"{item.get('description', '')} | "
                f"{item.get('status', '')} | "
                f"Raw Logs: {item.get('raw_log_count', 0)} | "
                f"{item.get('mdb_path', '')}"
            )

            if item.get("error"):
                lines.append(f"Error: {item.get('error')}")

        lines.append("")
        lines.append("VALIDATION SUMMARY")
        lines.append("-" * 80)

        validation_summary = result.get("validation_summary", {})

        for key, value in validation_summary.items():
            lines.append(f"{key}: {value}")

        txt_result = result.get("txt_result")

        lines.append("")
        lines.append("TXT RESULT")
        lines.append("-" * 80)

        if txt_result:
            lines.append(
                f"Output folder       : "
                f"{txt_result.get('output_folder', '')}"
            )
            lines.append(
                f"Total TXT files     : "
                f"{txt_result.get('total_files', 0)}"
            )
            lines.append(
                f"Total TXT records   : "
                f"{txt_result.get('total_records', 0)}"
            )

            for file_info in txt_result.get("generated_files", []):
                lines.append(
                    f"{file_info.get('file_path', '')} "
                    f"({file_info.get('record_count', 0)} rows)"
                )
        else:
            lines.append("TXT generation skipped.")

        report_result = result.get("report_result")

        lines.append("")
        lines.append("REPORT RESULT")
        lines.append("-" * 80)

        if report_result:
            lines.append(
                f"Report file         : "
                f"{report_result.get('report_file', '')}"
            )
        else:
            lines.append("Report generation skipped.")

        lines.append("")
        lines.append("=" * 80)
        lines.append("Process finished.")
        lines.append("=" * 80)

        file_path.write_text(
            "\n".join(lines),
            encoding="utf-8",
        )

    def _write_summary_json(
        self,
        file_path: Path,
        result: dict[str, Any],
    ) -> None:
        """
        Write machine-readable summary JSON.
        """
        json_data = self._make_json_safe(result)

        file_path.write_text(
            json.dumps(
                json_data,
                indent=4,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

    def _make_json_safe(
        self,
        value: Any,
    ) -> Any:
        """
        Convert Python objects into JSON-safe values.
        """
        if isinstance(value, dict):
            return {
                str(key): self._make_json_safe(item)
                for key, item in value.items()
            }

        if isinstance(value, list):
            return [
                self._make_json_safe(item)
                for item in value
            ]

        if isinstance(value, tuple):
            return [
                self._make_json_safe(item)
                for item in value
            ]

        if isinstance(value, Path):
            return str(value)

        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")

        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")

        return value

    @staticmethod
    def _format_datetime(value: Any) -> str:
        """
        Format datetime value for Process.log.
        """
        if isinstance(value, datetime):
            return value.strftime("%m/%d/%Y")

        if isinstance(value, date):
            return value.strftime("%m/%d/%Y")

        if value is None:
            return ""

        return str(value)

    @staticmethod
    def _normalize_workflow(workflow: str) -> str:
        """
        Normalize workflow label.
        """
        normalized = workflow.strip().lower()

        if normalized == "ho":
            return "HO"

        if normalized == "branch":
            return "Branch"

        raise ValueError(
            "workflow must be 'HO' or 'Branch'."
        )
