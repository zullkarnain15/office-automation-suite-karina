"""
=========================================================
Office Automation Suite - Karina (OAS-K)
---------------------------------------------------------
File        : config_manager.py
Module      : Shared
Version     : 1.0.0
Author      : OpenAI & Zulkarnain Shiddiq
Python      : 3.14+
=========================================================
Configuration Manager

This module contains configuration readers for OAS-K.

Current readers:
- AttendanceConfigurationReader
- HRISConfigurationReader

=========================================================
"""

from __future__ import annotations

from dataclasses import dataclass
from dataclasses import field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from shared.logger import get_logger

logger = get_logger(__name__)


# =========================================================
# ATTENDANCE CONFIGURATION
# =========================================================


@dataclass(slots=True)
class AttendanceMDBConfig:
    """
    Attendance MDB configuration item.
    """

    active: bool
    code: str
    description: str
    mdb_path: Path


@dataclass(slots=True)
class AttendanceConfiguration:
    """
    Attendance configuration result.
    """

    configuration_file: Path
    general: dict[str, Any]
    output: dict[str, Any]
    ho_mdb_list: list[AttendanceMDBConfig]
    branch_mdb_list: list[AttendanceMDBConfig]

    def get_output_folder(self) -> Path | None:
        """
        Return configured Attendance output folder, if available.
        """
        output_value = self._get_first_value(
            self.general,
            (
                "OutputFolder",
                "Output_Folder",
                "Output_Root",
            ),
        )

        if output_value is None:
            output_value = self._get_first_value(
                self.output,
                ("Output_Root",),
            )

        if output_value is None:
            return None

        output_path = Path(str(output_value).strip())

        if not str(output_path):
            return None

        if output_path.is_absolute():
            return output_path

        return self.configuration_file.parent.parent / output_path

    @staticmethod
    def _get_first_value(
        values: dict[str, Any],
        keys: tuple[str, ...],
    ) -> Any:
        """
        Return the first non-empty configuration value for keys.
        """
        for key in keys:
            value = values.get(key)

            if value is None:
                continue

            value_text = str(value).strip()

            if value_text:
                return value

        return None


class AttendanceConfigurationReader:
    """
    Reader for OAS-K Attendance Configuration workbook.
    """

    REQUIRED_SHEETS: tuple[str, ...] = (
        "General",
        "MDB_HO",
        "MDB_Branch",
        "Output",
    )

    def __init__(
        self,
        configuration_file: str | Path,
    ) -> None:
        self.configuration_file = Path(configuration_file)

    def read(self) -> AttendanceConfiguration:
        """
        Read Attendance Configuration workbook.
        """
        if not self.configuration_file.exists():
            raise FileNotFoundError(
                f"Configuration file not found: "
                f"{self.configuration_file}"
            )

        if self.configuration_file.suffix.lower() != ".xlsx":
            raise ValueError(
                "Attendance Configuration must be an .xlsx file."
            )

        logger.info(
            "Reading Attendance Configuration: %s",
            self.configuration_file,
        )

        workbook = load_workbook(
            filename=self.configuration_file,
            data_only=True,
            read_only=True,
        )

        self._validate_required_sheets(workbook.sheetnames)

        general = self._read_key_value_sheet(
            workbook["General"]
        )

        output = self._read_key_value_sheet(
            workbook["Output"]
        )

        ho_mdb_list = self._read_ho_mdb_sheet(
            workbook["MDB_HO"]
        )

        branch_mdb_list = self._read_branch_mdb_sheet(
            workbook["MDB_Branch"]
        )

        workbook.close()

        logger.info(
            "Attendance Configuration loaded. "
            "HO MDB: %s, Branch MDB: %s",
            len(ho_mdb_list),
            len(branch_mdb_list),
        )

        return AttendanceConfiguration(
            configuration_file=self.configuration_file,
            general=general,
            output=output,
            ho_mdb_list=ho_mdb_list,
            branch_mdb_list=branch_mdb_list,
        )

    def _validate_required_sheets(
        self,
        sheet_names: list[str],
    ) -> None:
        """
        Validate required sheets.
        """
        missing_sheets = [
            sheet_name
            for sheet_name in self.REQUIRED_SHEETS
            if sheet_name not in sheet_names
        ]

        if missing_sheets:
            raise ValueError(
                "Missing required sheet(s): "
                + ", ".join(missing_sheets)
            )

    def _read_key_value_sheet(
        self,
        sheet: Any,
    ) -> dict[str, Any]:
        """
        Read sheet with columns:
        Parameter | Value | Description
        """
        result: dict[str, Any] = {}

        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return result

        for row in rows[1:]:
            if row is None:
                continue

            parameter = row[0] if len(row) > 0 else None
            value = row[1] if len(row) > 1 else None

            if parameter is None:
                continue

            parameter_text = str(parameter).strip()

            if not parameter_text:
                continue

            result[parameter_text] = value

        return result

    def _read_ho_mdb_sheet(
        self,
        sheet: Any,
    ) -> list[AttendanceMDBConfig]:
        """
        Read MDB_HO sheet.

        Expected columns:
        Active | Company | Description | MDB_Path
        """
        mdb_list: list[AttendanceMDBConfig] = []

        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) <= 1:
            return mdb_list

        for row in rows[1:]:
            active = self._to_bool(row[0] if len(row) > 0 else None)

            if not active:
                continue

            company = self._to_text(row[1] if len(row) > 1 else "")
            description = self._to_text(row[2] if len(row) > 2 else "")
            mdb_path_text = self._to_text(row[3] if len(row) > 3 else "")

            if not mdb_path_text:
                continue

            mdb_list.append(
                AttendanceMDBConfig(
                    active=active,
                    code=company,
                    description=description,
                    mdb_path=Path(mdb_path_text),
                )
            )

        return mdb_list

    def _read_branch_mdb_sheet(
        self,
        sheet: Any,
    ) -> list[AttendanceMDBConfig]:
        """
        Read MDB_Branch sheet.

        Expected columns:
        Active | Branch_Code | Branch_Name | MDB_Path
        """
        mdb_list: list[AttendanceMDBConfig] = []

        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) <= 1:
            return mdb_list

        for row in rows[1:]:
            active = self._to_bool(row[0] if len(row) > 0 else None)

            if not active:
                continue

            branch_code = self._to_text(row[1] if len(row) > 1 else "")
            branch_name = self._to_text(row[2] if len(row) > 2 else "")
            mdb_path_text = self._to_text(row[3] if len(row) > 3 else "")

            if not mdb_path_text:
                continue

            mdb_list.append(
                AttendanceMDBConfig(
                    active=active,
                    code=branch_code,
                    description=branch_name,
                    mdb_path=Path(mdb_path_text),
                )
            )

        return mdb_list

    @staticmethod
    def _to_bool(value: Any) -> bool:
        """
        Convert Excel value to boolean.
        """
        if value is None:
            return False

        value_text = str(value).strip().lower()

        return value_text in (
            "y",
            "yes",
            "true",
            "1",
            "active",
        )

    @staticmethod
    def _to_text(value: Any) -> str:
        """
        Convert Excel value to clean string.
        """
        if value is None:
            return ""

        return str(value).strip()


# =========================================================
# OUTLOOK - REVISI CONFIGURATION
# =========================================================


@dataclass(slots=True)
class OutlookSenderConfig:
    """
    Outlook sender master item for HO or Branch workflow.
    """

    active: bool
    workflow: str
    sender_name: str
    sender_email: str
    required_cc_email: str
    company: str = ""
    branch_code: str = ""


@dataclass(slots=True)
class OutlookSubjectRule:
    """
    Outlook subject validation rule.
    """

    active: bool
    workflow: str
    subject_pattern: str


@dataclass(slots=True)
class OutlookAttachmentRule:
    """
    Outlook attachment extension rule.
    """

    active: bool
    workflow: str
    allowed_extensions: list[str]


@dataclass(slots=True)
class OutlookValidationRule:
    """
    Outlook data validation rule.
    """

    active: bool
    rule_code: str
    workflow: str
    rule_value: str


@dataclass(slots=True)
class OutlookReplyTemplate:
    """
    Outlook reply email template.
    """

    active: bool
    reply_code: str
    recipient_type: str
    trigger: str
    subject_template: str
    body_template: str


@dataclass(slots=True)
class OutlookRevisiConfiguration:
    """
    Outlook - Revisi configuration result.
    """

    configuration_file: Path
    general: dict[str, Any]
    ho_senders: list[OutlookSenderConfig]
    branch_senders: list[OutlookSenderConfig]
    subject_rules: list[OutlookSubjectRule]
    attachment_rules: list[OutlookAttachmentRule]
    validation_rules: list[OutlookValidationRule]
    reply_templates: list[OutlookReplyTemplate]

    def get_output_root(self) -> Path | None:
        """
        Return configured Outlook - Revisi output root, if available.
        """
        output_value = self._get_first_value(
            self.general,
            ("Output_Root", "OutputFolder", "Output_Folder"),
        )

        if output_value is None:
            return None

        output_path = Path(str(output_value).strip())

        if not str(output_path):
            return None

        if output_path.is_absolute():
            resolved_path = output_path
        else:
            resolved_path = self._get_application_root() / output_path

        if resolved_path.name.casefold() == "output":
            return resolved_path / "Outlook-Revisi"

        return resolved_path

    def get_pic_hr_emails(self) -> list[str]:
        """
        Return semicolon-separated PIC HR recipients.
        """
        return OutlookRevisiConfigurationReader.split_emails(
            self.general.get("PIC_HR_Emails")
        )

    def get_spv_pic_hr_emails(self) -> list[str]:
        """
        Return semicolon-separated SPV/PIC HR CC recipients.
        """
        return OutlookRevisiConfigurationReader.split_emails(
            self.general.get("SPV_PIC_HR_Emails")
        )

    def _get_application_root(self) -> Path:
        """
        Resolve project/EXE root for config/outlook relative paths.
        """
        parent = self.configuration_file.parent

        if parent.name.lower() == "outlook":
            config_parent = parent.parent
            if config_parent.name.lower() == "config":
                return config_parent.parent

        return self.configuration_file.parent

    @staticmethod
    def _get_first_value(
        values: dict[str, Any],
        keys: tuple[str, ...],
    ) -> Any:
        """
        Return the first non-empty configuration value for keys.
        """
        for key in keys:
            value = values.get(key)

            if value is None:
                continue

            value_text = str(value).strip()

            if value_text:
                return value

        return None


class OutlookRevisiConfigurationReader:
    """
    Reader for OAS-K Outlook - Revisi Configuration workbook.
    """

    REQUIRED_SHEETS: tuple[str, ...] = (
        "General",
        "HO_Sender_Master",
        "Branch_Sender_Master",
        "Subject_Rules",
        "Attachment_Rules",
        "Validation_Rules",
        "Reply_Templates",
    )

    def __init__(
        self,
        configuration_file: str | Path,
    ) -> None:
        self.configuration_file = Path(configuration_file)

    def read(self) -> OutlookRevisiConfiguration:
        """
        Read Outlook - Revisi Configuration workbook.
        """
        if not self.configuration_file.exists():
            raise FileNotFoundError(
                f"Outlook - Revisi Configuration file not found: "
                f"{self.configuration_file}"
            )

        if self.configuration_file.suffix.lower() != ".xlsx":
            raise ValueError(
                "Outlook - Revisi Configuration must be an .xlsx file."
            )

        logger.info(
            "Reading Outlook - Revisi Configuration: %s",
            self.configuration_file,
        )

        workbook = load_workbook(
            filename=self.configuration_file,
            data_only=True,
            read_only=True,
        )

        self._validate_required_sheets(workbook.sheetnames)

        general = self._read_key_value_sheet(workbook["General"])
        ho_senders = self._read_ho_sender_sheet(
            workbook["HO_Sender_Master"]
        )
        branch_senders = self._read_branch_sender_sheet(
            workbook["Branch_Sender_Master"]
        )
        subject_rules = self._read_subject_rules_sheet(
            workbook["Subject_Rules"]
        )
        attachment_rules = self._read_attachment_rules_sheet(
            workbook["Attachment_Rules"]
        )
        validation_rules = self._read_validation_rules_sheet(
            workbook["Validation_Rules"]
        )
        reply_templates = self._read_reply_templates_sheet(
            workbook["Reply_Templates"]
        )

        workbook.close()

        logger.info(
            "Outlook - Revisi Configuration loaded. "
            "HO Senders: %s, Branch Senders: %s",
            len(ho_senders),
            len(branch_senders),
        )

        return OutlookRevisiConfiguration(
            configuration_file=self.configuration_file,
            general=general,
            ho_senders=ho_senders,
            branch_senders=branch_senders,
            subject_rules=subject_rules,
            attachment_rules=attachment_rules,
            validation_rules=validation_rules,
            reply_templates=reply_templates,
        )

    def _validate_required_sheets(
        self,
        sheet_names: list[str],
    ) -> None:
        """
        Validate required sheets.
        """
        missing_sheets = [
            sheet_name
            for sheet_name in self.REQUIRED_SHEETS
            if sheet_name not in sheet_names
        ]

        if missing_sheets:
            raise ValueError(
                "Missing required Outlook - Revisi sheet(s): "
                + ", ".join(missing_sheets)
            )

    def _read_key_value_sheet(
        self,
        sheet: Any,
    ) -> dict[str, Any]:
        """
        Read a key-value sheet with Parameter | Value | Description columns.
        """
        rows = list(sheet.iter_rows(values_only=True))
        header_index = self._find_header_index(rows, ("Parameter", "Value"))
        result: dict[str, Any] = {}

        for row in rows[header_index + 1:]:
            parameter = self._cell(row, 0)

            if not parameter:
                continue

            result[parameter] = self._cell(row, 1)

        return result

    def _read_ho_sender_sheet(
        self,
        sheet: Any,
    ) -> list[OutlookSenderConfig]:
        """
        Read HO_Sender_Master sheet.
        """
        records = self._read_table_records(
            sheet,
            ("Active", "Sender_Name", "Sender_Email", "Required_CC_Email"),
        )
        senders: list[OutlookSenderConfig] = []

        for record in records:
            if not self._to_bool(record.get("Active")):
                continue

            sender_email = self._to_text(record.get("Sender_Email"))
            if not sender_email:
                continue

            senders.append(
                OutlookSenderConfig(
                    active=True,
                    workflow="HO",
                    sender_name=self._to_text(record.get("Sender_Name")),
                    sender_email=sender_email,
                    required_cc_email=self._to_text(
                        record.get("Required_CC_Email")
                    ),
                )
            )

        return senders

    def _read_branch_sender_sheet(
        self,
        sheet: Any,
    ) -> list[OutlookSenderConfig]:
        """
        Read Branch_Sender_Master sheet.
        """
        records = self._read_table_records(
            sheet,
            (
                "Active",
                "Company",
                "Branch_Code",
                "Sender_Name",
                "Sender_Email",
                "Required_CC_Email",
            ),
        )
        senders: list[OutlookSenderConfig] = []

        for record in records:
            if not self._to_bool(record.get("Active")):
                continue

            sender_email = self._to_text(record.get("Sender_Email"))
            if not sender_email:
                continue

            senders.append(
                OutlookSenderConfig(
                    active=True,
                    workflow="Branch",
                    company=self._to_text(record.get("Company")),
                    branch_code=self._to_text(record.get("Branch_Code")),
                    sender_name=self._to_text(record.get("Sender_Name")),
                    sender_email=sender_email,
                    required_cc_email=self._to_text(
                        record.get("Required_CC_Email")
                    ),
                )
            )

        return senders

    def _read_subject_rules_sheet(
        self,
        sheet: Any,
    ) -> list[OutlookSubjectRule]:
        """
        Read Subject_Rules sheet.
        """
        records = self._read_table_records(
            sheet,
            ("Active", "Workflow", "Subject_Pattern"),
        )
        rules: list[OutlookSubjectRule] = []

        for record in records:
            if not self._to_bool(record.get("Active")):
                continue

            workflow = self._normalize_workflow(record.get("Workflow"))
            subject_pattern = self._to_text(record.get("Subject_Pattern"))

            if not workflow or not subject_pattern:
                continue

            rules.append(
                OutlookSubjectRule(
                    active=True,
                    workflow=workflow,
                    subject_pattern=subject_pattern,
                )
            )

        return rules

    def _read_attachment_rules_sheet(
        self,
        sheet: Any,
    ) -> list[OutlookAttachmentRule]:
        """
        Read Attachment_Rules sheet.
        """
        records = self._read_table_records(
            sheet,
            ("Active", "Workflow", "Allowed_Extensions"),
        )
        rules: list[OutlookAttachmentRule] = []

        for record in records:
            if not self._to_bool(record.get("Active")):
                continue

            workflow = self._normalize_workflow(record.get("Workflow"))
            extensions = self.split_semicolon_values(
                record.get("Allowed_Extensions")
            )

            if not workflow or not extensions:
                continue

            rules.append(
                OutlookAttachmentRule(
                    active=True,
                    workflow=workflow,
                    allowed_extensions=[
                        extension.lower()
                        for extension in extensions
                    ],
                )
            )

        return rules

    def _read_validation_rules_sheet(
        self,
        sheet: Any,
    ) -> list[OutlookValidationRule]:
        """
        Read Validation_Rules sheet.
        """
        records = self._read_table_records(
            sheet,
            ("Active", "Rule_Code", "Workflow", "Rule_Value"),
        )
        rules: list[OutlookValidationRule] = []

        for record in records:
            if not self._to_bool(record.get("Active")):
                continue

            rule_code = self._to_text(record.get("Rule_Code")).upper()
            workflow = self._normalize_workflow(record.get("Workflow"))

            if not rule_code or not workflow:
                continue

            rules.append(
                OutlookValidationRule(
                    active=True,
                    rule_code=rule_code,
                    workflow=workflow,
                    rule_value=self._to_text(record.get("Rule_Value")),
                )
            )

        return rules

    def _read_reply_templates_sheet(
        self,
        sheet: Any,
    ) -> list[OutlookReplyTemplate]:
        """
        Read Reply_Templates sheet.
        """
        records = self._read_table_records(
            sheet,
            (
                "Active",
                "Reply_Code",
                "Recipient_Type",
                "Trigger",
                "Subject_Template",
                "Body_Template",
            ),
        )
        templates: list[OutlookReplyTemplate] = []

        for record in records:
            if not self._to_bool(record.get("Active")):
                continue

            reply_code = self._to_text(record.get("Reply_Code")).upper()
            trigger = self._to_text(record.get("Trigger")).upper()

            if not reply_code or not trigger:
                continue

            templates.append(
                OutlookReplyTemplate(
                    active=True,
                    reply_code=reply_code,
                    recipient_type=self._to_text(
                        record.get("Recipient_Type")
                    ).upper(),
                    trigger=trigger,
                    subject_template=self._to_text(
                        record.get("Subject_Template")
                    ),
                    body_template=self._to_text(
                        record.get("Body_Template")
                    ),
                )
            )

        return templates

    def _read_table_records(
        self,
        sheet: Any,
        required_headers: tuple[str, ...],
    ) -> list[dict[str, Any]]:
        """
        Read a table whose header row may appear after title/instruction rows.
        """
        rows = list(sheet.iter_rows(values_only=True))
        header_index = self._find_header_index(rows, required_headers)
        headers = [
            self._to_text(cell)
            for cell in rows[header_index]
        ]
        records: list[dict[str, Any]] = []

        for row in rows[header_index + 1:]:
            if not row or not any(self._to_text(cell) for cell in row):
                continue

            record: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                record[header] = self._cell(row, index)

            records.append(record)

        return records

    @classmethod
    def _find_header_index(
        cls,
        rows: list[tuple[Any, ...]],
        required_headers: tuple[str, ...],
    ) -> int:
        """
        Locate a header row by required column names.
        """
        required = {
            cls._normalize_header(header)
            for header in required_headers
        }

        for index, row in enumerate(rows):
            actual = {
                cls._normalize_header(cell)
                for cell in row
                if cls._to_text(cell)
            }

            if required.issubset(actual):
                return index

        raise ValueError(
            "Missing required column(s): "
            + ", ".join(required_headers)
        )

    @staticmethod
    def _cell(row: tuple[Any, ...], index: int) -> str:
        if index >= len(row):
            return ""

        return OutlookRevisiConfigurationReader._to_text(row[index])

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return OutlookRevisiConfigurationReader._to_text(value).lower()

    @staticmethod
    def _to_bool(value: Any) -> bool:
        """
        Convert Excel value to boolean.
        """
        if value is None:
            return False

        value_text = str(value).strip().lower()

        return value_text in (
            "y",
            "yes",
            "true",
            "1",
            "active",
        )

    @staticmethod
    def _to_text(value: Any) -> str:
        """
        Convert Excel value to clean string.
        """
        if value is None:
            return ""

        return str(value).strip()

    @staticmethod
    def _normalize_workflow(value: Any) -> str:
        """
        Normalize workflow value.
        """
        value_text = str(value or "").strip().lower()

        if value_text == "ho":
            return "HO"

        if value_text == "branch":
            return "Branch"

        if value_text == "all":
            return "All"

        return ""

    @staticmethod
    def split_semicolon_values(value: Any) -> list[str]:
        """
        Split semicolon-separated configuration values.
        """
        value_text = OutlookRevisiConfigurationReader._to_text(value)

        if not value_text:
            return []

        return [
            item.strip()
            for item in value_text.split(";")
            if item.strip()
        ]

    @staticmethod
    def split_emails(value: Any) -> list[str]:
        """
        Split semicolon-separated email addresses.
        """
        return OutlookRevisiConfigurationReader.split_semicolon_values(value)


# =========================================================
# HRIS CONFIGURATION
# =========================================================


@dataclass(slots=True)
class HRISRunControlConfig:
    """
    HRIS Run Control configuration item.
    """

    active: bool
    sequence: int
    workflow: str
    run_control_id: str
    description: str


@dataclass(slots=True)
class HRISAssistedStepConfig:
    """One configuration-driven assisted upload action."""

    active: bool
    sequence: int
    step_name: str
    action: str
    input_source: str
    method: str
    required: bool
    wait_after_seconds: float
    description: str


DEFAULT_HRIS_ASSISTED_STEPS: tuple[tuple[Any, ...], ...] = (
    ("Y", 1, "run_control_id", "click_type", "RUN_CONTROL_ID", "coordinate", True, 1, "Focus Run Control field and type current Run Control ID"),
    ("Y", 2, "start_date", "click_type", "START_DATE", "coordinate", True, 1, "Focus Start Date field and type Start Date"),
    ("Y", 3, "end_date", "click_type", "END_DATE", "coordinate", True, 1, "Focus End Date field and type End Date"),
    ("Y", 4, "add_attachment", "click", "NONE", "coordinate", True, 1, "Click Add Attachment"),
    ("Y", 5, "choose_file", "attach_file", "TXT_FILE_PATH", "assisted", True, 1, "Attach or choose TXT file for current item"),
    ("Y", 6, "upload", "click", "NONE", "coordinate", True, 2, "Click Upload"),
    ("Y", 7, "ok_after_upload", "click", "NONE", "coordinate", True, 1, "Click OK after Upload"),
    ("Y", 8, "run", "click", "NONE", "coordinate", True, 2, "Click Run"),
    ("Y", 9, "ok_after_run", "click", "NONE", "coordinate", True, 1, "Click OK after Run"),
)
HRIS_POST_UPLOAD_ASSISTED_STEP_NAMES = (
    "upload",
    "ok_after_upload",
    "run",
    "ok_after_run",
)


def resolve_hris_macro_steps(
    steps: list[HRISAssistedStepConfig],
) -> list[HRISAssistedStepConfig]:
    """Resolve canonical macro steps while supporting legacy four-step configs."""
    steps_by_name = {step.step_name: step for step in steps}
    canonical_steps = [
        steps_by_name[name]
        for name in HRIS_POST_UPLOAD_ASSISTED_STEP_NAMES
        if name in steps_by_name
    ]
    return canonical_steps or list(steps)


HRIS_ASSISTED_ACTIONS = {
    "click", "click_type", "type", "press",
    "attach_file", "wait", "manual_continue",
}
HRIS_ASSISTED_METHODS = {"coordinate", "playwright", "manual", "assisted"}
HRIS_ASSISTED_INPUT_SOURCES = {
    "NONE", "RUN_CONTROL_ID", "START_DATE", "END_DATE", "TXT_FILE_PATH",
}


@dataclass(slots=True)
class HRISConfiguration:
    """
    HRIS configuration result.
    """

    configuration_file: Path
    general: dict[str, Any]
    browser: dict[str, Any]
    upload: dict[str, Any]
    ho_run_controls: list[HRISRunControlConfig]
    branch_run_controls: list[HRISRunControlConfig]
    assisted_steps: list[HRISAssistedStepConfig] = field(default_factory=list)

    def get_output_folder(self) -> Path | None:
        """
        Return configured HRIS output folder, if available.
        """
        output_value = self._get_first_value(
            self.general,
            (
                "Folder_Upload_Path",
                "OutputFolder",
                "Output_Folder",
                "Output_Root",
            ),
        )

        if output_value is None:
            output_value = self._get_first_value(
                self.upload,
                (
                    "Folder_Upload_Path",
                    "OutputFolder",
                    "Output_Folder",
                    "Output_Root",
                ),
            )

        if output_value is None:
            return None

        output_path = Path(str(output_value).strip())

        if not str(output_path):
            return None

        if output_path.is_absolute():
            return output_path

        return self.configuration_file.parent.parent / output_path

    @staticmethod
    def _get_first_value(
        values: dict[str, Any],
        keys: tuple[str, ...],
    ) -> Any:
        """
        Return the first non-empty configuration value for keys.
        """
        for key in keys:
            value = values.get(key)

            if value is None:
                continue

            value_text = str(value).strip()

            if value_text:
                return value

        return None


class HRISConfigurationReader:
    """
    Reader for OAS-K HRIS Configuration workbook.
    """

    REQUIRED_SHEETS: tuple[str, ...] = (
        "General",
        "Run_Control",
        "Browser",
        "Upload",
        "Reference",
    )

    def __init__(
        self,
        configuration_file: str | Path,
    ) -> None:
        self.configuration_file = Path(configuration_file)

    def read(self) -> HRISConfiguration:
        """
        Read HRIS Configuration workbook.
        """
        if not self.configuration_file.exists():
            raise FileNotFoundError(
                f"HRIS Configuration file not found: "
                f"{self.configuration_file}"
            )

        if self.configuration_file.suffix.lower() != ".xlsx":
            raise ValueError(
                "HRIS Configuration must be an .xlsx file."
            )

        logger.info(
            "Reading HRIS Configuration: %s",
            self.configuration_file,
        )

        workbook = load_workbook(
            filename=self.configuration_file,
            data_only=True,
            read_only=True,
        )

        self._validate_required_sheets(workbook.sheetnames)

        general = self._read_key_value_sheet(
            workbook["General"]
        )

        browser = self._read_key_value_sheet(
            workbook["Browser"]
        )

        upload = self._read_key_value_sheet(
            workbook["Upload"]
        )

        run_controls = self._read_run_control_sheet(
            workbook["Run_Control"]
        )

        if "Assisted_Steps" in workbook.sheetnames:
            assisted_steps = self._read_assisted_steps_sheet(
                workbook["Assisted_Steps"]
            )
        else:
            logger.warning(
                "Assisted_Steps sheet is missing; using safe default steps."
            )
            assisted_steps = self._default_assisted_steps()

        workbook.close()

        ho_run_controls = [
            item
            for item in run_controls
            if item.workflow == "HO"
        ]

        branch_run_controls = [
            item
            for item in run_controls
            if item.workflow == "Branch"
        ]

        logger.info(
            "HRIS Configuration loaded. "
            "HO Run Control: %s, Branch Run Control: %s",
            len(ho_run_controls),
            len(branch_run_controls),
        )

        return HRISConfiguration(
            configuration_file=self.configuration_file,
            general=general,
            browser=browser,
            upload=upload,
            ho_run_controls=ho_run_controls,
            branch_run_controls=branch_run_controls,
            assisted_steps=assisted_steps,
        )

    def _validate_required_sheets(
        self,
        sheet_names: list[str],
    ) -> None:
        """
        Validate required sheets.
        """
        missing_sheets = [
            sheet_name
            for sheet_name in self.REQUIRED_SHEETS
            if sheet_name not in sheet_names
        ]

        if missing_sheets:
            raise ValueError(
                "Missing required HRIS sheet(s): "
                + ", ".join(missing_sheets)
            )

    def _read_key_value_sheet(
        self,
        sheet: Any,
    ) -> dict[str, Any]:
        """
        Read sheet with columns:
        Parameter | Value | Description
        """
        result: dict[str, Any] = {}

        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return result

        for row in rows[1:]:
            if row is None:
                continue

            parameter = row[0] if len(row) > 0 else None
            value = row[1] if len(row) > 1 else None

            if parameter is None:
                continue

            parameter_text = str(parameter).strip()

            if not parameter_text:
                continue

            result[parameter_text] = value

        return result

    def _read_run_control_sheet(
        self,
        sheet: Any,
    ) -> list[HRISRunControlConfig]:
        """
        Read Run_Control sheet.

        Expected columns:
        Active | Sequence | Workflow | Run_Control_ID | Description
        """
        run_controls: list[HRISRunControlConfig] = []

        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) <= 1:
            return run_controls

        for row in rows[1:]:
            if row is None:
                continue

            active = self._to_bool(row[0] if len(row) > 0 else None)

            if not active:
                continue

            sequence = self._to_int(row[1] if len(row) > 1 else None)
            workflow = self._normalize_workflow(
                row[2] if len(row) > 2 else None
            )
            run_control_id = self._to_text(
                row[3] if len(row) > 3 else ""
            )
            description = self._to_text(
                row[4] if len(row) > 4 else ""
            )

            if sequence <= 0:
                continue

            if workflow not in ("HO", "Branch"):
                continue

            if not run_control_id:
                continue

            run_controls.append(
                HRISRunControlConfig(
                    active=active,
                    sequence=sequence,
                    workflow=workflow,
                    run_control_id=run_control_id,
                    description=description,
                )
            )

        run_controls.sort(
            key=lambda item: (
                item.workflow,
                item.sequence,
            )
        )

        return run_controls

    def _read_assisted_steps_sheet(
        self,
        sheet: Any,
    ) -> list[HRISAssistedStepConfig]:
        """Read active Assisted_Steps rows and order them by Sequence."""
        steps: list[HRISAssistedStepConfig] = []
        rows = list(sheet.iter_rows(values_only=True))

        for row in rows[1:]:
            if not row or not self._to_bool(row[0] if len(row) > 0 else None):
                continue

            sequence = self._to_int(row[1] if len(row) > 1 else None)
            step_name = self._to_text(row[2] if len(row) > 2 else None)
            if sequence <= 0 or not step_name:
                continue

            step = HRISAssistedStepConfig(
                    active=True,
                    sequence=sequence,
                    step_name=step_name,
                    action=self._to_text(row[3] if len(row) > 3 else None).lower(),
                    input_source=self._to_text(row[4] if len(row) > 4 else "NONE").upper(),
                    method=self._to_text(row[5] if len(row) > 5 else "manual").lower(),
                    required=self._to_bool(row[6] if len(row) > 6 else None),
                    wait_after_seconds=self._to_float(row[7] if len(row) > 7 else None),
                    description=self._to_text(row[8] if len(row) > 8 else None),
                )
            self._validate_assisted_step(step)
            steps.append(step)

        steps.sort(key=lambda item: item.sequence)
        return steps

    def _default_assisted_steps(self) -> list[HRISAssistedStepConfig]:
        return [
            HRISAssistedStepConfig(
                active=self._to_bool(row[0]),
                sequence=self._to_int(row[1]),
                step_name=self._to_text(row[2]),
                action=self._to_text(row[3]),
                input_source=self._to_text(row[4]),
                method=self._to_text(row[5]),
                required=self._to_bool(row[6]),
                wait_after_seconds=self._to_float(row[7]),
                description=self._to_text(row[8]),
            )
            for row in DEFAULT_HRIS_ASSISTED_STEPS
        ]

    @staticmethod
    def _validate_assisted_step(step: HRISAssistedStepConfig) -> None:
        if step.action not in HRIS_ASSISTED_ACTIONS:
            raise ValueError(
                f"Invalid Assisted_Steps Action for {step.step_name}: "
                f"{step.action}"
            )
        if step.method not in HRIS_ASSISTED_METHODS:
            raise ValueError(
                f"Invalid Assisted_Steps Method for {step.step_name}: "
                f"{step.method}"
            )
        if step.input_source not in HRIS_ASSISTED_INPUT_SOURCES:
            raise ValueError(
                f"Invalid Assisted_Steps Input_Source for {step.step_name}: "
                f"{step.input_source}"
            )

    @staticmethod
    def get_run_controls_by_workflow(
        configuration: HRISConfiguration,
        workflow: str,
    ) -> list[HRISRunControlConfig]:
        """
        Return active Run Control list by workflow.
        """
        workflow_label = HRISConfigurationReader._normalize_workflow(
            workflow
        )

        if workflow_label == "HO":
            return configuration.ho_run_controls

        if workflow_label == "Branch":
            return configuration.branch_run_controls

        raise ValueError(
            "workflow must be 'HO' or 'Branch'."
        )

    @staticmethod
    def _to_bool(value: Any) -> bool:
        """
        Convert Excel value to boolean.
        """
        if value is None:
            return False

        value_text = str(value).strip().lower()

        return value_text in (
            "y",
            "yes",
            "true",
            "1",
            "active",
        )

    @staticmethod
    def _to_int(value: Any) -> int:
        """
        Convert Excel value to integer.
        """
        if value is None:
            return 0

        try:
            return int(value)
        except (TypeError, ValueError):
            return 0

    @staticmethod
    def _to_float(value: Any) -> float:
        if value is None:
            return 0.0
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _to_text(value: Any) -> str:
        """
        Convert Excel value to clean string.

        Important:
        Run_Control_ID must remain text.
        Example:
        001 must stay 001.
        02 must stay 02.
        """
        if value is None:
            return ""

        return str(value).strip()

    @staticmethod
    def _normalize_workflow(value: Any) -> str:
        """
        Normalize workflow value.
        """
        value_text = str(value or "").strip().lower()

        if value_text == "ho":
            return "HO"

        if value_text == "branch":
            return "Branch"

        return ""
