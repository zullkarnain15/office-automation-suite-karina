"""Read Outlook-Revisi reports as supporting revision data."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from threading import Event
from typing import Any

from openpyxl import load_workbook

from utilities.attendance_reconciliation.models import InvalidSourceRecord
from utilities.attendance_reconciliation.models import RevisionRecord
from utilities.attendance_reconciliation.models import ScanLogEntry
from utilities.attendance_reconciliation.models import SourceScanResult
from utilities.attendance_reconciliation.models import check_cancelled
from utilities.attendance_reconciliation.normalizer import normalize_date
from utilities.attendance_reconciliation.normalizer import normalize_nik
from utilities.attendance_reconciliation.normalizer import normalize_time
from utilities.attendance_reconciliation.normalizer import normalize_workflow
from utilities.attendance_reconciliation.scanner import detect_workflow
from utilities.attendance_reconciliation.scanner import discover_reports
from utilities.attendance_reconciliation.validators import OUTLOOK_VALID_HEADERS
from utilities.attendance_reconciliation.validators import validate_headers


class OutlookReportReader:
    """Scan Valid_Data while retaining Data_Anomaly only for audit."""

    def read_folder(
        self,
        root: Path,
        workflow: str,
        start_date: date,
        end_date: date,
        cancel_event: Event | None = None,
    ) -> SourceScanResult:
        result = SourceScanResult(source_type="Outlook-Revisi")
        paths, preliminary = discover_reports(root, "Outlook-Revisi", cancel_event)
        result.log_entries.extend(preliminary)
        for path in paths:
            check_cancelled(cancel_event)
            self._read_workbook(
                path, workflow, start_date, end_date, result, cancel_event
            )
        return result

    def _read_workbook(
        self,
        path: Path,
        workflow: str,
        start_date: date,
        end_date: date,
        result: SourceScanResult,
        cancel_event: Event | None,
    ) -> None:
        detected = detect_workflow(path)
        if detected and detected != workflow:
            result.log_entries.append(
                ScanLogEntry(
                    "Outlook-Revisi", path, detected, "WRONG_WORKFLOW",
                    f"Expected {workflow}, found {detected}.",
                )
            )
            return
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except PermissionError as error:
            result.log_entries.append(
                ScanLogEntry(
                    "Outlook-Revisi", path, detected, "LOCKED_FILE", str(error)
                )
            )
            return
        except Exception as error:
            result.log_entries.append(
                ScanLogEntry(
                    "Outlook-Revisi", path, detected, "CORRUPT_WORKBOOK", str(error)
                )
            )
            return

        records_before = len(result.records)
        invalid_before = len(result.invalid_records)
        records_read = 0
        dates_seen: list[date] = []
        try:
            missing_sheets = {
                "Valid_Data", "Data_Anomaly"
            } - set(workbook.sheetnames)
            if missing_sheets:
                result.log_entries.append(
                    ScanLogEntry(
                        "Outlook-Revisi", path, detected, "MISSING_SHEET",
                        "Missing sheet(s): " + ", ".join(sorted(missing_sheets)),
                    )
                )
                return

            sheet = workbook["Valid_Data"]
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if headers is None:
                result.log_entries.append(
                    ScanLogEntry(
                        "Outlook-Revisi", path, detected, "MISSING_COLUMN",
                        "Valid_Data is empty.",
                    )
                )
                return
            missing = validate_headers(headers, OUTLOOK_VALID_HEADERS)
            if missing:
                result.log_entries.append(
                    ScanLogEntry(
                        "Outlook-Revisi", path, detected, "MISSING_COLUMN",
                        "Valid_Data missing column(s): " + ", ".join(missing),
                    )
                )
                return
            columns = _column_map(headers)
            for row_number, row in enumerate(rows, 2):
                check_cancelled(cancel_event)
                if not row or not any(value not in (None, "") for value in row):
                    continue
                records_read += 1
                date_in = normalize_date(_value(row, columns, "Date_In"))
                date_out = normalize_date(_value(row, columns, "Date_Out"))
                record_date = date_in or date_out
                if record_date is not None:
                    dates_seen.append(record_date)
                if record_date is not None and not (
                    start_date <= record_date <= end_date
                ):
                    continue
                self._append_valid_record(
                    result, path, detected or workflow, workflow,
                    row_number, row, columns, date_in, date_out, record_date,
                )

            self._read_audit_anomalies(
                workbook["Data_Anomaly"], path, workflow, start_date,
                end_date, result, cancel_event,
            )
        finally:
            workbook.close()

        added = len(result.records) - records_before
        invalid_added = len(result.invalid_records) - invalid_before
        status = "USED" if added or invalid_added else "OUTSIDE_PERIOD"
        reason = "" if status == "USED" else "No records in selected period."
        result.log_entries.append(
            ScanLogEntry(
                "Outlook-Revisi", path, detected or workflow, status, reason,
                records_read,
                min(dates_seen) if dates_seen else None,
                max(dates_seen) if dates_seen else None,
            )
        )

    def _append_valid_record(
        self,
        result: SourceScanResult,
        path: Path,
        detected_workflow: str,
        selected_workflow: str,
        row_number: int,
        row: tuple[Any, ...],
        columns: dict[str, int],
        date_in: date | None,
        date_out: date | None,
        record_date: date | None,
    ) -> None:
        row_workflow = normalize_workflow(_value(row, columns, "Workflow"))
        workflow = row_workflow or detected_workflow
        nik = normalize_nik(_value(row, columns, "NIK"))
        reason = ""
        if workflow != selected_workflow:
            reason = f"Record workflow {workflow or '(empty)'} is not selected."
        elif not nik or record_date is None:
            reason = "NIK or revision date is invalid."
        elif date_in is not None and date_out is not None and date_in != date_out:
            reason = "CROSS_DATE_REVISION: Date_In and Date_Out differ."

        if reason:
            result.invalid_records.append(
                InvalidSourceRecord(
                    "Outlook-Revisi", path, "Valid_Data", row_number,
                    workflow, nik, str(record_date or ""), reason,
                )
            )
            return

        time_in_raw = _value(row, columns, "Time_In")
        time_out_raw = _value(row, columns, "Time_Out")
        time_in = normalize_time(time_in_raw)
        time_out = normalize_time(time_out_raw)
        if (time_in_raw not in (None, "") and time_in is None) or (
            time_out_raw not in (None, "") and time_out is None
        ):
            result.invalid_records.append(
                InvalidSourceRecord(
                    "Outlook-Revisi", path, "Valid_Data", row_number,
                    workflow, nik, str(record_date), "Revision time is invalid.",
                )
            )
            return

        attachment = str(_value(row, columns, "Attachment_Name") or "").strip()
        source_row = _to_int(_value(row, columns, "Source_Row"), row_number)
        result.records.append(
            RevisionRecord(
                workflow=workflow,
                nik=nik,
                attendance_date=record_date,
                revision_in=time_in,
                revision_out=time_out,
                attachment_name=attachment,
                source_report=path,
                source_row=source_row,
                date_in=date_in,
                date_out=date_out,
                source_reports=[path],
                source_rows=[source_row],
                attachments=[attachment] if attachment else [],
            )
        )

    def _read_audit_anomalies(
        self,
        sheet: Any,
        path: Path,
        workflow: str,
        start_date: date,
        end_date: date,
        result: SourceScanResult,
        cancel_event: Event | None,
    ) -> None:
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if headers is None:
            return
        columns = _column_map(headers)
        for row_number, row in enumerate(rows, 2):
            check_cancelled(cancel_event)
            if not row or not any(value not in (None, "") for value in row):
                continue
            date_in = normalize_date(_value(row, columns, "Date_In"))
            date_out = normalize_date(_value(row, columns, "Date_Out"))
            record_date = date_in or date_out
            if record_date is not None and not (
                start_date <= record_date <= end_date
            ):
                continue
            result.audit_anomalies.append(
                InvalidSourceRecord(
                    source_type="Outlook-Revisi",
                    source_report=path,
                    source_sheet="Data_Anomaly",
                    source_row=_to_int(
                        _value(row, columns, "Source_Row"), row_number
                    ),
                    workflow=normalize_workflow(
                        _value(row, columns, "Workflow")
                    ) or workflow,
                    nik=normalize_nik(_value(row, columns, "NIK")),
                    raw_date=str(record_date or ""),
                    reason=str(
                        _value(row, columns, "Anomaly_Reason")
                        or "Outlook source anomaly (audit only)."
                    ),
                    code=str(
                        _value(row, columns, "Anomaly_Code")
                        or "OUTLOOK_DATA_ANOMALY"
                    ),
                )
            )


def _column_map(headers: tuple[Any, ...]) -> dict[str, int]:
    return {
        str(value or "").strip(): index
        for index, value in enumerate(headers)
        if str(value or "").strip()
    }


def _value(row: tuple[Any, ...], columns: dict[str, int], name: str) -> Any:
    index = columns.get(name)
    return row[index] if index is not None and index < len(row) else None


def _to_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default
