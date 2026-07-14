"""Read Attendance module reports without modifying their workbooks."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from threading import Event
from typing import Any

from openpyxl import load_workbook

from utilities.attendance_reconciliation.models import InvalidSourceRecord
from utilities.attendance_reconciliation.models import MachineRecord
from utilities.attendance_reconciliation.models import ScanLogEntry
from utilities.attendance_reconciliation.models import SourceScanResult
from utilities.attendance_reconciliation.models import check_cancelled
from utilities.attendance_reconciliation.normalizer import interpret_single_tap
from utilities.attendance_reconciliation.normalizer import normalize_date
from utilities.attendance_reconciliation.normalizer import normalize_nik
from utilities.attendance_reconciliation.normalizer import normalize_time
from utilities.attendance_reconciliation.scanner import detect_workflow
from utilities.attendance_reconciliation.scanner import discover_reports
from utilities.attendance_reconciliation.validators import ATTENDANCE_ANOMALY_HEADERS
from utilities.attendance_reconciliation.validators import ATTENDANCE_DETAIL_HEADERS
from utilities.attendance_reconciliation.validators import validate_headers


class AttendanceReportReader:
    """Scan and normalize Attendance_Detail and Anomaly sheets."""

    def read_folder(
        self,
        root: Path,
        workflow: str,
        start_date: date,
        end_date: date,
        cancel_event: Event | None = None,
    ) -> SourceScanResult:
        result = SourceScanResult(source_type="Attendance")
        paths, preliminary = discover_reports(root, "Attendance", cancel_event)
        result.log_entries.extend(preliminary)
        for path in paths:
            check_cancelled(cancel_event)
            self._read_workbook(
                path, workflow, start_date, end_date, result, cancel_event
            )
        return result

    def _read_workbook(
        self,
        path: Path,
        workflow: str,
        start_date: date,
        end_date: date,
        result: SourceScanResult,
        cancel_event: Event | None,
    ) -> None:
        detected = detect_workflow(path)
        if detected and detected != workflow:
            result.log_entries.append(
                ScanLogEntry(
                    "Attendance", path, detected, "WRONG_WORKFLOW",
                    f"Expected {workflow}, found {detected}.",
                )
            )
            return

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except PermissionError as error:
            result.log_entries.append(
                ScanLogEntry(
                    "Attendance", path, detected, "LOCKED_FILE", str(error)
                )
            )
            return
        except Exception as error:
            result.log_entries.append(
                ScanLogEntry(
                    "Attendance", path, detected, "CORRUPT_WORKBOOK", str(error)
                )
            )
            return

        records_before = len(result.records)
        invalid_before = len(result.invalid_records)
        records_read = 0
        dates_seen: list[date] = []
        try:
            required_sheets = {"Attendance_Detail", "Anomaly"}
            missing_sheets = sorted(required_sheets - set(workbook.sheetnames))
            if missing_sheets:
                result.log_entries.append(
                    ScanLogEntry(
                        "Attendance", path, detected, "MISSING_SHEET",
                        "Missing sheet(s): " + ", ".join(missing_sheets),
                    )
                )
                return

            for sheet_name, is_anomaly, required_headers in (
                ("Attendance_Detail", False, ATTENDANCE_DETAIL_HEADERS),
                ("Anomaly", True, ATTENDANCE_ANOMALY_HEADERS),
            ):
                sheet = workbook[sheet_name]
                rows = sheet.iter_rows(values_only=True)
                headers = next(rows, None)
                if headers is None:
                    result.log_entries.append(
                        ScanLogEntry(
                            "Attendance", path, detected, "MISSING_COLUMN",
                            f"{sheet_name} is empty.",
                        )
                    )
                    return
                missing = validate_headers(headers, required_headers)
                if missing:
                    result.log_entries.append(
                        ScanLogEntry(
                            "Attendance", path, detected, "MISSING_COLUMN",
                            f"{sheet_name} missing column(s): " + ", ".join(missing),
                        )
                    )
                    return
                column_map = {
                    str(value or "").strip(): index
                    for index, value in enumerate(headers)
                    if str(value or "").strip()
                }
                for row_number, row in enumerate(rows, 2):
                    check_cancelled(cancel_event)
                    if not row or not any(value not in (None, "") for value in row):
                        continue
                    records_read += 1
                    raw_date = _value(row, column_map, "Tanggal")
                    attendance_date = normalize_date(raw_date)
                    if attendance_date is not None:
                        dates_seen.append(attendance_date)
                    if attendance_date is not None and not (
                        start_date <= attendance_date <= end_date
                    ):
                        continue
                    self._append_record(
                        result=result,
                        path=path,
                        workflow=detected or workflow,
                        sheet_name=sheet_name,
                        row_number=row_number,
                        row=row,
                        column_map=column_map,
                        attendance_date=attendance_date,
                        raw_date=raw_date,
                        is_anomaly=is_anomaly,
                    )
        finally:
            workbook.close()

        added = len(result.records) - records_before
        invalid_added = len(result.invalid_records) - invalid_before
        status = "USED" if added or invalid_added else "OUTSIDE_PERIOD"
        reason = "" if status == "USED" else "No records in selected period."
        result.log_entries.append(
            ScanLogEntry(
                "Attendance",
                path,
                detected or workflow,
                status,
                reason,
                records_read,
                min(dates_seen) if dates_seen else None,
                max(dates_seen) if dates_seen else None,
            )
        )

    def _append_record(
        self,
        result: SourceScanResult,
        path: Path,
        workflow: str,
        sheet_name: str,
        row_number: int,
        row: tuple[Any, ...],
        column_map: dict[str, int],
        attendance_date: date | None,
        raw_date: Any,
        is_anomaly: bool,
    ) -> None:
        nik = normalize_nik(_value(row, column_map, "NIK"))
        if not nik or attendance_date is None:
            result.invalid_records.append(
                InvalidSourceRecord(
                    "Attendance", path, sheet_name, row_number, workflow, nik,
                    str(raw_date or ""),
                    "NIK or attendance date is invalid.",
                )
            )
            return

        check_in_raw = _value(row, column_map, "Jam Masuk")
        check_out_raw = _value(row, column_map, "Jam Keluar")
        machine_in = normalize_time(check_in_raw)
        machine_out = normalize_time(check_out_raw)
        pair_status = str(_value(row, column_map, "Pair Status") or "").strip()
        original_time = None
        original_column = ""
        interpreted_in = None
        interpreted_out = None
        interpretation_rule = ""
        internal_status = ""

        if is_anomaly and pair_status.casefold() == "single_tap":
            (
                machine_in,
                machine_out,
                original_time,
                original_column,
                interpretation_rule,
                internal_status,
            ) = interpret_single_tap(check_in_raw, check_out_raw)
            interpreted_in = machine_in
            interpreted_out = machine_out
            if internal_status == "INVALID_SINGLE_TAP_TIME":
                result.invalid_records.append(
                    InvalidSourceRecord(
                        "Attendance", path, sheet_name, row_number, workflow, nik,
                        str(raw_date or ""), "SINGLE_TAP time is missing or invalid.",
                    )
                )
                return
        elif (check_in_raw not in (None, "") and machine_in is None) or (
            check_out_raw not in (None, "") and machine_out is None
        ):
            result.invalid_records.append(
                InvalidSourceRecord(
                    "Attendance", path, sheet_name, row_number, workflow, nik,
                    str(raw_date or ""), "Machine time is invalid.",
                )
            )
            return

        record = MachineRecord(
            workflow=workflow,
            nik=nik,
            name=str(_value(row, column_map, "Nama") or "").strip(),
            attendance_date=attendance_date,
            machine_in=machine_in,
            machine_out=machine_out,
            tap_count=_to_int(_value(row, column_map, "Tap Count")),
            pair_status=pair_status,
            validation_status=str(
                _value(row, column_map, "Validation Status") or ""
            ).strip(),
            remarks=str(
                _value(row, column_map, "Validation Remarks") or ""
            ).strip(),
            source_mdb=str(_value(row, column_map, "Source MDB") or "").strip(),
            source_report=path,
            source_row=row_number,
            is_anomaly=is_anomaly,
            original_anomaly_time=original_time,
            original_anomaly_column=original_column,
            interpreted_machine_in=interpreted_in,
            interpreted_machine_out=interpreted_out,
            interpretation_rule=interpretation_rule,
            internal_status=internal_status,
            source_reports=[path],
        )
        result.records.append(record)


def _value(row: tuple[Any, ...], columns: dict[str, int], name: str) -> Any:
    index = columns.get(name)
    return row[index] if index is not None and index < len(row) else None


def _to_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0
