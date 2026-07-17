"""Generate the streaming reconciliation workbook."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from threading import Event
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utilities.attendance_reconciliation.matcher import STATUS_GUIDE
from utilities.attendance_reconciliation.models import ALL_COMPARISON_STATUSES
from utilities.attendance_reconciliation.models import ComparisonRecord
from utilities.attendance_reconciliation.models import ConflictRecord
from utilities.attendance_reconciliation.models import ReconciliationRequest
from utilities.attendance_reconciliation.models import ReconciliationScan
from utilities.attendance_reconciliation.models import SourceAuditRecord
from utilities.attendance_reconciliation.models import STATUS_INVALID_SOURCE_DATA
from utilities.attendance_reconciliation.models import check_cancelled
from utilities.attendance_reconciliation.normalizer import format_date
from utilities.attendance_reconciliation.normalizer import format_time


SHEET_ORDER = (
    "Guide_Status",
    "Dashboard",
    "Summary_Per_Karyawan",
    "Comparison_Detail",
    "Machine_Anomaly",
    "Revision_Only",
    "Conflict_Review",
    "Duplicate_Source",
    "Scan_Log",
)

HEADER_FILL = "1F4E78"
STATUS_COLORS = {
    "normal": "E2F0D9",
    "attention": "FFF2CC",
    "anomaly": "FCE4D6",
    "revision": "DDEBF7",
    "conflict": "F4CCCC",
    "neutral": "E7E6E6",
}
STATUS_FILLS = {
    color: PatternFill("solid", fgColor=color)
    for color in STATUS_COLORS.values()
}


def status_color(status: str) -> str:
    if status in {
        "MACHINE_COMPLETE_NO_REVISION",
        "MACHINE_COMPLETE_REVISION_MATCH",
    }:
        return STATUS_COLORS["normal"]
    if status == "MACHINE_COMPLETE_REVISION_DIFFERENT":
        return STATUS_COLORS["attention"]
    if status.startswith("MACHINE_ANOMALY"):
        return STATUS_COLORS["anomaly"]
    if status == "REVISION_ONLY":
        return STATUS_COLORS["revision"]
    if "CONFLICT" in status or "INVALID" in status or "CROSS_DATE" in status:
        return STATUS_COLORS["conflict"]
    return STATUS_COLORS["neutral"]


def format_duration(seconds: float) -> str:
    """Format a non-negative duration as HH:MM:SS."""
    total_seconds = max(0, int(round(seconds)))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


class ReconciliationExcelWriter:
    def write(
        self,
        path: Path,
        job_id: str,
        request: ReconciliationRequest,
        scan: ReconciliationScan,
        comparisons: list[ComparisonRecord],
        conflicts: list[ConflictRecord],
        duplicates: list[SourceAuditRecord],
        summary: dict[str, Any],
        cancel_event: Event | None = None,
        process_started_at: datetime | None = None,
    ) -> Path:
        path.parent.mkdir(parents=True, exist_ok=True)
        temporary = path.with_suffix(".tmp.xlsx")
        # Write-only mode streams rows directly to the XLSX archive instead of
        # retaining millions of Cell objects in memory. Large reconciliation
        # reports can exceed eight million cells.
        workbook = Workbook(write_only=True)
        for name in SHEET_ORDER:
            workbook.create_sheet(name)
        try:
            self._guide(workbook["Guide_Status"], job_id, request)
            check_cancelled(cancel_event)
            self._employee_summary(
                workbook["Summary_Per_Karyawan"],
                scan,
                comparisons,
            )
            check_cancelled(cancel_event)
            self._comparison(workbook["Comparison_Detail"], comparisons)
            check_cancelled(cancel_event)
            self._machine_anomaly(
                workbook["Machine_Anomaly"], scan, comparisons
            )
            check_cancelled(cancel_event)
            self._revision_only(workbook["Revision_Only"], comparisons)
            self._conflicts(workbook["Conflict_Review"], conflicts)
            self._duplicates(workbook["Duplicate_Source"], duplicates)
            self._scan_log(workbook["Scan_Log"], scan)
            check_cancelled(cancel_event)
            self._stamp_process_duration(summary, process_started_at)
            self._dashboard(workbook["Dashboard"], summary)
            check_cancelled(cancel_event)
            workbook.save(temporary)
            temporary.replace(path)
        except Exception:
            if temporary.exists():
                temporary.unlink()
            raise
        finally:
            workbook.close()
        return path

    def _guide(self, sheet: Any, job_id: str, request: ReconciliationRequest) -> None:
        headers = ["Identity", "Value"]
        self._prepare_sheet(
            sheet,
            [
                "Status Code",
                "Kategori",
                "Penjelasan Bahasa Indonesia",
                "Tindakan",
                "Warna",
            ],
        )
        identity = [
            ("Report Name", "Comparison-Attendance Reconciliation"),
            ("Comparison Job ID", job_id),
            ("Workflow", request.workflow),
            ("Start Date", format_date(request.start_date)),
            ("End Date", format_date(request.end_date)),
            ("Attendance Source Root", str(request.attendance_path)),
            ("Outlook-Revisi Source Root", str(request.outlook_path)),
            ("Generated At", f"{datetime.now():%Y-%m-%d %H:%M:%S}"),
            ("Engine Version", "1.0.0"),
        ]
        self._append_header(sheet, headers)
        for row in identity:
            sheet.append(row)
        sheet.append([])
        sheet.append(["Prinsip Utama"])
        sheet.append(["Attendance merupakan sumber data utama dan paling valid."])
        sheet.append(["Outlook hanya data revisi pendukung."])
        sheet.append(["Report hanya menjelaskan hasil perbandingan."])
        sheet.append(["Report tidak mengubah data Attendance asli."])
        sheet.append([])
        status_headers = [
            "Status Code", "Kategori", "Penjelasan Bahasa Indonesia",
            "Tindakan", "Warna",
        ]
        self._append_header(sheet, status_headers)
        for status in ALL_COMPARISON_STATUSES:
            if status == STATUS_INVALID_SOURCE_DATA:
                category, description, review = (
                    "Konflik",
                    "Data sumber tidak dapat dinormalisasi atau tidak valid.",
                    True,
                )
            else:
                category, description, review = STATUS_GUIDE[status]
            row = [
                status,
                category,
                description,
                "Review manual" if review else "Tidak perlu tindakan",
                status_color(status),
            ]
            self._append_colored_row(sheet, row, status_color(status))
        self._finish_sheet(
            sheet,
            len(status_headers),
            18 + len(ALL_COMPARISON_STATUSES),
        )

    def _dashboard(self, sheet: Any, summary: dict[str, Any]) -> None:
        headers = ["Metric", "Value"]
        self._prepare_sheet(sheet, headers)
        self._append_header(sheet, headers)
        labels = [
            ("Workflow", "workflow"),
            ("Selected Start Date", "start_date"),
            ("Selected End Date", "end_date"),
            ("Attendance Reports Found", "attendance_reports_found"),
            ("Attendance Reports Used", "attendance_reports_used"),
            ("Attendance Reports Skipped", "attendance_reports_skipped"),
            ("Outlook Reports Found", "outlook_reports_found"),
            ("Outlook Reports Used", "outlook_reports_used"),
            ("Outlook Reports Skipped", "outlook_reports_skipped"),
            ("Attendance Records", "attendance_records"),
            ("Machine Complete", "machine_complete"),
            ("Machine Anomalies", "machine_anomalies"),
            ("Outlook Valid Revisions", "outlook_valid_revisions"),
            ("Revision Only", "revision_only"),
            ("Machine Source Conflict", "machine_source_conflict"),
            ("Multiple Revision Conflict", "multiple_revision_conflict"),
            ("Invalid Source Data", "invalid_source_data"),
            ("Total Comparison Records", "total_comparison_records"),
            ("Warnings", "warnings"),
            ("Process Started At", "process_started_at"),
            ("Process Completed At", "process_completed_at"),
            ("Process Duration", "process_duration"),
            ("Process Duration Seconds", "process_duration_seconds"),
        ]
        for label, key in labels:
            value = summary.get(key, "")
            if isinstance(value, list):
                value = "\n".join(value)
            sheet.append([label, value])
        self._finish_sheet(sheet, len(headers), 1 + len(labels))

    @staticmethod
    def _stamp_process_duration(
        summary: dict[str, Any],
        process_started_at: datetime | None,
    ) -> None:
        if process_started_at is None:
            return
        completed_at = datetime.now()
        seconds = max(
            0.0,
            (completed_at - process_started_at).total_seconds(),
        )
        summary["process_started_at"] = process_started_at.isoformat(
            timespec="seconds"
        )
        summary["process_completed_at"] = completed_at.isoformat(
            timespec="seconds"
        )
        summary["process_duration"] = format_duration(seconds)
        summary["process_duration_seconds"] = round(seconds, 2)

    def _employee_summary(
        self,
        sheet: Any,
        scan: ReconciliationScan,
        comparisons: list[ComparisonRecord],
    ) -> None:
        """Aggregate employee-level audit metrics without changing matching."""
        headers = [
            "NIK",
            "Nama",
            "Total Hari Attendance",
            "Total Hari Revisi",
            "Machine Complete",
            "Machine Anomaly",
            "Revision Match",
            "Revision Different",
            "Revision Only",
            "Conflict",
            "Total Perlu Review",
            "Status Akhir",
        ]
        self._prepare_sheet(sheet, headers)
        self._append_header(sheet, headers)

        employees: dict[str, dict[str, Any]] = {}

        def employee(nik: str) -> dict[str, Any]:
            return employees.setdefault(
                nik,
                {
                    "name": "",
                    "attendance_days": set(),
                    "revision_days": set(),
                    "machine_complete": 0,
                    "machine_anomaly": 0,
                    "revision_match": 0,
                    "revision_different": 0,
                    "revision_only": 0,
                    "conflict": 0,
                    "review": 0,
                    "normal": 0,
                },
            )

        for record in scan.attendance.records:
            if not record.nik:
                continue
            item = employee(record.nik)
            item["attendance_days"].add(record.attendance_date)
            if record.name and not item["name"]:
                item["name"] = record.name

        for record in scan.outlook.records:
            if not record.nik:
                continue
            employee(record.nik)["revision_days"].add(record.attendance_date)

        for comparison in comparisons:
            item = employee(comparison.key.nik)
            if comparison.machine is not None:
                if comparison.machine.is_anomaly:
                    item["machine_anomaly"] += 1
                else:
                    item["machine_complete"] += 1
            if comparison.status == "MACHINE_COMPLETE_REVISION_MATCH":
                item["revision_match"] += 1
            elif comparison.status == "MACHINE_COMPLETE_REVISION_DIFFERENT":
                item["revision_different"] += 1
            elif comparison.status == "REVISION_ONLY":
                item["revision_only"] += 1
            if "CONFLICT" in comparison.status:
                item["conflict"] += 1
            if comparison.review_required:
                item["review"] += 1
            else:
                item["normal"] += 1

        invalid_records = [
            *scan.attendance.invalid_records,
            *scan.outlook.invalid_records,
            *scan.outlook.audit_anomalies,
        ]
        for invalid in invalid_records:
            if not invalid.nik:
                continue
            item = employee(invalid.nik)
            item["conflict"] += 1
            item["review"] += 1

        for nik in sorted(employees, key=str.casefold):
            item = employees[nik]
            if item["review"] == 0:
                final_status = "NORMAL"
                color = STATUS_COLORS["normal"]
            elif item["normal"] > 0:
                final_status = "PARTIAL"
                color = STATUS_COLORS["attention"]
            else:
                final_status = "REVIEW REQUIRED"
                color = STATUS_COLORS["conflict"]
            row = [
                nik,
                item["name"],
                len(item["attendance_days"]),
                len(item["revision_days"]),
                item["machine_complete"],
                item["machine_anomaly"],
                item["revision_match"],
                item["revision_different"],
                item["revision_only"],
                item["conflict"],
                item["review"],
                final_status,
            ]
            self._append_status_row(
                sheet,
                row,
                status_column=12,
                color=color,
            )
        self._finish_sheet(sheet, len(headers), 1 + len(employees))

    def _comparison(
        self, sheet: Any, comparisons: list[ComparisonRecord]
    ) -> None:
        headers = [
            "No", "Workflow", "NIK", "Nama", "Tanggal", "Machine In",
            "Machine Out", "Machine Tap Count", "Machine Pair Status",
            "Machine Validation Status", "Machine Remarks", "Source MDB",
            "Attendance Source Report", "Attendance Source Count",
            "Original Anomaly Time", "Original Anomaly Column",
            "Interpreted Machine In", "Interpreted Machine Out",
            "Interpretation Rule", "Revision In", "Revision Out",
            "Revision Status", "Attachment Name", "Outlook Source Report",
            "Outlook Source Row", "Revision Duplicate Count",
            "Comparison Status", "Status Description", "Category",
            "Review Required", "Review Note",
        ]
        self._prepare_sheet(sheet, headers)
        self._append_header(sheet, headers)
        for number, item in enumerate(comparisons, 1):
            machine = item.machine
            revision = item.revision
            row = [
                number,
                item.key.workflow,
                item.key.nik,
                machine.name if machine else "",
                format_date(item.key.attendance_date),
                format_time(machine.machine_in) if machine else "",
                format_time(machine.machine_out) if machine else "",
                machine.tap_count if machine else "",
                machine.pair_status if machine else "",
                machine.validation_status if machine else "",
                machine.remarks if machine else "",
                machine.source_mdb if machine else "",
                "\n".join(str(path) for path in machine.source_reports)
                if machine else "",
                machine.source_count if machine else 0,
                format_time(machine.original_anomaly_time) if machine else "",
                machine.original_anomaly_column if machine else "",
                format_time(machine.interpreted_machine_in) if machine else "",
                format_time(machine.interpreted_machine_out) if machine else "",
                machine.interpretation_rule if machine else "",
                format_time(revision.revision_in) if revision else "",
                format_time(revision.revision_out) if revision else "",
                "COMPLETE" if revision and revision.complete else (
                    "INCOMPLETE" if revision else "NOT_AVAILABLE"
                ),
                "\n".join(revision.attachments) if revision else "",
                "\n".join(str(path) for path in revision.source_reports)
                if revision else "",
                ", ".join(str(row) for row in revision.source_rows)
                if revision else "",
                revision.duplicate_count if revision else 0,
                item.status,
                item.status_description,
                item.category,
                "YES" if item.review_required else "NO",
                item.review_note,
            ]
            # Coloring only the status cell preserves the visual status cue
            # without creating a style object for all 31 detail cells.
            self._append_status_row(
                sheet,
                row,
                status_column=27,
                color=status_color(item.status),
            )
        self._finish_sheet(sheet, len(headers), 1 + len(comparisons))

    def _machine_anomaly(
        self,
        sheet: Any,
        scan: ReconciliationScan,
        comparisons: list[ComparisonRecord],
    ) -> None:
        headers = [
            "No", "Workflow", "NIK", "Nama", "Tanggal", "Original In",
            "Original Out", "Tap Count", "Pair Status", "Validation Status",
            "Remarks", "Original Anomaly Time", "Original Anomaly Column",
            "Interpreted Machine In", "Interpreted Machine Out",
            "Interpretation Rule", "Internal Status", "Revision Status",
            "Comparison Status", "Source Report", "Source Row", "Review Note",
        ]
        self._prepare_sheet(sheet, headers)
        self._append_header(sheet, headers)
        status_by_key = {item.key: item for item in comparisons}
        anomalies = [item for item in scan.attendance.records if item.is_anomaly]
        for number, record in enumerate(anomalies, 1):
            comparison = status_by_key.get(record.key)
            sheet.append([
                number, record.workflow, record.nik, record.name,
                format_date(record.attendance_date),
                format_time(record.machine_in), format_time(record.machine_out),
                record.tap_count, record.pair_status, record.validation_status,
                record.remarks, format_time(record.original_anomaly_time),
                record.original_anomaly_column,
                format_time(record.interpreted_machine_in),
                format_time(record.interpreted_machine_out),
                record.interpretation_rule, record.internal_status,
                "AVAILABLE" if comparison and comparison.revision else "NOT_AVAILABLE",
                comparison.status if comparison else STATUS_INVALID_SOURCE_DATA,
                str(record.source_report), record.source_row,
                comparison.review_note if comparison else "Review manual diperlukan.",
            ])
        number = len(anomalies)
        for invalid in scan.attendance.invalid_records:
            if invalid.source_sheet != "Anomaly":
                continue
            number += 1
            sheet.append([
                number,
                invalid.workflow,
                invalid.nik,
                "",
                invalid.raw_date,
                "",
                "",
                "",
                "",
                "INVALID",
                invalid.reason,
                "",
                "",
                "",
                "",
                "",
                "INVALID_SINGLE_TAP_TIME"
                if "SINGLE_TAP" in invalid.reason
                else STATUS_INVALID_SOURCE_DATA,
                "NOT_AVAILABLE",
                STATUS_INVALID_SOURCE_DATA,
                str(invalid.source_report),
                invalid.source_row,
                "Review manual diperlukan.",
            ])
        self._finish_sheet(sheet, len(headers), 1 + number)

    def _revision_only(
        self, sheet: Any, comparisons: list[ComparisonRecord]
    ) -> None:
        headers = [
            "No", "Workflow", "NIK", "Tanggal", "Revision In",
            "Revision Out", "Attachment Name", "Outlook Source Report",
            "Outlook Source Row", "Status", "Description", "Review Note",
        ]
        self._prepare_sheet(sheet, headers)
        self._append_header(sheet, headers)
        rows = [item for item in comparisons if item.status == "REVISION_ONLY"]
        for number, item in enumerate(rows, 1):
            revision = item.revision
            sheet.append([
                number, item.key.workflow, item.key.nik,
                format_date(item.key.attendance_date),
                format_time(revision.revision_in) if revision else "",
                format_time(revision.revision_out) if revision else "",
                "\n".join(revision.attachments) if revision else "",
                "\n".join(str(path) for path in revision.source_reports)
                if revision else "",
                ", ".join(str(row) for row in revision.source_rows)
                if revision else "",
                item.status,
                "Data mesin tidak ditemukan dalam report dan periode yang dipilih.",
                item.review_note,
            ])
        self._finish_sheet(sheet, len(headers), 1 + len(rows))

    def _conflicts(self, sheet: Any, conflicts: list[ConflictRecord]) -> None:
        headers = [
            "No", "Conflict Type", "Record Key", "Source Type",
            "Source Report", "Source Row", "Values", "Reason",
            "Review Required",
        ]
        self._prepare_sheet(sheet, headers)
        self._append_header(sheet, headers)
        for number, item in enumerate(conflicts, 1):
            sheet.append([
                number, item.conflict_type,
                item.key.display() if item.key else "",
                item.source_type, item.source_report, item.source_row,
                item.values, item.reason, "YES",
            ])
        self._finish_sheet(sheet, len(headers), 1 + len(conflicts))

    def _duplicates(
        self, sheet: Any, duplicates: list[SourceAuditRecord]
    ) -> None:
        headers = [
            "No", "Duplicate Type", "Record Key", "Source Count",
            "Sources", "Values",
        ]
        self._prepare_sheet(sheet, headers)
        self._append_header(sheet, headers)
        for number, item in enumerate(duplicates, 1):
            sheet.append([
                number, item.duplicate_type, item.key.display(),
                item.source_count, "\n".join(item.sources), "\n".join(item.values),
            ])
        self._finish_sheet(sheet, len(headers), 1 + len(duplicates))

    def _scan_log(self, sheet: Any, scan: ReconciliationScan) -> None:
        headers = [
            "No", "Source Type", "File Path", "Workflow Detected", "Status",
            "Reason", "Records Read", "Period Min", "Period Max", "Timestamp",
        ]
        self._prepare_sheet(sheet, headers)
        self._append_header(sheet, headers)
        entries = [*scan.attendance.log_entries, *scan.outlook.log_entries]
        for number, item in enumerate(entries, 1):
            sheet.append([
                number, item.source_type, str(item.file_path),
                item.workflow_detected, item.status, item.reason,
                item.records_read, format_date(item.period_min),
                format_date(item.period_max),
                f"{item.timestamp:%Y-%m-%d %H:%M:%S}",
            ])
        self._finish_sheet(sheet, len(headers), 1 + len(entries))

    @staticmethod
    def _prepare_sheet(sheet: Any, headers: list[str]) -> None:
        """Set worksheet metadata before write-only row streaming begins."""
        sheet.freeze_panes = "A2"
        for column, header in enumerate(headers, 1):
            width = min(max(len(str(header)) + 2, 10), 32)
            sheet.column_dimensions[get_column_letter(column)].width = width

    @staticmethod
    def _append_header(sheet: Any, values: list[Any]) -> None:
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill = PatternFill("solid", fgColor=HEADER_FILL)
        font = Font(bold=True, color="FFFFFF")
        alignment = Alignment(vertical="top", wrap_text=True)
        cells: list[WriteOnlyCell] = []
        for value in values:
            cell = WriteOnlyCell(sheet, value=value)
            cell.border = border
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            cells.append(cell)
        sheet.append(cells)

    @staticmethod
    def _append_colored_row(
        sheet: Any,
        values: list[Any],
        color: str,
    ) -> None:
        fill = STATUS_FILLS[color]
        cells: list[WriteOnlyCell] = []
        for value in values:
            cell = WriteOnlyCell(sheet, value=value)
            cell.fill = fill
            cells.append(cell)
        sheet.append(cells)

    @staticmethod
    def _append_status_row(
        sheet: Any,
        values: list[Any],
        status_column: int,
        color: str,
    ) -> None:
        row = list(values)
        status_index = status_column - 1
        status_cell = WriteOnlyCell(sheet, value=row[status_index])
        status_cell.fill = STATUS_FILLS[color]
        row[status_index] = status_cell
        sheet.append(row)

    @staticmethod
    def _finish_sheet(
        sheet: Any,
        column_count: int,
        row_count: int,
    ) -> None:
        if column_count and row_count:
            last_column = get_column_letter(column_count)
            sheet.auto_filter.ref = f"A1:{last_column}{row_count}"
