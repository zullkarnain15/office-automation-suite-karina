"""Streaming Excel audit report for Attachment Consolidation."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill

from utilities.attachment_consolidation.models import ConsolidationResult
from utilities.attachment_consolidation.statuses import STATUS_GUIDE


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")


class ConsolidationReportWriter:
    SHEET_ORDER = (
        "Guide_Status",
        "Valid_Records",
        "Invalid_Records",
        "File_Inventory",
        "Process_Summary",
    )

    def write(self, result: ConsolidationResult) -> Path:
        path = result.artifacts.report_file
        temporary = path.with_suffix(".tmp.xlsx")
        workbook = Workbook(write_only=True)
        sheets = {
            name: workbook.create_sheet(name)
            for name in self.SHEET_ORDER
        }
        try:
            self._guide(sheets["Guide_Status"], result)
            self._valid(sheets["Valid_Records"], result)
            self._invalid(sheets["Invalid_Records"], result)
            self._inventory(sheets["File_Inventory"], result)
            self._summary(sheets["Process_Summary"], result)
            workbook.save(temporary)
            temporary.replace(path)
        except Exception:
            if temporary.exists():
                temporary.unlink()
            raise
        finally:
            workbook.close()
        return path

    def _guide(self, sheet: Any, result: ConsolidationResult) -> None:
        self._header(sheet, ["Identity", "Value"])
        for row in (
            ("Module", "Attachment Consolidation"),
            ("Job ID", result.artifacts.job_id),
            ("Mode", result.request.mode),
            ("Workflow", result.request.workflow),
            ("Source Root", str(result.request.source_root)),
            ("Output Folder", str(result.artifacts.job_folder)),
            ("Non-Destructive", "Yes"),
            ("Automatic Deduplication", "No"),
        ):
            sheet.append(row)
        sheet.append([])
        self._header(sheet, ["Status Code", "Level", "Description"])
        for row in STATUS_GUIDE:
            sheet.append(row)

    def _valid(self, sheet: Any, result: ConsolidationResult) -> None:
        self._header(
            sheet,
            [
                "No",
                "Source_File",
                "Relative_Path",
                "Source_Row",
                "Workflow",
                "NIK",
                "Date_In",
                "Time_In",
                "Date_Out",
                "Time_Out",
                "Output_TXT",
                "Status",
            ],
        )
        relative_paths = {
            item.path.resolve(): item.relative_path
            for item in result.scan.files
        }
        for number, record in enumerate(result.records, 1):
            sheet.append(
                [
                    number,
                    record.source_file.name,
                    relative_paths.get(
                        record.source_file.resolve(),
                        record.source_file.name,
                    ),
                    record.source_row,
                    result.request.workflow,
                    record.nik,
                    record.date_in.strftime("%m/%d/%Y"),
                    record.time_in,
                    record.date_out.strftime("%m/%d/%Y"),
                    record.time_out,
                    record.output_file.name if record.output_file else "",
                    "VALID",
                ]
            )

    def _invalid(self, sheet: Any, result: ConsolidationResult) -> None:
        self._header(
            sheet,
            [
                "No",
                "Source_File",
                "Relative_Path",
                "Source_Row",
                "Workflow",
                "NIK",
                "Date_In",
                "Time_In",
                "Date_Out",
                "Time_Out",
                "Status_Code",
                "Reason",
                "Raw_Value",
            ],
        )
        relative_paths = {
            item.path.resolve(): item.relative_path
            for item in result.scan.files
        }
        for number, anomaly in enumerate(result.anomalies, 1):
            sheet.append(
                [
                    number,
                    anomaly.source_file.name,
                    relative_paths.get(
                        anomaly.source_file.resolve(),
                        anomaly.source_file.name,
                    ),
                    anomaly.source_row,
                    result.request.workflow,
                    anomaly.nik,
                    anomaly.date_in,
                    anomaly.time_in,
                    anomaly.date_out,
                    anomaly.time_out,
                    anomaly.code,
                    anomaly.reason,
                    anomaly.raw_value,
                ]
            )

    def _inventory(self, sheet: Any, result: ConsolidationResult) -> None:
        self._header(
            sheet,
            [
                "No",
                "Relative_Path",
                "Extension",
                "Size_Bytes",
                "Status",
                "Row_Read",
                "Valid_Row",
                "Invalid_Row",
                "Empty_Row_Dropped",
                "Output_TXT",
                "Message",
            ],
        )
        for number, item in enumerate(result.file_results, 1):
            sheet.append(
                [
                    number,
                    item.scanned.relative_path,
                    item.scanned.extension,
                    item.scanned.size_bytes,
                    item.status,
                    item.row_read,
                    item.valid_count,
                    item.invalid_count,
                    item.empty_row_dropped,
                    "\n".join(path.name for path in item.output_files),
                    item.message or item.scanned.reason,
                ]
            )

    def _summary(self, sheet: Any, result: ConsolidationResult) -> None:
        self._header(sheet, ["Metric", "Value"])
        processable = len(result.scan.processable_files)
        skipped = len(result.scan.files) - processable
        rows: Iterable[tuple[str, object]] = (
            ("Job_ID", result.artifacts.job_id),
            ("Mode", result.request.mode),
            ("Workflow", result.request.workflow),
            ("Source_Root", str(result.request.source_root)),
            ("Output_Folder", str(result.artifacts.job_folder)),
            ("Scan_Subfolders", result.request.scan_subfolders),
            ("Started_At", result.started_at.isoformat(timespec="seconds")),
            ("Finished_At", result.finished_at.isoformat(timespec="seconds")),
            ("Duration_Seconds", round(result.duration_seconds, 3)),
            ("Final_Status", self._final_status(result)),
            ("Files_Found", len(result.scan.files)),
            ("Files_Processable", processable),
            ("Files_Skipped", skipped),
            ("Valid_Records", len(result.records)),
            ("Invalid_Records", len(result.anomalies)),
            ("TXT_Max_Lines", result.max_lines),
            ("TXT_Files_Generated", len(result.output_files)),
            ("Error_Message", result.error_message),
        )
        for row in rows:
            sheet.append(row)

    @staticmethod
    def _final_status(result: ConsolidationResult) -> str:
        if result.cancelled:
            return "CANCELLED"
        if result.success:
            return "SUCCESS"
        return "FAILED"

    @staticmethod
    def _header(sheet: Any, values: list[str]) -> None:
        cells = []
        for value in values:
            cell = WriteOnlyCell(sheet, value=value)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cells.append(cell)
        sheet.append(cells)
