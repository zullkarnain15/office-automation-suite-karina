"""
=========================================================
Office Automation Suite - Karina (OAS-K)
---------------------------------------------------------
File        : gui.py
Module      : HRIS
Version     : 1.0.0
Python      : 3.14+
=========================================================
HRIS Upload GUI - Sprint 6.23
=========================================================
"""

from __future__ import annotations

import calendar
from datetime import date
from datetime import datetime
from shutil import copy2

from openpyxl import load_workbook
import os
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from config.app_config import DATE_FORMAT
from hris.engine import HRISFullUploadEngine
from shared.config_manager import HRISConfigurationReader
from shared.logger import get_logger

logger = get_logger(__name__)

VS_CODE_BACKGROUND = "#1E1E1E"
VS_CODE_PANEL = "#252526"
VS_CODE_SURFACE = "#2D2D30"
VS_CODE_SURFACE_ACTIVE = "#3E3E42"
VS_CODE_INPUT = "#1B1B1B"
VS_CODE_BORDER = "#3C3C3C"
VS_CODE_TEXT = "#D4D4D4"
VS_CODE_MUTED_TEXT = "#9CDCFE"
VS_CODE_ACCENT = "#007ACC"
VS_CODE_ACCENT_HOVER = "#0E639C"


class HRISUploadGUI:
    """HRIS Upload GUI."""

    def __init__(self, root: tk.Tk | tk.Toplevel) -> None:
        self.root = root
        self.root.title("OAS-K - HRIS Upload")
        self.root.geometry("920x680")
        self.root.minsize(900, 640)

        self.config_file_var = tk.StringVar()
        self.txt_folder_var = tk.StringVar()
        self.output_folder_var = tk.StringVar()
        self.use_config_output_var = tk.BooleanVar(value=True)
        self.workflow_var = tk.StringVar(value="HO")
        self.local_mock_mode_var = tk.BooleanVar(value=False)
        self.start_date_var = tk.StringVar()
        self.end_date_var = tk.StringVar()

        self.status_var = tk.StringVar(value="Ready.")
        self.job_id_var = tk.StringVar(value="-")
        self.success_count_var = tk.StringVar(value="0")
        self.failed_count_var = tk.StringVar(value="0")
        self.report_folder_var = tk.StringVar(value="-")

        self._configure_dark_style()
        self._build_ui()
        self.toggle_output_source()

    def _configure_dark_style(self) -> None:
        """Configure HRIS GUI dark theme."""

        style = ttk.Style(self.root)

        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(
            ".",
            background=VS_CODE_BACKGROUND,
            foreground=VS_CODE_TEXT,
            fieldbackground=VS_CODE_INPUT,
            bordercolor=VS_CODE_BORDER,
            lightcolor=VS_CODE_BORDER,
            darkcolor=VS_CODE_BORDER,
            font=("Segoe UI", 9),
        )
        style.configure("TFrame", background=VS_CODE_BACKGROUND)
        style.configure("TLabel", background=VS_CODE_BACKGROUND, foreground=VS_CODE_TEXT)
        style.configure(
            "Muted.TLabel",
            background=VS_CODE_PANEL,
            foreground=VS_CODE_MUTED_TEXT,
        )
        style.configure(
            "TLabelframe",
            background=VS_CODE_PANEL,
            foreground=VS_CODE_MUTED_TEXT,
            bordercolor=VS_CODE_BORDER,
        )
        style.configure(
            "TLabelframe.Label",
            background=VS_CODE_PANEL,
            foreground=VS_CODE_MUTED_TEXT,
        )
        style.configure(
            "TEntry",
            fieldbackground=VS_CODE_INPUT,
            foreground=VS_CODE_TEXT,
            insertcolor=VS_CODE_TEXT,
            bordercolor=VS_CODE_BORDER,
        )
        style.map(
            "TEntry",
            fieldbackground=[
                ("disabled", VS_CODE_SURFACE),
                ("readonly", VS_CODE_INPUT),
            ],
            foreground=[
                ("disabled", VS_CODE_MUTED_TEXT),
            ],
        )
        style.configure(
            "TButton",
            background=VS_CODE_SURFACE,
            foreground=VS_CODE_TEXT,
            bordercolor=VS_CODE_BORDER,
            focusthickness=1,
            focuscolor=VS_CODE_ACCENT,
        )
        style.map(
            "TButton",
            background=[
                ("active", VS_CODE_SURFACE_ACTIVE),
                ("disabled", VS_CODE_PANEL),
            ],
            foreground=[
                ("disabled", VS_CODE_MUTED_TEXT),
            ],
        )
        style.configure(
            "TCheckbutton",
            background=VS_CODE_PANEL,
            foreground=VS_CODE_TEXT,
            indicatorbackground=VS_CODE_INPUT,
            indicatorforeground=VS_CODE_TEXT,
        )
        style.map(
            "TCheckbutton",
            background=[("active", VS_CODE_PANEL)],
            foreground=[("active", VS_CODE_MUTED_TEXT)],
        )
        style.configure(
            "TRadiobutton",
            background=VS_CODE_PANEL,
            foreground=VS_CODE_TEXT,
            indicatorbackground=VS_CODE_INPUT,
        )
        style.map(
            "TRadiobutton",
            background=[("active", VS_CODE_PANEL)],
            foreground=[("active", VS_CODE_MUTED_TEXT)],
        )

    def _build_ui(self) -> None:
        self.root.configure(bg=VS_CODE_BACKGROUND)
        container = ttk.Frame(self.root, padding=16)
        container.pack(fill="both", expand=True)

        ttk.Label(
            container,
            text="HRIS Upload Module",
            font=("Segoe UI", 18, "bold"),
        ).pack(anchor="w", pady=(0, 4))

        ttk.Label(
            container,
            text=(
                "Upload HRIS TXT files using configured Run Control IDs. "
                "Login HRIS is manual; password is not stored."
            ),
            font=("Segoe UI", 10),
        ).pack(anchor="w", pady=(0, 16))

        self._build_input_frame(container)
        self._build_action_frame(container)
        self._build_result_frame(container)
        self._build_log_frame(container)

    def _build_input_frame(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(parent, text="Input", padding=12)
        frame.pack(fill="x", pady=(0, 12))
        frame.columnconfigure(1, weight=1)

        self._add_file_row(frame, 0, "HRIS Configuration File", self.config_file_var, self._browse_config_file)
        self._add_file_row(frame, 1, "TXT Folder", self.txt_folder_var, self._browse_txt_folder)

        ttk.Label(frame, text="Output Folder").grid(row=2, column=0, sticky="w", padx=(0, 10), pady=6)
        self.output_entry = ttk.Entry(frame, textvariable=self.output_folder_var)
        self.output_entry.grid(row=2, column=1, sticky="ew", pady=6)
        self.output_browse_button = ttk.Button(
            frame,
            text="Browse",
            command=self._browse_output_folder,
        )
        self.output_browse_button.grid(row=2, column=2, padx=(8, 0), pady=6)

        ttk.Checkbutton(
            frame,
            text="Same as Configuration File",
            variable=self.use_config_output_var,
            command=self.toggle_output_source,
        ).grid(
            row=3,
            column=1,
            sticky="w",
            pady=(0, 6),
        )

        ttk.Label(frame, text="Workflow").grid(row=4, column=0, sticky="w", padx=(0, 10), pady=6)
        workflow_frame = ttk.Frame(frame)
        workflow_frame.grid(row=4, column=1, sticky="w", pady=6)
        ttk.Radiobutton(workflow_frame, text="HO", value="HO", variable=self.workflow_var).pack(side="left", padx=(0, 16))
        ttk.Radiobutton(workflow_frame, text="Branch", value="Branch", variable=self.workflow_var).pack(side="left")

        ttk.Label(frame, text="Date Range").grid(row=5, column=0, sticky="w", padx=(0, 10), pady=6)
        date_frame = ttk.Frame(frame)
        date_frame.grid(row=5, column=1, sticky="w", pady=6)
        ttk.Label(date_frame, text="Start Date").pack(side="left", padx=(0, 6))
        ttk.Entry(date_frame, textvariable=self.start_date_var, width=16).pack(side="left", padx=(0, 16))
        ttk.Button(
            date_frame,
            text="...",
            width=3,
            command=lambda: self.open_date_picker("start"),
        ).pack(side="left", padx=(0, 16))
        ttk.Label(date_frame, text="End Date").pack(side="left", padx=(0, 6))
        ttk.Entry(date_frame, textvariable=self.end_date_var, width=16).pack(side="left")
        ttk.Button(
            date_frame,
            text="...",
            width=3,
            command=lambda: self.open_date_picker("end"),
        ).pack(side="left", padx=(6, 0))

        ttk.Label(
            frame,
            text="Date format: MM/DD/YYYY. Example: 03/30/2026",
            style="Muted.TLabel",
        ).grid(row=6, column=1, sticky="w", pady=(0, 6))

        local_mock_check = ttk.Checkbutton(
            frame,
            text="Local Mock Test Mode",
            variable=self.local_mock_mode_var,
        )

        local_mock_check.grid(
            row=7,
            column=1,
            sticky="w",
            pady=(4, 6),
        )

        mock_hint_label = ttk.Label(
            frame,
            text=(
                "Use this only on personal laptop testing. "
                "It does not change the real HRIS configuration file."
            ),
            style="Muted.TLabel",
        )

        mock_hint_label.grid(
            row=8,
            column=1,
            sticky="w",
            pady=(0, 6),
        )

    def _build_action_frame(self, parent: ttk.Frame) -> None:
        frame = ttk.Frame(parent)
        frame.pack(fill="x", pady=(0, 12))

        self.start_button = ttk.Button(frame, text="Start Upload", command=self._start_upload)
        self.start_button.pack(side="left", padx=(0, 8))
        ttk.Button(frame, text="Reset", command=self._reset_form).pack(side="left", padx=(0, 8))
        ttk.Button(frame, text="Open Report Folder", command=self._open_report_folder).pack(side="left")

    def _build_result_frame(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(parent, text="Result", padding=12)
        frame.pack(fill="x", pady=(0, 12))
        frame.columnconfigure(1, weight=1)

        rows = [
            ("Status", self.status_var),
            ("Job ID", self.job_id_var),
            ("Success", self.success_count_var),
            ("Failed", self.failed_count_var),
            ("Report Folder", self.report_folder_var),
        ]

        for row_index, (label_text, variable) in enumerate(rows):
            ttk.Label(frame, text=label_text, font=("Segoe UI", 9, "bold")).grid(
                row=row_index,
                column=0,
                sticky="w",
                padx=(0, 10),
                pady=3,
            )
            ttk.Label(frame, textvariable=variable, wraplength=680).grid(
                row=row_index,
                column=1,
                sticky="w",
                pady=3,
            )

    def _build_log_frame(self, parent: ttk.Frame) -> None:
        frame = ttk.LabelFrame(parent, text="Process Log", padding=12)
        frame.pack(fill="both", expand=True)

        self.log_text = tk.Text(
            frame,
            height=12,
            wrap="word",
            state="disabled",
            bg=VS_CODE_INPUT,
            fg=VS_CODE_TEXT,
            insertbackground=VS_CODE_TEXT,
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=VS_CODE_BORDER,
        )
        self.log_text.pack(side="left", fill="both", expand=True)

        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.log_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_text.configure(yscrollcommand=scrollbar.set)

    def _add_file_row(
        self,
        frame: ttk.LabelFrame,
        row: int,
        label: str,
        variable: tk.StringVar,
        command: object,
    ) -> None:
        ttk.Label(frame, text=label).grid(row=row, column=0, sticky="w", padx=(0, 10), pady=6)
        ttk.Entry(frame, textvariable=variable).grid(row=row, column=1, sticky="ew", pady=6)
        ttk.Button(frame, text="Browse", command=command).grid(row=row, column=2, padx=(8, 0), pady=6)

    def _browse_config_file(self) -> None:
        file_path = filedialog.askopenfilename(
            title="Select HRIS Configuration File",
            filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")],
        )
        if file_path:
            self.config_file_var.set(file_path)

            if self.use_config_output_var.get():
                self.load_output_folder_from_configuration(
                    show_warning=True
                )

    def _browse_txt_folder(self) -> None:
        folder_path = filedialog.askdirectory(title="Select TXT Folder")
        if folder_path:
            self.txt_folder_var.set(folder_path)

    def _browse_output_folder(self) -> None:
        folder_path = filedialog.askdirectory(title="Select Output Folder")
        if folder_path:
            self.output_folder_var.set(folder_path)

    def toggle_output_source(self) -> None:
        """Toggle output folder source between config and manual."""

        if self.use_config_output_var.get():
            self.output_browse_button.configure(state="disabled")
            self.output_entry.configure(state="normal")
            self.load_output_folder_from_configuration(
                show_warning=False
            )
            self.output_entry.configure(state="disabled")

            if hasattr(self, "log_text"):
                self._append_log(
                    "Output folder source: HRIS Configuration."
                )
        else:
            self.output_entry.configure(state="normal")
            self.output_browse_button.configure(state="normal")

            if hasattr(self, "log_text"):
                self._append_log("Output folder source: Manual selection.")

    def load_output_folder_from_configuration(
        self,
        show_warning: bool,
    ) -> bool:
        """Load output folder from selected HRIS Configuration."""

        config_file = self.config_file_var.get().strip()

        if not config_file:
            return False

        try:
            configuration_reader = HRISConfigurationReader(config_file)
            configuration = configuration_reader.read()
            output_folder = configuration.get_output_folder()
        except Exception as exc:
            if hasattr(self, "log_text"):
                self._append_log(
                    f"Failed to read output folder from configuration: {exc}"
                )

            if show_warning:
                messagebox.showwarning(
                    "HRIS Configuration",
                    "Output Folder could not be read from "
                    "HRIS Configuration.\n\n"
                    f"{exc}",
                )

            return False

        if output_folder is None:
            if hasattr(self, "log_text"):
                self._append_log(
                    "Folder_Upload_Path is not set in HRIS Configuration."
                )

            if show_warning:
                messagebox.showwarning(
                    "HRIS Configuration",
                    "Folder_Upload_Path is not set in "
                    "HRIS Configuration.",
                )

            return False

        self.output_entry.configure(state="normal")
        self.output_folder_var.set(str(output_folder))

        if self.use_config_output_var.get():
            self.output_entry.configure(state="disabled")

        if hasattr(self, "log_text"):
            self._append_log(
                f"Output folder loaded from configuration: {output_folder}"
            )

        return True

    def _start_upload(self) -> None:
        if not self._validate_inputs():
            return

        self.start_button.configure(state="disabled")
        self._clear_log()
        self._append_log("Starting HRIS upload...")

        thread = threading.Thread(target=self._run_upload_worker, daemon=True)
        thread.start()

    def open_date_picker(self, target: str) -> None:
        """Open simple calendar date picker."""

        initial_date = self._get_initial_picker_date(target)

        picker = tk.Toplevel(self.root)
        picker.title("Select Date")
        picker.geometry("320x300")
        picker.resizable(False, False)
        picker.transient(self.root)
        picker.grab_set()
        picker.configure(bg=VS_CODE_BACKGROUND)

        selected_year = tk.IntVar(value=initial_date.year)
        selected_month = tk.IntVar(value=initial_date.month)

        header_frame = tk.Frame(picker, bg=VS_CODE_BACKGROUND)
        header_frame.pack(fill="x", pady=8)

        calendar_frame = tk.Frame(picker, bg=VS_CODE_BACKGROUND)
        calendar_frame.pack(pady=5)

        def previous_month() -> None:
            month = selected_month.get()
            year = selected_year.get()

            if month == 1:
                selected_month.set(12)
                selected_year.set(year - 1)
            else:
                selected_month.set(month - 1)

            refresh_calendar()

        def next_month() -> None:
            month = selected_month.get()
            year = selected_year.get()

            if month == 12:
                selected_month.set(1)
                selected_year.set(year + 1)
            else:
                selected_month.set(month + 1)

            refresh_calendar()

        def select_date(day_number: int) -> None:
            selected_date = date(
                selected_year.get(),
                selected_month.get(),
                day_number,
            )

            formatted_date = selected_date.strftime(DATE_FORMAT)

            if target == "start":
                self.start_date_var.set(formatted_date)
                self._append_log(f"Start Date selected: {formatted_date}")
            else:
                self.end_date_var.set(formatted_date)
                self._append_log(f"End Date selected: {formatted_date}")

            picker.destroy()

        def refresh_calendar() -> None:
            for widget in calendar_frame.winfo_children():
                widget.destroy()

            month_name = calendar.month_name[selected_month.get()]

            title_label.config(
                text=f"{month_name} {selected_year.get()}"
            )

            day_names = [
                "Mon",
                "Tue",
                "Wed",
                "Thu",
                "Fri",
                "Sat",
                "Sun",
            ]

            for column_index, day_name in enumerate(day_names):
                tk.Label(
                    calendar_frame,
                    text=day_name,
                    width=4,
                    bg=VS_CODE_BACKGROUND,
                    fg=VS_CODE_MUTED_TEXT,
                    font=("Segoe UI", 9),
                ).grid(
                    row=0,
                    column=column_index,
                    padx=2,
                    pady=2,
                )

            month_calendar = calendar.monthcalendar(
                selected_year.get(),
                selected_month.get(),
            )

            for row_index, week in enumerate(
                month_calendar,
                start=1,
            ):
                for column_index, day_number in enumerate(week):
                    if day_number == 0:
                        tk.Label(
                            calendar_frame,
                            text="",
                            width=4,
                            bg=VS_CODE_BACKGROUND,
                        ).grid(
                            row=row_index,
                            column=column_index,
                            padx=2,
                            pady=2,
                        )
                    else:
                        tk.Button(
                            calendar_frame,
                            text=str(day_number),
                            width=4,
                            command=lambda day=day_number: select_date(
                                day
                            ),
                            bg=VS_CODE_SURFACE,
                            fg=VS_CODE_TEXT,
                            activebackground=VS_CODE_SURFACE_ACTIVE,
                            activeforeground=VS_CODE_TEXT,
                            relief="flat",
                            bd=0,
                        ).grid(
                            row=row_index,
                            column=column_index,
                            padx=2,
                            pady=2,
                        )

        tk.Button(
            header_frame,
            text="<",
            width=4,
            command=previous_month,
            bg=VS_CODE_SURFACE,
            fg=VS_CODE_TEXT,
            activebackground=VS_CODE_SURFACE_ACTIVE,
            activeforeground=VS_CODE_TEXT,
            relief="flat",
            bd=0,
        ).pack(
            side="left",
            padx=10,
        )

        title_label = tk.Label(
            header_frame,
            text="",
            font=("Segoe UI", 12, "bold"),
            bg=VS_CODE_BACKGROUND,
            fg=VS_CODE_TEXT,
        )

        title_label.pack(
            side="left",
            expand=True,
        )

        tk.Button(
            header_frame,
            text=">",
            width=4,
            command=next_month,
            bg=VS_CODE_SURFACE,
            fg=VS_CODE_TEXT,
            activebackground=VS_CODE_SURFACE_ACTIVE,
            activeforeground=VS_CODE_TEXT,
            relief="flat",
            bd=0,
        ).pack(
            side="right",
            padx=10,
        )

        refresh_calendar()

    def _get_initial_picker_date(self, target: str) -> date:
        """Return initial date for date picker."""

        if target == "start":
            value = self.start_date_var.get().strip()
        else:
            value = self.end_date_var.get().strip()

        try:
            parsed_date = datetime.strptime(
                value,
                DATE_FORMAT,
            )
            return parsed_date.date()
        except ValueError:
            return date.today()

    def _run_upload_worker(self) -> None:
        """
        Run HRIS upload engine.
        """
        try:
            configuration_file = Path(
                self.config_file_var.get().strip()
            )

            if self.local_mock_mode_var.get():
                self.root.after(
                    0,
                    lambda: self._append_log(
                        "Local Mock Test Mode enabled."
                    ),
                )

                configuration_file = self._prepare_local_mock_config(
                    source_config_file=configuration_file,
                )

            engine = HRISFullUploadEngine(
                configuration_file=configuration_file,
                txt_folder=self.txt_folder_var.get(),
                output_root=self.output_folder_var.get(),
                workflow=self.workflow_var.get(),
                start_date=self.start_date_var.get(),
                end_date=self.end_date_var.get(),
                wait_for_manual_login=not self.local_mock_mode_var.get(),
            )

            result = engine.run()

            self.root.after(
                0,
                lambda: self._handle_upload_result(result),
            )

        except Exception as error:
            logger.exception("HRIS upload failed from GUI.")
            self.root.after(
                0,
                lambda upload_error=error: self._handle_upload_error(
                    upload_error
                ),
            )

    def _prepare_local_mock_config(
        self,
        source_config_file: Path,
    ) -> Path:
        """
        Create temporary HRIS configuration for local mock GUI test.

        This does not modify the real HRIS configuration file.
        """
        mock_site_folder = Path(
            r"D:\Python Project\OfficeAutomationSuite-Karina\temp\hris_mock_site"
        )

        mock_site_folder.mkdir(
            parents=True,
            exist_ok=True,
        )

        mock_upload_file = mock_site_folder / "gui_local_mock_upload.html"

        html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Mock HRIS GUI Upload - OAS-K</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            background: #f5f5f5;
            padding: 40px;
        }

        .card {
            width: 680px;
            margin: auto;
            background: white;
            border: 1px solid #ddd;
            border-radius: 8px;
            padding: 24px;
        }

        h1 {
            color: #1F4E78;
            font-size: 22px;
        }

        button {
            display: block;
            width: 100%;
            margin-top: 12px;
            padding: 10px;
            background: #1F4E78;
            color: white;
            border: none;
            border-radius: 4px;
            font-weight: bold;
            cursor: pointer;
        }

        input {
            width: 100%;
            padding: 8px;
            margin-top: 6px;
            box-sizing: border-box;
        }

        label {
            display: block;
            margin-top: 12px;
            font-weight: bold;
        }

        .section {
            margin-top: 18px;
            padding: 14px;
            border: 1px solid #ddd;
            display: none;
        }

        .success {
            color: #2E8B57;
            font-weight: bold;
            display: none;
        }
    </style>
</head>
<body>
    <div class="card">
        <h1>Mock HRIS Home</h1>
        <p>Local GUI mock upload page for OAS-K.</p>

        <button onclick="document.getElementById('tao').style.display='block';">
            Time Attendance & Overtime
        </button>

        <div id="tao" class="section">
            <button onclick="document.getElementById('upload').style.display='block';">
                Overtime Upload Attendance
            </button>
        </div>

        <div id="upload" class="section">
            <h2>Upload Attendance Page</h2>

            <label>Run Control ID</label>
            <input id="run_control_id" type="text">

            <label>Start Date</label>
            <input id="start_date" type="text">

            <label>End Date</label>
            <input id="end_date" type="text">

            <label>Attachment</label>
            <input id="attachment_file" type="file">

            <button id="upload_button" onclick="document.getElementById('upload_ok_button').style.display='block';">
                Upload
            </button>

            <button id="upload_ok_button" style="display:none;" onclick="document.getElementById('run_button').style.display='block';">
                OK
            </button>

            <button id="run_button" style="display:none;" onclick="document.getElementById('run_ok_button').style.display='block';">
                Run
            </button>

            <button id="run_ok_button" style="display:none;" onclick="document.getElementById('success_message').style.display='block';">
                OK
            </button>

            <p id="success_message" class="success">
                Upload simulated successfully.
            </p>
        </div>
    </div>
</body>
</html>
"""

        mock_upload_file.write_text(
            html,
            encoding="utf-8",
        )

        temp_config_file = Path(
            r"D:\Python Project\OfficeAutomationSuite-Karina\temp\OAS-K_HRIS_Configuration_GUI_Local_Test.xlsx"
        )

        temp_config_file.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        copy2(
            source_config_file,
            temp_config_file,
        )

        local_url = mock_upload_file.resolve().as_uri()

        workbook = load_workbook(temp_config_file)
        sheet = workbook["General"]

        for row in sheet.iter_rows(min_row=2):
            parameter = row[0].value

            if str(parameter).strip() == "HRIS_URL":
                row[1].value = local_url
                break

        workbook.save(temp_config_file)
        workbook.close()

        self.root.after(
            0,
            lambda: self._append_log(
                f"Temporary mock config created: {temp_config_file}"
            ),
        )

        self.root.after(
            0,
            lambda: self._append_log(
                f"Mock HRIS URL: {local_url}"
            ),
        )

        return temp_config_file

    def _handle_upload_result(self, result: object) -> None:
        self.start_button.configure(state="normal")
        self.status_var.set("SUCCESS" if result.success else "FAILED")
        self.job_id_var.set(result.job_id or "-")
        self.success_count_var.set(str(result.success_count))
        self.failed_count_var.set(str(result.failed_count))
        self.report_folder_var.set(str(result.report_folder or "-"))
        self._append_log(result.message)

        if result.success:
            messagebox.showinfo("HRIS Upload", "HRIS upload completed successfully.")
        else:
            messagebox.showerror("HRIS Upload", result.message)

    def _handle_upload_error(self, error: Exception) -> None:
        self.start_button.configure(state="normal")
        self.status_var.set("FAILED")
        self._append_log(str(error))
        messagebox.showerror("HRIS Upload Error", str(error))

    def _validate_inputs(self) -> bool:
        config_file = Path(self.config_file_var.get().strip())
        txt_folder = Path(self.txt_folder_var.get().strip())
        start_date = self.start_date_var.get().strip()
        end_date = self.end_date_var.get().strip()

        if not config_file.exists():
            messagebox.showerror("Validation Error", "HRIS Configuration file is required.")
            return False
        if config_file.suffix.lower() != ".xlsx":
            messagebox.showerror("Validation Error", "HRIS Configuration file must be .xlsx.")
            return False
        if not txt_folder.exists() or not txt_folder.is_dir():
            messagebox.showerror("Validation Error", "TXT Folder is required.")
            return False

        if self.use_config_output_var.get():
            self.load_output_folder_from_configuration(
                show_warning=False
            )

        output_folder = Path(self.output_folder_var.get().strip())

        if not output_folder.exists() or not output_folder.is_dir():
            messagebox.showerror(
                "Validation Error",
                "Output Folder is required.\n\n"
                "Set Folder_Upload_Path in HRIS Configuration or "
                "turn off Same as Configuration File and select it manually.",
            )
            return False
        if not start_date or not end_date:
            messagebox.showerror("Validation Error", "Start Date and End Date are required.")
            return False
        return True

    def _reset_form(self) -> None:
        self.config_file_var.set("")
        self.txt_folder_var.set("")
        self.output_folder_var.set("")
        self.workflow_var.set("HO")
        self.use_config_output_var.set(True)
        self.local_mock_mode_var.set(False)
        self.start_date_var.set("")
        self.end_date_var.set("")
        self.status_var.set("Ready.")
        self.job_id_var.set("-")
        self.success_count_var.set("0")
        self.failed_count_var.set("0")
        self.report_folder_var.set("-")
        self._clear_log()
        self.toggle_output_source()

    def _open_report_folder(self) -> None:
        report_folder = self.report_folder_var.get().strip()
        if not report_folder or report_folder == "-":
            messagebox.showinfo("Report Folder", "No report folder available yet.")
            return

        folder_path = Path(report_folder)
        if not folder_path.exists():
            messagebox.showerror("Report Folder", "Report folder does not exist.")
            return

        os.startfile(folder_path)

    def _append_log(self, message: str) -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"{message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _clear_log(self) -> None:
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")


def main() -> None:
    root = tk.Tk()
    HRISUploadGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
