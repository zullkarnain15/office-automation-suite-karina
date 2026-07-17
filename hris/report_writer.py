"""
=========================================================
Office Automation Suite - Karina (OAS-K)
---------------------------------------------------------
File        : report_writer.py
Module      : HRIS
Version     : 1.0.0
Author      : OpenAI & Zulkarnain Shiddiq
Python      : 3.14+
=========================================================
HRIS Upload Report Writer

Sprint 6.10:
- Generate Upload_Report_<Workflow>_<JOB_ID>.xlsx
- Write upload plan detail
- No browser automation yet

=========================================================
"""

from __future__ import annotations

from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from hris.artifact_writer import HRISJobArtifacts
from hris.job_manager import HRISUploadPlan
from shared.logger import get_logger

logger = get_logger(__name__)


class HRISUploadReportWriter:
    """
    HRIS Upload Report Writer.
    """

    SHEET_UPLOAD_DETAIL = "Upload_Detail"

    def write_upload_report(
        self,
        artifacts: HRISJobArtifacts,
        upload_plan: HRISUploadPlan,
    ) -> Path:
        """
        Write HRIS upload report workbook.
        """
        report_file = (
            artifacts.job_report_folder
            / f"Upload_Report_{artifacts.workflow}_{artifacts.job_id}.xlsx"
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.SHEET_UPLOAD_DETAIL

        self._write_header(
            worksheet=worksheet,
            artifacts=artifacts,
            upload_plan=upload_plan,
        )

        self._write_table(
            worksheet=worksheet,
            upload_plan=upload_plan,
        )

        self._format_sheet(worksheet)

        report_file.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        workbook.save(report_file)
        workbook.close()

        logger.info(
            "HRIS upload report created: %s",
            report_file,
        )

        return report_file

    def _write_header(
        self,
        worksheet: Any,
        artifacts: HRISJobArtifacts,
        upload_plan: HRISUploadPlan,
    ) -> None:
        """
        Write report summary header.
        """
        worksheet["A1"] = "Office Automation Suite - Karina (OAS-K)"
        worksheet["A2"] = "HRIS Upload Report"
        worksheet["A4"] = "Job ID"
        worksheet["B4"] = artifacts.job_id
        worksheet["A5"] = "Workflow"
        worksheet["B5"] = artifacts.workflow
        worksheet["A6"] = "TXT Folder"
        worksheet["B6"] = str(upload_plan.txt_folder)
        worksheet["A7"] = "Total TXT Files"
        worksheet["B7"] = upload_plan.total_txt_files
        worksheet["A8"] = "Total Run Control"
        worksheet["B8"] = upload_plan.total_run_controls
        worksheet["A9"] = "Status"
        worksheet["B9"] = "READY"
        worksheet["A10"] = "Created At"
        worksheet["B10"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        worksheet.merge_cells("A1:H1")
        worksheet.merge_cells("A2:H2")

    def _write_table(
        self,
        worksheet: Any,
        upload_plan: HRISUploadPlan,
    ) -> None:
        """
        Write upload plan table.
        """
        headers = [
            "No",
            "Workflow",
            "TXT File",
            "TXT Path",
            "Run Control ID",
            "Run Control Description",
            "Status",
            "Message",
            "Moved To",
            "Verification",
            "Process Instance",
        ]

        header_row = 12

        for column_index, header in enumerate(headers, start=1):
            cell = worksheet.cell(
                row=header_row,
                column=column_index,
                value=header,
            )
            cell.font = Font(
                bold=True,
                color="FFFFFF",
            )
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="1F4E78",
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for row_index, item in enumerate(
            upload_plan.plan_items,
            start=header_row + 1,
        ):
            worksheet.cell(
                row=row_index,
                column=1,
                value=item.sequence,
            )
            worksheet.cell(
                row=row_index,
                column=2,
                value=item.workflow,
            )
            worksheet.cell(
                row=row_index,
                column=3,
                value=item.txt_file_name,
            )
            worksheet.cell(
                row=row_index,
                column=4,
                value=str(item.txt_file_path),
            )
            worksheet.cell(
                row=row_index,
                column=5,
                value=item.run_control_id,
            )
            worksheet.cell(
                row=row_index,
                column=6,
                value=item.run_control_description,
            )
            worksheet.cell(
                row=row_index,
                column=7,
                value=item.status,
            )
            worksheet.cell(
                row=row_index,
                column=8,
                value=item.message,
            )
            worksheet.cell(
                row=row_index,
                column=9,
                value="",
            )
            worksheet.cell(
                row=row_index,
                column=10,
                value=item.verification_status,
            )
            worksheet.cell(
                row=row_index,
                column=11,
                value=item.process_instance,
            )

    def update_report_after_upload(
        self,
        report_file: str | Path,
        summary: dict[str, Any],
    ) -> None:
        """
        Update Upload_Report_<Workflow>_<JOB_ID>.xlsx after upload and file movement.

        This updates:
        - Status
        - Message
        - Moved To
        - Verification
        - Process Instance
        """
        report_path = Path(report_file)

        if not report_path.exists():
            raise FileNotFoundError(
                f"Upload report file not found: {report_path}"
            )

        workbook = load_workbook(report_path)
        worksheet = workbook[self.SHEET_UPLOAD_DETAIL]
        worksheet.freeze_panes = None

        plan_items = summary.get("plan_items", [])

        item_map = {
            str(item.get("txt_file_name", "")).strip(): item
            for item in plan_items
        }

        header_row = 12

        columns = self._get_report_columns(
            worksheet=worksheet,
            header_row=header_row,
        )

        required_columns = (
            "TXT File",
            "Status",
            "Message",
            "Moved To",
        )

        for column_name in required_columns:
            if column_name not in columns:
                raise ValueError(
                    f"Required report column not found: {column_name}"
                )

        for column_name in ("Verification", "Process Instance"):
            if column_name not in columns:
                column_index = worksheet.max_column + 1
                header_cell = worksheet.cell(
                    row=header_row,
                    column=column_index,
                    value=column_name,
                )
                header_cell.font = Font(bold=True, color="FFFFFF")
                header_cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="1F4E78",
                )
                header_cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                worksheet.column_dimensions[
                    get_column_letter(column_index)
                ].width = 18 if column_name == "Verification" else 20
                columns[column_name] = column_index

        data_start_row = header_row + 1
        max_row = worksheet.max_row

        for row_index in range(data_start_row, max_row + 1):
            txt_file_cell = worksheet.cell(
                row=row_index,
                column=columns["TXT File"],
            )

            txt_file_name = str(txt_file_cell.value or "").strip()

            if not txt_file_name:
                continue

            item = item_map.get(txt_file_name)

            if item is None:
                continue

            worksheet.cell(
                row=row_index,
                column=columns["Status"],
                value=item.get("status", ""),
            )

            worksheet.cell(
                row=row_index,
                column=columns["Message"],
                value=item.get("message", ""),
            )

            worksheet.cell(
                row=row_index,
                column=columns["Moved To"],
                value=item.get("moved_to", ""),
            )
            if "Verification" in columns:
                worksheet.cell(
                    row=row_index,
                    column=columns["Verification"],
                    value=item.get("verification_status", ""),
                )
            if "Process Instance" in columns:
                worksheet.cell(
                    row=row_index,
                    column=columns["Process Instance"],
                    value=item.get("process_instance", ""),
                )

        worksheet["B9"] = summary.get("status", "")
        workbook.save(report_path)
        workbook.close()

        logger.info(
            "HRIS upload report updated after upload: %s",
            report_path,
        )

    def _get_report_columns(
        self,
        worksheet: Any,
        header_row: int,
    ) -> dict[str, int]:
        """
        Return report column mapping from header row.
        """
        columns: dict[str, int] = {}

        for cell in worksheet[header_row]:
            header = str(cell.value or "").strip()

            if header:
                columns[header] = cell.column

        return columns

    def _format_sheet(
        self,
        worksheet: Any,
    ) -> None:
        """
        Apply readable Excel formatting.
        """
        title_font = Font(
            bold=True,
            size=14,
        )

        subtitle_font = Font(
            bold=True,
            size=12,
        )

        worksheet["A1"].font = title_font
        worksheet["A2"].font = subtitle_font

        for row in range(4, 11):
            worksheet.cell(row=row, column=1).font = Font(bold=True)

        column_widths = {
            1: 8,
            2: 14,
            3: 28,
            4: 70,
            5: 18,
            6: 32,
            7: 16,
            8: 40,
            9: 50,
            10: 18,
            11: 20,
        }

        for column_index, width in column_widths.items():
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = width

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        worksheet.auto_filter.ref = "A12:K12"
