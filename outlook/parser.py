"""
Outlook - Revisi attachment parsing and TXT output helpers.
"""

from __future__ import annotations

import csv
import secrets
from dataclasses import dataclass, field
from datetime import date
from datetime import datetime
from io import StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_HO_COLUMNS = (
    "NIK",
    "DATE IN",
    "TIME IN",
    "DATE OUT",
    "TIMEOUT",
)


@dataclass(slots=True)
class AttendanceRevisionRecord:
    """Normalized attendance revision row."""

    nik: str
    date_in: date
    time_in: str
    date_out: date
    time_out: str
    source_file: Path
    source_row: int
    output_file: Path | None = None

    def to_txt_row(self) -> list[str]:
        """Return HRIS-ready quoted TXT row values."""
        return [
            self.date_in.strftime("%m/%d/%Y"),
            self.nik,
            self.date_in.strftime("%m/%d/%Y"),
            self.time_in,
            self.date_out.strftime("%m/%d/%Y"),
            self.time_out,
        ]


@dataclass(slots=True)
class OutlookDataAnomaly:
    """Structured invalid source row retained for the process report."""

    source_file: Path
    source_row: int
    nik: str
    date_in: str
    time_in: str
    date_out: str
    time_out: str
    code: str
    reason: str
    raw_value: str = ""


@dataclass(slots=True)
class ParseResult:
    """Attachment parse result."""

    records: list[AttendanceRevisionRecord]
    errors: list[str]
    anomalies: list[OutlookDataAnomaly] = field(default_factory=list)
    row_read: int = 0
    empty_row_dropped: int = 0

    @property
    def valid(self) -> bool:
        return not self.errors


class OutlookAttachmentParser:
    """Parse Outlook - Revisi attachments into normalized rows."""

    def parse(
        self,
        path: str | Path,
        workflow: str,
    ) -> ParseResult:
        """Parse an attachment according to workflow."""
        attachment_path = Path(path)
        workflow_label = self._normalize_workflow(workflow)

        if workflow_label == "HO":
            return self.parse_ho_excel(attachment_path)

        if workflow_label == "Branch":
            return self.parse_branch_txt(attachment_path)

        return ParseResult([], [f"Unsupported workflow: {workflow}"])

    def parse_ho_excel(self, path: Path) -> ParseResult:
        """Parse HO Excel attachment."""
        errors: list[str] = []
        anomalies: list[OutlookDataAnomaly] = []
        records: list[AttendanceRevisionRecord] = []
        row_read = 0
        empty_row_dropped = 0

        try:
            workbook = load_workbook(path, data_only=True, read_only=True)
            sheet = workbook.active
            rows = iter(sheet.iter_rows(values_only=True))
        except Exception as error:
            reason = f"Failed to read Excel attachment: {error}"
            return ParseResult(
                [], [reason],
                [OutlookDataAnomaly(path, 0, "", "", "", "", "", "FILE_READ_FAILED", reason)],
            )

        try:
            header_index: int | None = None
            header_row: tuple[Any, ...] | None = None
            saw_row = False
            required_headers = {
                self._normalize_header(value)
                for value in REQUIRED_HO_COLUMNS
            }
            for index, row in enumerate(rows):
                saw_row = True
                actual_headers = {
                    self._normalize_header(value)
                    for value in row
                    if self._to_text(value)
                }
                if required_headers.issubset(actual_headers):
                    header_index = index
                    header_row = row
                    break

            if not saw_row:
                reason = "Excel attachment is empty."
                return ParseResult(
                    [], [reason],
                    [OutlookDataAnomaly(path, 0, "", "", "", "", "", "ATTACHMENT_STRUCTURE_INVALID", reason)],
                )

            if header_index is None or header_row is None:
                reason = (
                    "Missing required HO column(s): "
                    + ", ".join(REQUIRED_HO_COLUMNS)
                )
                return ParseResult(
                    [],
                    [reason],
                    [
                        OutlookDataAnomaly(
                            path, 0, "", "", "", "", "",
                            "MISSING_REQUIRED_COLUMN", reason,
                        )
                    ],
                )
            headers = [self._normalize_header(value) for value in header_row]
            column_map = {
                header: index
                for index, header in enumerate(headers)
                if header
            }

            for row_number, row in enumerate(rows, header_index + 2):
                row_read += 1
                if self._is_empty_ho_data_row(row, column_map):
                    empty_row_dropped += 1
                    continue

                record = self._record_from_values(
                    nik=self._value_at(row, column_map["NIK"]),
                    date_in=self._value_at(row, column_map["DATE IN"]),
                    time_in=self._value_at(row, column_map["TIME IN"]),
                    date_out=self._value_at(row, column_map["DATE OUT"]),
                    time_out=self._value_at(row, column_map["TIMEOUT"]),
                    source_file=path,
                    source_row=row_number,
                    errors=errors,
                    anomalies=anomalies,
                )

                if record is not None:
                    records.append(record)
        finally:
            try:
                workbook.close()
            except Exception:
                pass

        return ParseResult(records, errors, anomalies, row_read, empty_row_dropped)

    @classmethod
    def _is_empty_ho_data_row(
        cls,
        row: tuple[Any, ...],
        column_map: dict[str, int],
    ) -> bool:
        """Ignore unused template rows with at most one meaningful value."""
        if not row:
            return True

        values = [
            cls._value_at(row, column_map[column])
            for column in REQUIRED_HO_COLUMNS
        ]
        populated_count = sum(
            not cls._is_empty_template_value(value)
            for value in values
        )
        return populated_count <= 1

    @staticmethod
    def _is_empty_template_value(value: Any) -> bool:
        if value is None or not str(value).strip():
            return True

        if isinstance(value, (int, float)) and value == 0:
            return True

        if (
            hasattr(value, "hour")
            and hasattr(value, "minute")
            and not isinstance(value, datetime)
            and value.hour == 0
            and value.minute == 0
            and getattr(value, "second", 0) == 0
        ):
            return True

        return False

    def parse_branch_txt(self, path: Path) -> ParseResult:
        """Parse Branch TXT attachment with six comma-separated columns."""
        errors: list[str] = []
        anomalies: list[OutlookDataAnomaly] = []
        records: list[AttendanceRevisionRecord] = []
        row_read = 0
        empty_row_dropped = 0

        try:
            text = self._decode_branch_text(path)
            sample = text[:4096]
            delimiter = "\t" if "\t" in sample and "," not in sample else ","
            with StringIO(text, newline="") as handle:
                reader = csv.reader(handle, delimiter=delimiter)
                for row_number, row in enumerate(reader, 1):
                    row_read += 1
                    if not row or not any(value.strip() for value in row):
                        empty_row_dropped += 1
                        continue

                    values = [value.strip().strip('"') for value in row]
                    if row_number == 1 and self._looks_like_header(values):
                        continue

                    if len(values) != 6:
                        reason = (
                            f"{path.name} row {row_number}: expected 6 columns, "
                            f"got {len(values)}."
                        )
                        errors.append(reason)
                        anomalies.append(
                            OutlookDataAnomaly(
                                path, row_number, "", "", "", "", "",
                                "COLUMN_COUNT", reason, self._raw_value(values),
                            )
                        )
                        continue

                    record = self._record_from_values(
                        nik=values[1],
                        date_in=values[0],
                        time_in=values[3],
                        date_out=values[4],
                        time_out=values[5],
                        source_file=path,
                        source_row=row_number,
                        errors=errors,
                        anomalies=anomalies,
                    )

                    if record is not None:
                        records.append(record)
        except Exception as error:
            reason = f"Failed to read TXT attachment: {error}"
            return ParseResult(
                [], [reason],
                [OutlookDataAnomaly(path, 0, "", "", "", "", "", "FILE_READ_FAILED", reason)],
                row_read,
                empty_row_dropped,
            )

        return ParseResult(records, errors, anomalies, row_read, empty_row_dropped)

    @staticmethod
    def _decode_branch_text(path: Path) -> str:
        """Decode Branch TXT without silently discarding invalid bytes."""
        raw_data = path.read_bytes()
        encodings = (
            ("utf-16",)
            if raw_data.startswith((b"\xff\xfe", b"\xfe\xff"))
            else ("utf-8-sig", "cp1252")
        )
        decode_errors: list[str] = []

        for encoding in encodings:
            try:
                return raw_data.decode(encoding)
            except UnicodeDecodeError as error:
                decode_errors.append(f"{encoding}: {error}")

        raise UnicodeError(
            f"Unable to decode {path.name}. Tried: "
            + "; ".join(decode_errors)
        )

    def _record_from_values(
        self,
        nik: Any,
        date_in: Any,
        time_in: Any,
        date_out: Any,
        time_out: Any,
        source_file: Path,
        source_row: int,
        errors: list[str],
        anomalies: list[OutlookDataAnomaly],
    ) -> AttendanceRevisionRecord | None:
        nik_text = self._to_text(nik)
        date_in_value = self._parse_date(date_in)
        date_out_value = self._parse_date(date_out)
        time_in_text = self._parse_time(time_in)
        time_out_text = self._parse_time(time_out)
        prefix = f"{source_file.name} row {source_row}"

        failure: tuple[str, str] | None = None
        if not nik_text:
            failure = ("NIK_REQUIRED", f"{prefix}: NIK is required.")
        elif date_in_value is None:
            failure = ("DATE_FORMAT", f"{prefix}: DATE IN is invalid.")
        elif date_out_value is None:
            failure = ("DATE_FORMAT", f"{prefix}: DATE OUT is invalid.")
        elif time_in_text is None:
            failure = ("TIME_FORMAT", f"{prefix}: TIME IN is invalid.")
        elif time_out_text is None:
            failure = ("TIME_FORMAT", f"{prefix}: TIMEOUT is invalid.")

        if failure is not None:
            code, reason = failure
            errors.append(reason)
            anomalies.append(
                OutlookDataAnomaly(
                    source_file=source_file,
                    source_row=source_row,
                    nik=nik_text,
                    date_in=self._to_text(date_in),
                    time_in=self._to_text(time_in),
                    date_out=self._to_text(date_out),
                    time_out=self._to_text(time_out),
                    code=code,
                    reason=reason,
                    raw_value=self._raw_value(
                        [date_in, nik, date_in, time_in, date_out, time_out]
                    ),
                )
            )
            return None

        return AttendanceRevisionRecord(
            nik=nik_text,
            date_in=date_in_value,
            time_in=time_in_text,
            date_out=date_out_value,
            time_out=time_out_text,
            source_file=source_file,
            source_row=source_row,
        )

    @staticmethod
    def _raw_value(values: list[Any]) -> str:
        """Return a compact audit value without retaining binary content."""
        return " | ".join(str(value or "") for value in values)[:1000]

    @classmethod
    def _find_header_index(
        cls,
        rows: list[tuple[Any, ...]],
        required_headers: tuple[str, ...],
    ) -> int:
        required = {
            cls._normalize_header(header)
            for header in required_headers
        }

        for index, row in enumerate(rows):
            actual = {
                cls._normalize_header(value)
                for value in row
                if cls._to_text(value)
            }
            if required.issubset(actual):
                return index

        raise ValueError(
            "Missing required HO column(s): "
            + ", ".join(required_headers)
        )

    @staticmethod
    def _value_at(row: tuple[Any, ...], index: int) -> Any:
        if index >= len(row):
            return None
        return row[index]

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return OutlookAttachmentParser._to_text(value).upper().replace("_", " ")

    @staticmethod
    def _to_text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _looks_like_header(values: list[str]) -> bool:
        normalized = {value.strip().upper().replace("_", " ") for value in values}
        return bool({"NIK", "DATE IN", "TIME IN", "DATE OUT", "TIMEOUT"} & normalized)

    @staticmethod
    def _parse_date(value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        value_text = OutlookAttachmentParser._to_text(value)

        if not value_text:
            return None

        for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(value_text, fmt).date()
            except ValueError:
                continue

        return None

    @staticmethod
    def _parse_time(value: Any) -> str | None:
        if isinstance(value, datetime):
            return value.strftime("%H:%M")

        if hasattr(value, "hour") and hasattr(value, "minute"):
            return f"{value.hour:02d}:{value.minute:02d}"

        value_text = OutlookAttachmentParser._to_text(value)

        if not value_text:
            return None

        for fmt in ("%H:%M", "%H:%M:%S"):
            try:
                return datetime.strptime(value_text, fmt).strftime("%H:%M")
            except ValueError:
                continue

        return None

    @staticmethod
    def _normalize_workflow(value: str) -> str:
        value_text = str(value or "").strip().lower()
        if value_text == "ho":
            return "HO"
        if value_text == "branch":
            return "Branch"
        return ""


class OutlookTxtWriter:
    """Write normalized Outlook - Revisi rows into split TXT files."""

    def write(
        self,
        records: list[AttendanceRevisionRecord],
        output_folder: str | Path,
        workflow: str,
        max_lines: int,
        job_id: str,
        prefix: str = "Outlook_Revisi",
    ) -> list[Path]:
        """Write records into one or more TXT files."""
        if not records:
            return []

        output_path = Path(output_folder)
        output_path.mkdir(parents=True, exist_ok=True)
        workflow_label = OutlookAttachmentParser._normalize_workflow(workflow)
        safe_prefix = f"{prefix}_{workflow_label or workflow}"
        date_suffix = str(job_id).split("_", maxsplit=1)[0]
        limit = max(max_lines, 1)
        files: list[Path] = []
        remaining = list(records)
        existing_files = self._existing_output_files(
            output_path,
            safe_prefix,
            date_suffix,
        )

        if existing_files:
            current_file = existing_files[-1]
            current_count = self._line_count(current_file)
            available = max(limit - current_count, 0)
            if available:
                chunk = remaining[:available]
                self._write_chunk(current_file, chunk, mode="a")
                files.append(current_file)
                remaining = remaining[len(chunk):]

        file_index = self._next_output_index(
            output_path,
            safe_prefix,
            date_suffix,
        )
        used_random_suffixes = self._existing_random_suffixes(output_path)
        for start in range(0, len(remaining), limit):
            chunk = remaining[start:start + limit]
            random_suffix = self._unique_random_suffix(
                used_random_suffixes
            )
            used_random_suffixes.add(random_suffix)
            file_path = output_path / (
                f"{safe_prefix}_{file_index:03d}_{date_suffix}_"
                f"{random_suffix}.txt"
            )
            self._write_chunk(file_path, chunk, mode="w")
            files.append(file_path)
            file_index += 1

        return files

    @staticmethod
    def _write_chunk(
        file_path: Path,
        records: list[AttendanceRevisionRecord],
        mode: str,
    ) -> None:
        with file_path.open(mode, encoding="utf-8", newline="") as handle:
            writer = csv.writer(
                handle,
                quoting=csv.QUOTE_ALL,
                lineterminator="\n",
            )
            for record in records:
                writer.writerow(record.to_txt_row())
                record.output_file = file_path

    @staticmethod
    def _line_count(path: Path) -> int:
        with path.open("r", encoding="utf-8", newline="") as handle:
            return sum(1 for _line in handle)

    @staticmethod
    def _next_output_index(
        output_path: Path,
        safe_prefix: str,
        date_suffix: str,
    ) -> int:
        """Return the next free job-wide TXT sequence number."""
        existing = OutlookTxtWriter._existing_output_files(
            output_path,
            safe_prefix,
            date_suffix,
        )
        indexes = [
            OutlookTxtWriter._output_index(path, safe_prefix, date_suffix)
            for path in existing
        ]
        return max((index for index in indexes if index is not None), default=0) + 1

    @staticmethod
    def _existing_output_files(
        output_path: Path,
        safe_prefix: str,
        date_suffix: str,
    ) -> list[Path]:
        """Return legacy and randomized TXT files in sequence order."""
        candidates = {
            *output_path.glob(f"{safe_prefix}_*_{date_suffix}.txt"),
            *output_path.glob(f"{safe_prefix}_*_{date_suffix}_???.txt"),
        }
        return sorted(
            candidates,
            key=lambda path: (
                OutlookTxtWriter._output_index(
                    path,
                    safe_prefix,
                    date_suffix,
                )
                or 0,
                path.name,
            ),
        )

    @staticmethod
    def _output_index(
        path: Path,
        safe_prefix: str,
        date_suffix: str,
    ) -> int | None:
        stem = path.stem
        prefix = f"{safe_prefix}_"
        if not stem.startswith(prefix):
            return None
        parts = stem[len(prefix):].split("_")
        if len(parts) not in {2, 3} or parts[1] != date_suffix:
            return None
        try:
            return int(parts[0])
        except ValueError:
            return None

    @staticmethod
    def _existing_random_suffixes(output_path: Path) -> set[str]:
        return {
            path.stem.rsplit("_", maxsplit=1)[-1]
            for path in output_path.glob("*.txt")
            if (
                len(path.stem.rsplit("_", maxsplit=1)[-1]) == 3
                and path.stem.rsplit("_", maxsplit=1)[-1].isdigit()
            )
        }

    @staticmethod
    def _unique_random_suffix(used_suffixes: set[str]) -> str:
        """Return an unused zero-padded random number from 000 through 999."""
        if len(used_suffixes) >= 1000:
            raise RuntimeError("All three-digit TXT filename suffixes are in use.")

        for _attempt in range(32):
            candidate = f"{secrets.randbelow(1000):03d}"
            if candidate not in used_suffixes:
                return candidate

        for number in range(1000):
            candidate = f"{number:03d}"
            if candidate not in used_suffixes:
                return candidate
        raise RuntimeError("Unable to reserve a unique TXT filename suffix.")
