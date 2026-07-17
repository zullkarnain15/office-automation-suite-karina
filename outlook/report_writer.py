"""Excel process report writer for Outlook - Revisi."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from shared.logger import get_logger

logger = get_logger(__name__)


class OutlookProcessReportWriter:
    """Write the seven-sheet audit report from engine result data only."""

    TABLE_FILL = PatternFill("solid", fgColor="2F75B5")
    TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
    SUCCESS_FILL = PatternFill("solid", fgColor="C6EFCE")
    FAILED_FILL = PatternFill("solid", fgColor="FFC7CE")
    NEUTRAL_FILL = PatternFill("solid", fgColor="D9E1F2")
    WARNING_FILL = PatternFill("solid", fgColor="FFEB9C")

    def write(self, result: Any, configuration: Any) -> Path:
        """Create a report without re-reading source attachments or TXT files."""
        if result.report_file is None:
            raise ValueError("Report output path is not available.")

        report_path = Path(result.report_file)
        workbook = Workbook()
        dashboard = workbook.active
        dashboard.title = "Dashboard"
        self._write_dashboard(dashboard, result, configuration)
        self._write_email_result(workbook.create_sheet("Email_Result"), result)
        self._write_attachment_result(
            workbook.create_sheet("Attachment_Result"), result
        )
        self._write_valid_data(workbook.create_sheet("Valid_Data"), result)
        self._write_employee_summary(
            workbook.create_sheet("Summary_Per_Karyawan"), result
        )
        self._write_anomalies(workbook.create_sheet("Data_Anomaly"), result)
        self._write_output_summary(
            workbook.create_sheet("Output_Summary"), result
        )

        report_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(report_path)
        workbook.close()
        return report_path

    def _write_dashboard(
        self, sheet: Any, result: Any, configuration: Any
    ) -> None:
        general = configuration.general
        start = result.start_time
        end = result.end_time
        duration = (end - start).total_seconds() if start and end else 0
        messages = self._report_messages(result)
        attachment_count = sum(
            1
            for message in messages
            for item in message.attachment_results
            if item.file_status != "SKIPPED"
        )
        reply_sent = sum(1 for item in messages if item.reply_result == "SENT")
        reply_failed = sum(1 for item in messages if item.reply_result == "FAILED")
        moved = sum(1 for item in messages if item.move_result == "MOVED")
        not_moved = sum(
            1 for item in messages if item.move_result in {"NOT_MOVED", "FAILED"}
        )
        sent_copy_moved = sum(
            1 for item in messages
            if item.sent_copy_status == "MOVED_TO_SENT"
        )
        sent_copy_pending = sum(
            1 for item in messages
            if item.sent_copy_status == "COPY_PENDING"
        )
        sent_copy_failed = sum(
            1 for item in messages
            if item.sent_copy_status in {"BCC_REJECTED", "MOVE_FAILED"}
        )

        sheet["A1"] = str(
            general.get("Module_Display_Name", "Outlook - Revisi")
        )
        sheet["A2"] = "Outlook Process Report"
        sheet.merge_cells("A1:D1")
        sheet.merge_cells("A2:D2")
        sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
        sheet["A1"].fill = self.TITLE_FILL
        sheet["A2"].font = Font(bold=True, size=12)

        identity = [
            ("Module", general.get("Module_Display_Name", "Outlook - Revisi")),
            ("Job_ID", result.job_id),
            ("Workflow", result.workflow),
            ("Payroll_Period", general.get("Payroll_Period", "")),
            ("Mailbox", general.get("Mailbox_SMTP", "")),
            ("Source_Folder", general.get("Source_Folder", "Inbox")),
            ("Start_Time", self._datetime(start)),
            ("End_Time", self._datetime(end)),
            ("Duration", f"{duration:.2f} seconds"),
            ("Final_Status", result.final_status),
            ("Output_Folder", str(result.output_folder)),
        ]
        summary = [
            ("Email_Found", result.target_email),
            ("Email_Target_Workflow", result.target_email),
            ("Email_Success", result.success_email),
            ("Email_Failed", result.failed_email),
            ("Attachment_Processed", attachment_count),
            ("Valid_Row_Count", result.valid_row_count),
            ("Anomaly_Row_Count", result.anomaly_row_count),
            ("TXT_File_Count", result.output_txt_count),
            ("Reply_Sent_Count", reply_sent),
            ("Reply_Failed_Count", reply_failed),
            ("Email_Moved_Count", moved),
            ("Email_Not_Moved_Count", not_moved),
            ("Sent_Copy_Moved_Count", sent_copy_moved),
            ("Sent_Copy_Pending_Count", sent_copy_pending),
            ("Sent_Copy_Failed_Count", sent_copy_failed),
            ("Reconciliation_Status", result.reconciliation_status),
            ("Reconciliation_Issue", "\n".join(result.reconciliation_issues)),
        ]
        self._write_key_values(sheet, 4, 1, "Process Identity", identity)
        self._write_key_values(sheet, 4, 3, "Process Summary", summary)
        sheet.column_dimensions["A"].width = 26
        sheet.column_dimensions["B"].width = 55
        sheet.column_dimensions["C"].width = 32
        sheet.column_dimensions["D"].width = 55
        final_cell = sheet.cell(14, 2)
        final_cell.fill = self._status_fill(result.final_status)
        sheet.freeze_panes = "A4"

    def _write_email_result(self, sheet: Any, result: Any) -> None:
        headers = [
            "No", "Received_Time", "Workflow_Selected",
            "Workflow_Detected", "Sender_Name", "Sender_SMTP", "Required_CC",
            "Actual_CC", "Subject", "Expected_Subject", "Attachment_Count",
            "Validation_Sender", "Validation_CC", "Validation_Subject",
            "Validation_Attachment", "Validation_Data", "Final_Result",
            "Failure_Code", "Failure_Reason", "Reply_Result", "Reply_From",
            "Move_Result", "Processed_Time", "Output_Folder", "Sent_Copy_ID",
            "Sent_Copy_Status", "Sent_Copy_Detail",
        ]
        rows = []
        for number, item in enumerate(self._report_messages(result), 1):
            rows.append([
                number, self._datetime(item.received_time),
                result.workflow, item.detected_workflow, item.sender_name,
                item.sender_email, item.required_cc, item.actual_cc, item.subject,
                item.expected_subject, item.attachment_count,
                item.validation_sender, item.validation_cc,
                item.validation_subject, item.validation_attachment,
                item.validation_data, item.status, item.failure_code,
                "\n".join(item.errors), item.reply_result, item.reply_from,
                item.move_result, self._datetime(item.processed_time),
                str(result.output_folder), item.sent_copy_id,
                item.sent_copy_status, item.sent_copy_detail,
            ])
        self._write_table(
            sheet,
            headers,
            rows,
            status_columns=(17, 20, 22, 26),
        )

    def _write_attachment_result(self, sheet: Any, result: Any) -> None:
        headers = [
            "No", "Workflow", "Original_File_Name",
            "Saved_File_Name", "File_Extension", "File_Size_KB", "File_Status",
            "Row_Read", "Row_Valid", "Row_Anomaly", "Duplicate_Row",
            "Empty_Row_Dropped", "Output_TXT", "Error_Message",
        ]
        rows = []
        number = 1
        for message in self._report_messages(result):
            for item in message.attachment_results:
                if item.file_status == "SKIPPED":
                    continue
                rows.append([
                    number, message.workflow,
                    item.original_file_name, item.saved_file_name,
                    item.file_extension, item.file_size_kb, item.file_status,
                    item.row_read, item.row_valid, item.row_anomaly,
                    item.duplicate_row, item.empty_row_dropped,
                    "\n".join(path.name for path in item.output_txt),
                    item.error_message,
                ])
                number += 1
        self._write_table(sheet, headers, rows, status_columns=(7,))

    def _write_valid_data(self, sheet: Any, result: Any) -> None:
        headers = [
            "No", "Attachment_Name", "Source_Row", "Workflow",
            "NIK", "Date_In", "Time_In", "Date_Out", "Time_Out", "Output_TXT",
        ]
        rows = []
        number = 1
        for message in self._report_messages(result):
            for record in message.valid_records:
                rows.append([
                    number, record.source_file.name,
                    record.source_row, message.workflow, record.nik,
                    record.date_in.strftime("%m/%d/%Y"), record.time_in,
                    record.date_out.strftime("%m/%d/%Y"), record.time_out,
                    record.output_file.name if record.output_file else "",
                ])
                number += 1
        self._write_table(sheet, headers, rows)
        for cell in sheet["E"][1:]:
            cell.number_format = "@"

    def _write_anomalies(self, sheet: Any, result: Any) -> None:
        headers = [
            "No", "Attachment_Name", "Source_Row", "Workflow",
            "NIK", "Date_In", "Time_In", "Date_Out", "Time_Out",
            "Anomaly_Code", "Anomaly_Reason", "Raw_Value",
        ]
        rows = []
        number = 1
        for message in self._report_messages(result):
            for anomaly in message.anomalies:
                rows.append([
                    number, anomaly.source_file.name,
                    anomaly.source_row, message.workflow, anomaly.nik,
                    anomaly.date_in, anomaly.time_in, anomaly.date_out,
                    anomaly.time_out, anomaly.code, anomaly.reason,
                    anomaly.raw_value[:1000],
                ])
                number += 1
        self._write_table(sheet, headers, rows)
        for cell in sheet["E"][1:]:
            cell.number_format = "@"

    def _write_employee_summary(self, sheet: Any, result: Any) -> None:
        """Write revision totals grouped by employee NIK."""
        headers = [
            "NIK",
            "Total Hari Revisi",
            "Valid For TXT",
            "Anomaly",
            "Total File Sumber",
            "Status",
        ]
        summary: dict[str, dict[str, Any]] = {}

        def employee_item(nik: Any) -> dict[str, Any]:
            employee_nik = str(nik or "").strip() or "(EMPTY)"
            return summary.setdefault(
                employee_nik,
                {
                    "dates": set(),
                    "valid": 0,
                    "anomaly": 0,
                    "source_files": set(),
                },
            )

        for message in self._report_messages(result):
            for record in message.valid_records:
                item = employee_item(record.nik)
                date_key = self._revision_date_key(record.date_in)
                if date_key:
                    item["dates"].add(date_key)
                item["valid"] += 1
                item["source_files"].add(str(record.source_file))

            for anomaly in message.anomalies:
                item = employee_item(anomaly.nik)
                date_key = self._revision_date_key(anomaly.date_in)
                if date_key:
                    item["dates"].add(date_key)
                item["anomaly"] += 1
                item["source_files"].add(str(anomaly.source_file))

        rows = []
        for nik, item in sorted(summary.items()):
            valid_count = item["valid"]
            anomaly_count = item["anomaly"]
            if valid_count and anomaly_count:
                status = "PARTIAL"
            elif valid_count:
                status = "VALID"
            else:
                status = "ANOMALY"
            rows.append(
                [
                    nik,
                    len(item["dates"]),
                    valid_count,
                    anomaly_count,
                    len(item["source_files"]),
                    status,
                ]
            )

        self._write_table(sheet, headers, rows, status_columns=(6,))
        for cell in sheet["A"][1:]:
            cell.number_format = "@"
        for row in sheet.iter_rows(
            min_row=2,
            min_col=2,
            max_col=5,
        ):
            for cell in row:
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="top")

    def _write_output_summary(self, sheet: Any, result: Any) -> None:
        headers = [
            "No", "Workflow", "Output_Type", "File_Name", "Full_Path",
            "Line_Count", "Created_Time", "File_Status", "Source_Email_Count",
            "Source_Attachment_Count",
        ]
        source_attachments = sum(
            1
            for message in self._report_messages(result)
            for item in message.attachment_results
            if item.file_status != "SKIPPED"
        )
        created = self._datetime(result.end_time or datetime.now())
        unique_txt: dict[Path, int] = {}
        for message in self._report_messages(result):
            for record in message.valid_records:
                if record.output_file is not None:
                    unique_txt[record.output_file] = (
                        unique_txt.get(record.output_file, 0) + 1
                    )
        rows = []
        number = 1
        for path, line_count in unique_txt.items():
            rows.append([
                number, result.workflow, "HRIS_TXT", path.name, str(path),
                line_count, created, "CREATED", result.success_email,
                source_attachments,
            ])
            number += 1
        artifacts = [
            ("EXCEL_REPORT", result.report_file, "CREATED"),
            ("PROCESS_LOG", result.process_log, "CREATED"),
            ("SUMMARY_JSON", result.summary_json, "CREATED"),
        ]
        for output_type, path, status in artifacts:
            if path is None:
                continue
            rows.append([
                number, result.workflow, output_type, Path(path).name,
                str(path), "", created, status, result.target_email,
                source_attachments,
            ])
            number += 1
        self._write_table(sheet, headers, rows, status_columns=(8,))

    def _write_key_values(
        self,
        sheet: Any,
        start_row: int,
        start_column: int,
        title: str,
        values: Iterable[tuple[str, Any]],
    ) -> None:
        title_cell = sheet.cell(start_row, start_column, title)
        title_cell.font = Font(bold=True, color="FFFFFF")
        title_cell.fill = self.TABLE_FILL
        sheet.cell(start_row, start_column + 1).fill = self.TABLE_FILL
        for offset, (label, value) in enumerate(values, 1):
            sheet.cell(start_row + offset, start_column, label).font = Font(bold=True)
            cell = sheet.cell(start_row + offset, start_column + 1, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _write_table(
        self,
        sheet: Any,
        headers: list[str],
        rows: list[list[Any]],
        status_columns: tuple[int, ...] = (),
    ) -> None:
        for column, header in enumerate(headers, 1):
            cell = sheet.cell(1, column, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self.TABLE_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_number, values in enumerate(rows, 2):
            for column, value in enumerate(values, 1):
                cell = sheet.cell(row_number, column, value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            for column in status_columns:
                sheet.cell(row_number, column).fill = self._status_fill(
                    str(sheet.cell(row_number, column).value or "")
                )
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column, header in enumerate(headers, 1):
            width = 14
            if any(word in header for word in ("Path", "Reason", "Message", "Subject")):
                width = 42
            elif any(word in header for word in ("Name", "Email", "CC", "Value")):
                width = 25
            sheet.column_dimensions[get_column_letter(column)].width = width

    def _status_fill(self, value: str) -> PatternFill:
        normalized = value.upper()
        if normalized in {
            "SUCCESS", "SENT", "MOVED", "MOVED_TO_SENT", "PASS",
            "CREATED", "COMPLETED", "VALID",
        }:
            return self.SUCCESS_FILL
        if normalized in {
            "FAILED", "FAIL", "BCC_REJECTED", "MOVE_FAILED", "ANOMALY",
        }:
            return self.FAILED_FILL
        if (
            "WARNING" in normalized
            or normalized in {"COPY_PENDING", "PARTIAL"}
        ):
            return self.WARNING_FILL
        return self.NEUTRAL_FILL

    @staticmethod
    def _report_messages(result: Any) -> list[Any]:
        """Exclude messages outside the selected workflow from the workbook."""
        return [
            item
            for item in result.message_results
            if item.status != "SKIPPED_OTHER_WORKFLOW"
        ]

    @staticmethod
    def _datetime(value: Any) -> str:
        if value is None:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return str(value)

    @staticmethod
    def _revision_date_key(value: Any) -> str | None:
        """Return a canonical DATE IN value for unique-day aggregation."""
        if value is None:
            return None
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")

        value_text = str(value).strip()
        if not value_text:
            return None

        for date_format in (
            "%m/%d/%Y",
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%d-%m-%Y",
            "%Y-%m-%d %H:%M:%S",
        ):
            try:
                return datetime.strptime(value_text, date_format).strftime(
                    "%Y-%m-%d"
                )
            except ValueError:
                continue
        return None
