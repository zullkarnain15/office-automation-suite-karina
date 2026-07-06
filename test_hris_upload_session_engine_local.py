"""
Temporary test file for Sprint 6.13 HRIS Upload Session Engine.

This test uses local mock HRIS login page.
It does not access real HRIS URL.

Run from project root:
py test_hris_upload_session_engine_local.py
"""

from __future__ import annotations

from pathlib import Path

from hris.session_engine import HRISUploadSessionEngine


CONFIG_FILE = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\config\hris\OAS-K_HRIS_Configuration.xlsx"
)

SAMPLE_TXT_FOLDER = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\temp\hris_txt_sample"
)

OUTPUT_ROOT = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\output"
)

MOCK_SITE_FOLDER = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\temp\hris_mock_site"
)

MOCK_LOGIN_FILE = MOCK_SITE_FOLDER / "login.html"

WORKFLOW = "HO"


def prepare_sample_txt_files() -> None:
    """
    Create sample TXT files for session engine test.
    """
    SAMPLE_TXT_FOLDER.mkdir(
        parents=True,
        exist_ok=True,
    )

    sample_files = [
        "Attendance_HO_001.txt",
        "Attendance_HO_002.txt",
        "Attendance_HO_003.txt",
    ]

    for file_name in sample_files:
        file_path = SAMPLE_TXT_FOLDER / file_name

        if not file_path.exists():
            file_path.write_text(
                '"03/30/2026","100808219","03/30/2026","13:20","03/30/2026","16:39"\n',
                encoding="utf-8",
            )


def prepare_mock_hris_page() -> None:
    """
    Create local mock HRIS page.
    """
    MOCK_SITE_FOLDER.mkdir(
        parents=True,
        exist_ok=True,
    )

    html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Mock HRIS Login - OAS-K</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            background: #f5f5f5;
            padding: 40px;
        }

        .card {
            width: 420px;
            margin: auto;
            background: white;
            border: 1px solid #ddd;
            border-radius: 8px;
            padding: 24px;
        }

        h1 {
            color: #1F4E78;
            font-size: 22px;
        }

        label {
            display: block;
            margin-top: 14px;
            font-weight: bold;
        }

        input {
            width: 100%;
            padding: 8px;
            margin-top: 6px;
            box-sizing: border-box;
        }

        button {
            margin-top: 18px;
            width: 100%;
            padding: 10px;
            background: #2E8B57;
            color: white;
            border: none;
            border-radius: 4px;
            font-weight: bold;
            cursor: pointer;
        }

        .success {
            margin-top: 18px;
            color: #2E8B57;
            font-weight: bold;
            display: none;
        }
    </style>
</head>
<body>
    <div class="card">
        <h1>Mock HRIS Login</h1>
        <p>Local mock page for OAS-K Sprint 6.13.</p>

        <label>User ID</label>
        <input id="username" type="text" value="test.user">

        <label>Password</label>
        <input id="password" type="password" value="password">

        <button onclick="document.getElementById('success').style.display='block';">
            Login
        </button>

        <div id="success" class="success">
            Login simulated successfully. Return to terminal and press ENTER.
        </div>
    </div>
</body>
</html>
"""

    MOCK_LOGIN_FILE.write_text(
        html,
        encoding="utf-8",
    )


def patch_config_url_to_local_mock() -> Path:
    """
    Create temporary HRIS configuration copy for local test.

    This does not modify the real HRIS configuration file.
    """
    from shutil import copy2
    from openpyxl import load_workbook

    local_url = MOCK_LOGIN_FILE.resolve().as_uri()

    temp_config_file = Path(
        r"D:\Python Project\OfficeAutomationSuite-Karina\temp\OAS-K_HRIS_Configuration_Local_Test.xlsx"
    )

    temp_config_file.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    copy2(
        CONFIG_FILE,
        temp_config_file,
    )

    workbook = load_workbook(temp_config_file)
    sheet = workbook["General"]

    for row in sheet.iter_rows(min_row=2):
        parameter = row[0].value

        if str(parameter).strip() == "HRIS_URL":
            row[1].value = local_url
            break

    workbook.save(temp_config_file)
    workbook.close()

    print(f"Temporary HRIS_URL set to: {local_url}")
    print(f"Temporary config file     : {temp_config_file}")

    return temp_config_file


def main() -> None:
    print("=" * 80)
    print("OAS-K Sprint 6.13 HRIS Upload Session Engine Local Test")
    print("=" * 80)

    prepare_sample_txt_files()
    prepare_mock_hris_page()
    temp_config_file = patch_config_url_to_local_mock()

    session_engine = HRISUploadSessionEngine(
        configuration_file=temp_config_file,
        txt_folder=SAMPLE_TXT_FOLDER,
        output_root=OUTPUT_ROOT,
        workflow=WORKFLOW,
    )

    result = session_engine.prepare_and_open_login()

    print()
    print("SESSION RESULT")
    print("-" * 80)
    print(f"Success          : {result.success}")
    print(f"Message          : {result.message}")
    print(f"Job ID           : {result.job_id}")
    print(f"Browser Opened   : {result.browser_opened}")
    print(f"Login Confirmed  : {result.login_confirmed}")
    print(f"Report Folder    : {result.report_folder}")

    print()
    print("=" * 80)
    print("HRIS Upload Session Engine local test finished.")
    print("=" * 80)


if __name__ == "__main__":
    main()