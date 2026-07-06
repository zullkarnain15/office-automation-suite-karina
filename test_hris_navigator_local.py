"""
Temporary test file for Sprint 6.14 HRIS Navigator.

This test uses local mock HRIS pages.
It does not access real HRIS URL.

Run from project root:
py test_hris_navigator_local.py
"""

from __future__ import annotations

from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook

from hris.browser import HRISBrowserManager
from hris.navigator import HRISNavigator
from shared.config_manager import HRISConfigurationReader


CONFIG_FILE = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\config\hris\OAS-K_HRIS_Configuration.xlsx"
)

TEMP_CONFIG_FILE = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\temp\OAS-K_HRIS_Configuration_Navigator_Test.xlsx"
)

MOCK_SITE_FOLDER = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\temp\hris_mock_site"
)

MOCK_LOGIN_FILE = MOCK_SITE_FOLDER / "login_navigator.html"


def prepare_mock_hris_site() -> None:
    """
    Create local mock HRIS site for navigator test.
    """
    MOCK_SITE_FOLDER.mkdir(
        parents=True,
        exist_ok=True,
    )

    html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Mock HRIS Navigation - OAS-K</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            background: #f5f5f5;
            padding: 40px;
        }

        .card {
            width: 520px;
            margin: auto;
            background: white;
            border: 1px solid #ddd;
            border-radius: 8px;
            padding: 24px;
        }

        h1 {
            color: #1F4E78;
            font-size: 22px;
        }

        button {
            display: block;
            width: 100%;
            margin-top: 12px;
            padding: 10px;
            background: #1F4E78;
            color: white;
            border: none;
            border-radius: 4px;
            font-weight: bold;
            cursor: pointer;
        }

        .section {
            margin-top: 18px;
            padding: 14px;
            border: 1px solid #ddd;
            display: none;
        }

        .success {
            color: #2E8B57;
            font-weight: bold;
        }
    </style>
</head>
<body>
    <div class="card">
        <h1>Mock HRIS Home</h1>
        <p>Local mock page for OAS-K Sprint 6.14.</p>

        <button onclick="document.getElementById('tao').style.display='block';">
            Time Attendance & Overtime
        </button>

        <div id="tao" class="section">
            <p>Time Attendance & Overtime menu opened.</p>

            <button onclick="document.getElementById('upload').style.display='block';">
                Overtime Upload Attendance
            </button>
        </div>

        <div id="upload" class="section">
            <h2 class="success">Upload Attendance Page</h2>
            <p>Navigator reached upload attendance page successfully.</p>
        </div>
    </div>
</body>
</html>
"""

    MOCK_LOGIN_FILE.write_text(
        html,
        encoding="utf-8",
    )


def prepare_temp_config() -> Path:
    """
    Create temporary config copy with local mock URL.
    """
    TEMP_CONFIG_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    copy2(
        CONFIG_FILE,
        TEMP_CONFIG_FILE,
    )

    local_url = MOCK_LOGIN_FILE.resolve().as_uri()

    workbook = load_workbook(TEMP_CONFIG_FILE)
    sheet = workbook["General"]

    for row in sheet.iter_rows(min_row=2):
        parameter = row[0].value

        if str(parameter).strip() == "HRIS_URL":
            row[1].value = local_url
            break

    workbook.save(TEMP_CONFIG_FILE)
    workbook.close()

    print(f"Temporary HRIS_URL set to: {local_url}")
    print(f"Temporary config file     : {TEMP_CONFIG_FILE}")

    return TEMP_CONFIG_FILE


def main() -> None:
    print("=" * 80)
    print("OAS-K Sprint 6.14 HRIS Navigator Local Test")
    print("=" * 80)

    prepare_mock_hris_site()
    temp_config_file = prepare_temp_config()

    reader = HRISConfigurationReader(temp_config_file)
    configuration = reader.read()

    browser_manager = HRISBrowserManager(
        configuration=configuration,
    )

    try:
        session = browser_manager.open_login_page()

        print()
        print("BROWSER OPENED")
        print("-" * 80)
        print(f"Current URL: {session.page.url}")

        navigator = HRISNavigator(
            page=session.page,
        )

        navigator.open_overtime_upload_attendance()

        is_upload_page_opened = navigator.verify_upload_page_opened()

        print()
        print("NAVIGATION RESULT")
        print("-" * 80)
        print(f"Upload Page Opened: {is_upload_page_opened}")

    finally:
        print()
        print("Closing browser session...")
        browser_manager.close()

    print()
    print("=" * 80)
    print("HRIS Navigator local test finished.")
    print("=" * 80)


if __name__ == "__main__":
    main()