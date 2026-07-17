from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from hris.report_writer import HRISUploadReportWriter


def _artifacts(tmp_path: Path):
    report_folder = tmp_path / "Report"
    report_folder.mkdir()
    return SimpleNamespace(
        job_id="HRIS_HO_20260713_141644",
        workflow="HO",
        job_report_folder=report_folder,
    )


def _upload_plan(tmp_path: Path):
    item = SimpleNamespace(
        sequence=1,
        workflow="HO",
        txt_file_name="Attendance_HO_001.txt",
        txt_file_path=tmp_path / "Attendance_HO_001.txt",
        run_control_id="001",
        run_control_description="HO Upload Slot 1",
        status="PENDING",
        message="",
        verification_status="",
        process_instance="",
    )
    return SimpleNamespace(
        txt_folder=tmp_path,
        total_txt_files=1,
        total_run_controls=10,
        plan_items=[item],
    )


def _summary() -> dict:
    return {
        "status": "COMPLETED",
        "plan_items": [
            {
                "txt_file_name": "Attendance_HO_001.txt",
                "status": "SUCCESS",
                "message": "Process Scheduler OK confirmed.",
                "moved_to": r"C:\output\HRIS\HO\Upload\Attendance_HO_001.txt",
                "verification_status": "SUBMITTED",
                "process_instance": "123456",
            }
        ],
    }


def test_report_includes_and_updates_verification_columns(tmp_path: Path) -> None:
    writer = HRISUploadReportWriter()
    report = writer.write_upload_report(
        _artifacts(tmp_path),
        _upload_plan(tmp_path),
    )

    writer.update_report_after_upload(report, _summary())

    workbook = load_workbook(report, data_only=True)
    sheet = workbook["Upload_Detail"]
    headers = [cell.value for cell in sheet[12]]
    assert headers[:11] == [
        "No",
        "Workflow",
        "TXT File",
        "TXT Path",
        "Run Control ID",
        "Run Control Description",
        "Status",
        "Message",
        "Moved To",
        "Verification",
        "Process Instance",
    ]
    assert sheet["G13"].value == "SUCCESS"
    assert sheet["J13"].value == "SUBMITTED"
    assert sheet["K13"].value == "123456"
    workbook.close()


def test_updater_repairs_legacy_report_without_verification_headers(
    tmp_path: Path,
) -> None:
    writer = HRISUploadReportWriter()
    report = writer.write_upload_report(
        _artifacts(tmp_path),
        _upload_plan(tmp_path),
    )
    workbook = load_workbook(report)
    sheet = workbook["Upload_Detail"]
    sheet["J12"] = None
    sheet["K12"] = None
    workbook.save(report)
    workbook.close()

    writer.update_report_after_upload(report, _summary())

    workbook = load_workbook(report, data_only=True)
    sheet = workbook["Upload_Detail"]
    columns = {
        str(cell.value): cell.column
        for cell in sheet[12]
        if cell.value
    }
    assert sheet.cell(13, columns["Verification"]).value == "SUBMITTED"
    assert sheet.cell(13, columns["Process Instance"]).value == "123456"
    workbook.close()
