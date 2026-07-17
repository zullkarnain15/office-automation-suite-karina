from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from threading import Event

import pytest
from openpyxl import Workbook, load_workbook

from utilities.attachment_consolidation.engine import (
    AttachmentConsolidationEngine,
)
from utilities.attachment_consolidation.job_manager import (
    ConsolidationJobManager,
)
from utilities.attachment_consolidation.models import (
    MODE_EXCEL,
    MODE_TXT,
    ConsolidationRequest,
)
from utilities.attachment_consolidation.scanner import AttachmentScanner
from utilities.attachment_consolidation.statuses import (
    FILE_HIDDEN_SYSTEM,
    FILE_PARTIAL,
    FILE_READY,
    FILE_TEMPORARY,
    FILE_UNSUPPORTED,
)


def _configuration(path: Path, max_lines: int = 10000) -> None:
    workbook = Workbook()
    workbook.active.title = "General"
    for name in (
        "HO_Sender_Master",
        "Branch_Sender_Master",
        "Subject_Rules",
        "Attachment_Rules",
        "Validation_Rules",
        "Reply_Templates",
    ):
        workbook.create_sheet(name)

    workbook["General"].append(["Title"])
    workbook["General"].append([])
    workbook["General"].append([])
    workbook["General"].append(["Parameter", "Value", "Description"])
    workbook["General"].append(["TXT_Max_Lines", max_lines, ""])

    workbook["HO_Sender_Master"].append(
        ["Active", "Sender_Name", "Sender_Email", "Required_CC_Email"]
    )
    workbook["Branch_Sender_Master"].append(
        [
            "Active",
            "Company",
            "Branch_Code",
            "Sender_Name",
            "Sender_Email",
            "Required_CC_Email",
        ]
    )
    workbook["Subject_Rules"].append(
        ["Active", "Workflow", "Subject_Pattern"]
    )
    workbook["Attachment_Rules"].append(
        ["Active", "Workflow", "Allowed_Extensions"]
    )
    workbook["Validation_Rules"].append(
        ["Active", "Rule_Code", "Workflow", "Rule_Value"]
    )
    workbook["Reply_Templates"].append(
        [
            "Active",
            "Reply_Code",
            "Recipient_Type",
            "Trigger",
            "Subject_Template",
            "Body_Template",
        ]
    )
    workbook.save(path)
    workbook.close()


def _excel_attachment(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["NIK", "DATE IN", "TIME IN", "DATE OUT", "TIMEOUT"])
    sheet.append(["001234", "07/01/2026", "08:00", "07/01/2026", "17:00"])
    sheet.append([None, "07/02/2026", "08:00", "07/02/2026", "17:00"])
    workbook.save(path)
    workbook.close()


def test_scanner_is_recursive_and_audits_excel_exclusions(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source"
    nested = source / "nested"
    nested.mkdir(parents=True)
    _excel_attachment(nested / "valid.xlsx")
    (source / "legacy.xls").write_bytes(b"legacy")
    (source / "~$locked.xlsx").write_bytes(b"temporary")
    _excel_attachment(source / ".hidden.xlsx")

    request = ConsolidationRequest(
        mode=MODE_EXCEL,
        workflow="HO",
        source_root=source,
        output_root=tmp_path / "output",
        scan_subfolders=True,
    )
    result = AttachmentScanner().scan(request)
    statuses = {
        item.relative_path: item.status
        for item in result.files
    }

    assert statuses["nested\\valid.xlsx"] == FILE_READY
    assert statuses["legacy.xls"] == FILE_UNSUPPORTED
    assert statuses["~$locked.xlsx"] == FILE_TEMPORARY
    assert statuses[".hidden.xlsx"] == FILE_HIDDEN_SYSTEM


def test_scanner_rejects_output_inside_source(tmp_path: Path) -> None:
    source = tmp_path / "source"
    source.mkdir()
    request = ConsolidationRequest(
        mode=MODE_TXT,
        workflow="Branch",
        source_root=source,
        output_root=source / "output",
    )

    with pytest.raises(ValueError, match="Output Root"):
        AttachmentScanner().scan(request)


def test_excel_consolidation_writes_only_valid_rows_and_audit(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source"
    source.mkdir()
    _excel_attachment(source / "revision.xlsx")
    config = tmp_path / "outlook.xlsx"
    _configuration(config, max_lines=1)
    request = ConsolidationRequest(
        mode=MODE_EXCEL,
        workflow="HO",
        source_root=source,
        output_root=tmp_path / "output",
        configuration_file=config,
    )

    engine = AttachmentConsolidationEngine()
    scan = engine.scan(request)
    result = engine.run(request, scan=scan)

    assert result.success
    assert len(result.records) == 1
    assert len(result.anomalies) == 1
    assert result.file_results[0].status == FILE_PARTIAL
    assert len(result.output_files) == 1
    assert re.fullmatch(
        r"Outlook_Revisi_HO_001_\d{8}_\d{3}\.txt",
        result.output_files[0].name,
    )
    assert result.output_files[0].read_text(encoding="utf-8") == (
        '"07/01/2026","001234","07/01/2026","08:00",'
        '"07/01/2026","17:00"\n'
    )

    workbook = load_workbook(result.artifacts.report_file, read_only=True)
    assert workbook.sheetnames == [
        "Guide_Status",
        "Valid_Records",
        "Invalid_Records",
        "File_Inventory",
        "Process_Summary",
    ]
    workbook.close()
    summary = json.loads(
        result.artifacts.summary_json.read_text(encoding="utf-8")
    )
    assert summary["status"] == "SUCCESS"
    assert summary["valid_records"] == 1
    assert summary["invalid_records"] == 1


def test_txt_consolidation_preserves_nik_splits_and_keeps_duplicates(
    tmp_path: Path,
) -> None:
    source = tmp_path / "source"
    source.mkdir()
    source_file = source / "revision.txt"
    source_file.write_bytes(
        (
            '"07/01/2026","001234","07/01/2026","08:00",'
            '"07/01/2026","17:00"\n'
            '"07/01/2026","001234","07/01/2026","08:00",'
            '"07/01/2026","17:00"\n'
            '"invalid","row"\n'
        ).encode("cp1252")
    )
    config = tmp_path / "outlook.xlsx"
    _configuration(config, max_lines=1)
    request = ConsolidationRequest(
        mode=MODE_TXT,
        workflow="Branch",
        source_root=source,
        output_root=tmp_path / "output",
        configuration_file=config,
    )

    result = AttachmentConsolidationEngine().run(request)

    assert result.success
    assert [record.nik for record in result.records] == ["001234", "001234"]
    assert len(result.output_files) == 2
    assert len(result.anomalies) == 1
    assert all(
        path.read_text(encoding="utf-8").count("\n") == 1
        for path in result.output_files
    )


def test_job_manager_does_not_overwrite_same_second(tmp_path: Path) -> None:
    now = datetime(2026, 7, 17, 10, 30, 45)
    manager = ConsolidationJobManager()

    first = manager.reserve(tmp_path, "HO", now=now)
    second = manager.reserve(tmp_path, "HO", now=now)

    assert first.job_id == "20260717_103045"
    assert second.job_id == "20260717_103045_02"
    assert first.job_folder != second.job_folder


def test_cancelled_process_keeps_audit_artifacts(tmp_path: Path) -> None:
    source = tmp_path / "source"
    source.mkdir()
    _excel_attachment(source / "revision.xlsx")
    config = tmp_path / "outlook.xlsx"
    _configuration(config)
    request = ConsolidationRequest(
        mode=MODE_EXCEL,
        workflow="HO",
        source_root=source,
        output_root=tmp_path / "output",
        configuration_file=config,
    )
    engine = AttachmentConsolidationEngine()
    scan = engine.scan(request)
    cancel = Event()
    cancel.set()

    result = engine.run(request, scan=scan, cancel_event=cancel)

    assert result.cancelled
    assert not result.success
    assert result.artifacts.report_file.is_file()
    assert result.artifacts.summary_json.is_file()
    assert result.file_results[0].status == "CANCELLED"
