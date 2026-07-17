from datetime import date
from pathlib import Path
import json
import re
import shutil
from types import SimpleNamespace

import pytest

from openpyxl import Workbook, load_workbook

import outlook.engine as outlook_engine
from outlook.downloader import OutlookAttachment
from outlook.downloader import OutlookMessage
from outlook.engine import OutlookRevisiEngine
from outlook.parser import AttendanceRevisionRecord
from outlook.parser import OutlookAttachmentParser
from outlook.parser import OutlookDataAnomaly
from outlook.parser import OutlookTxtWriter
from outlook.report_writer import OutlookProcessReportWriter
from shared.config_manager import OutlookRevisiConfigurationReader


def _configuration(path: Path, output_root: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "General"
    for name in (
        "HO_Sender_Master",
        "Branch_Sender_Master",
        "Subject_Rules",
        "Attachment_Rules",
        "Validation_Rules",
        "Reply_Templates",
    ):
        workbook.create_sheet(name)

    workbook["General"].append(["Title"])
    workbook["General"].append([])
    workbook["General"].append([])
    workbook["General"].append(["Parameter", "Value", "Description"])
    workbook["General"].append(["Integration_Method", "OOM_COM", ""])
    workbook["General"].append(["Mailbox_SMTP", "karina.hr.1@oto.co.id", ""])
    workbook["General"].append(["Source_Folder", "Inbox", ""])
    workbook["General"].append(["Reply_From_SMTP", "karina.hr.1@oto.co.id", ""])
    workbook["General"].append(["Output_Root", str(output_root), ""])
    workbook["General"].append(["Payroll_Period", "06-2026", ""])
    workbook["General"].append(["Auto_Reply_Enabled", "TRUE", ""])
    workbook["General"].append(["Send_Mode", "DRAFT", ""])
    workbook["General"].append(["TXT_Max_Lines", "1", ""])

    workbook["HO_Sender_Master"].append(["Title"])
    workbook["HO_Sender_Master"].append([])
    workbook["HO_Sender_Master"].append([])
    workbook["HO_Sender_Master"].append(
        ["Active", "Sender_Name", "Sender_Email", "Required_CC_Email"]
    )
    workbook["HO_Sender_Master"].append(
        ["Y", "HO User", "ho@example.com", "cc@example.com"]
    )

    workbook["Branch_Sender_Master"].append(["Title"])
    workbook["Branch_Sender_Master"].append([])
    workbook["Branch_Sender_Master"].append([])
    workbook["Branch_Sender_Master"].append(
        [
            "Active",
            "Company",
            "Branch_Code",
            "Sender_Name",
            "Sender_Email",
            "Required_CC_Email",
        ]
    )
    workbook["Branch_Sender_Master"].append(
        ["Y", "OTO", "AMU", "Branch User", "branch@example.com", "branchcc@example.com"]
    )

    workbook["Subject_Rules"].append(["Title"])
    workbook["Subject_Rules"].append([])
    workbook["Subject_Rules"].append([])
    workbook["Subject_Rules"].append(["Active", "Workflow", "Subject_Pattern"])
    workbook["Subject_Rules"].append(["Y", "HO", "ATT_REV {PERIOD}"])
    workbook["Subject_Rules"].append(
        ["Y", "Branch", "[{COMPANY}-{BRANCH_CODE}] Attendance {PERIOD}"]
    )

    workbook["Attachment_Rules"].append(["Title"])
    workbook["Attachment_Rules"].append([])
    workbook["Attachment_Rules"].append([])
    workbook["Attachment_Rules"].append(["Active", "Workflow", "Allowed_Extensions"])
    workbook["Attachment_Rules"].append(["Y", "HO", ".xlsx;.xls"])
    workbook["Attachment_Rules"].append(["Y", "Branch", ".txt"])

    workbook["Validation_Rules"].append(["Title"])
    workbook["Validation_Rules"].append([])
    workbook["Validation_Rules"].append([])
    workbook["Validation_Rules"].append(["Active", "Rule_Code", "Workflow", "Rule_Value"])
    workbook["Validation_Rules"].append(["Y", "NIK_REQUIRED", "All", "TRUE"])

    workbook["Reply_Templates"].append(["Title"])
    workbook["Reply_Templates"].append([])
    workbook["Reply_Templates"].append([])
    workbook["Reply_Templates"].append(
        [
            "Active",
            "Reply_Code",
            "Recipient_Type",
            "Trigger",
            "Subject_Template",
            "Body_Template",
        ]
    )
    workbook["Reply_Templates"].append(
        ["Y", "SUCCESS_SENDER", "SENDER", "PROCESS_SUCCESS", "Re: {ORIGINAL_SUBJECT}", "OK"]
    )
    workbook["Reply_Templates"].append(
        ["Y", "FAILED_CC", "SENDER", "CC_INVALID", "Re: {ORIGINAL_SUBJECT}", "CC BAD"]
    )
    workbook["Reply_Templates"].append(
        ["Y", "FAILED_GENERAL", "SENDER", "VALIDATION_FAILED", "Re: {ORIGINAL_SUBJECT}", "{ERROR_REASON}"]
    )

    workbook.save(path)


def _ho_attachment(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["NIK", "DATE IN", "TIME IN", "DATE OUT", "TIMEOUT"])
    sheet.append(["123456789", "06/01/2026", "08:00", "06/01/2026", "17:00"])
    sheet.append(["987654321", "06/02/2026", "09:00", "06/02/2026", "18:00"])
    workbook.save(path)


class FakeClient:
    def __init__(self, messages):
        self.messages = messages
        self.replies = []
        self.moves = []

        self.fetch_count = 0

    def fetch_messages(
        self,
        attachment_folder,
        limit=None,
        message_filter=None,
        attachment_filter=None,
    ):
        self.fetch_count += 1
        candidates = [
            message
            for message in self.messages
            if message_filter is None or message_filter(message)
        ]
        selected = candidates[:limit]
        output_folder = Path(attachment_folder)
        for message in selected:
            if attachment_filter is not None and not attachment_filter(message):
                message.attachments = []
                continue
            copied = []
            for attachment in message.attachments:
                target = output_folder / attachment.file_name
                shutil.copy2(attachment.path, target)
                copied.append(OutlookAttachment(attachment.file_name, target))
            message.attachments = copied
        return selected

    def send_reply(self, **kwargs):
        self.replies.append(kwargs)

    def send_mail(self, **kwargs):
        self.replies.append(kwargs)

    def move_to_folder(self, message, folder_name):
        self.moves.append((message, folder_name))


def test_parse_ho_excel_attachment(tmp_path: Path) -> None:
    attachment = tmp_path / "attendance.xlsx"
    _ho_attachment(attachment)

    result = OutlookAttachmentParser().parse_ho_excel(attachment)

    assert result.valid
    assert len(result.records) == 2
    assert result.records[0].to_txt_row() == [
        "06/01/2026",
        "123456789",
        "06/01/2026",
        "08:00",
        "06/01/2026",
        "17:00",
    ]


def test_parse_ho_ignores_empty_formula_template_rows(tmp_path: Path) -> None:
    attachment = tmp_path / "template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["NIK", "DATE IN", "TIME IN", "DATE OUT", "TIMEOUT"])
    sheet.append(["123456789", "06/01/2026", "08:00", "06/01/2026", "17:00"])
    sheet.append([None, None, None, 0, None])
    sheet["D3"].number_format = "hh:mm"
    workbook.save(attachment)

    result = OutlookAttachmentParser().parse_ho_excel(attachment)

    assert result.valid
    assert len(result.records) == 1


def test_parse_ho_ignores_row_with_only_one_populated_data_cell(
    tmp_path: Path,
) -> None:
    attachment = tmp_path / "one-cell.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["NIK", "DATE IN", "TIME IN", "DATE OUT", "TIMEOUT"])
    sheet.append(["123456789", None, None, None, None])
    workbook.save(attachment)

    result = OutlookAttachmentParser().parse_ho_excel(attachment)

    assert result.valid
    assert result.records == []
    assert result.anomalies == []
    assert result.empty_row_dropped == 1


def test_parse_ho_still_rejects_partially_filled_row_without_nik(
    tmp_path: Path,
) -> None:
    attachment = tmp_path / "partial.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["NIK", "DATE IN", "TIME IN", "DATE OUT", "TIMEOUT"])
    sheet.append([None, "06/01/2026", "08:00", "06/01/2026", "17:00"])
    workbook.save(attachment)

    result = OutlookAttachmentParser().parse_ho_excel(attachment)

    assert not result.valid
    assert result.errors == ["partial.xlsx row 2: NIK is required."]


def test_parse_branch_accepts_windows_1252_non_breaking_space(
    tmp_path: Path,
) -> None:
    attachment = tmp_path / "branch-ansi.txt"
    attachment.write_bytes(
        (
            "06/01/2026,123456789\N{NO-BREAK SPACE},06/01/2026,"
            "08:00,06/01/2026,17:00\n"
        ).encode("cp1252")
    )

    result = OutlookAttachmentParser().parse_branch_txt(attachment)

    assert result.valid
    assert len(result.records) == 1
    assert result.records[0].to_txt_row() == [
        "06/01/2026",
        "123456789",
        "06/01/2026",
        "08:00",
        "06/01/2026",
        "17:00",
    ]


def test_parse_branch_accepts_utf16_tab_delimited_attachment(
    tmp_path: Path,
) -> None:
    attachment = tmp_path / "branch-unicode.txt"
    attachment.write_text(
        "06/01/2026\t123456789\t06/01/2026\t08:00\t"
        "06/01/2026\t17:00\n",
        encoding="utf-16",
    )

    result = OutlookAttachmentParser().parse_branch_txt(attachment)

    assert result.valid
    assert len(result.records) == 1
    assert result.records[0].nik == "123456789"


def test_parse_branch_rejects_bytes_unknown_to_supported_encodings(
    tmp_path: Path,
) -> None:
    attachment = tmp_path / "branch-invalid.txt"
    attachment.write_bytes(b"\x81")

    result = OutlookAttachmentParser().parse_branch_txt(attachment)

    assert not result.valid
    assert result.records == []
    assert result.anomalies[0].code == "FILE_READ_FAILED"
    assert "utf-8-sig" in result.errors[0]
    assert "cp1252" in result.errors[0]


def test_outlook_txt_random_suffix_is_unique_with_collision_retry(
    tmp_path: Path,
    monkeypatch,
) -> None:
    random_values = iter((7, 7, 42))
    monkeypatch.setattr(
        "outlook.parser.secrets.randbelow",
        lambda _limit: next(random_values),
    )
    records = [
        AttendanceRevisionRecord(
            f"{index:09d}",
            date(2026, 7, 1),
            "08:00",
            date(2026, 7, 1),
            "17:00",
            tmp_path / "source.xlsx",
            index + 2,
        )
        for index in range(2)
    ]

    files = OutlookTxtWriter().write(
        records=records,
        output_folder=tmp_path / "TXT",
        workflow="HO",
        max_lines=1,
        job_id="20260710_153045",
    )

    assert [path.name for path in files] == [
        "Outlook_Revisi_HO_001_20260710_007.txt",
        "Outlook_Revisi_HO_002_20260710_042.txt",
    ]


def test_employee_summary_groups_unique_days_and_statuses(
    tmp_path: Path,
) -> None:
    source_a = tmp_path / "source-a.xlsx"
    source_b = tmp_path / "source-b.xlsx"
    valid_records = [
        AttendanceRevisionRecord(
            "123", date(2026, 6, 1), "08:00", date(2026, 6, 1),
            "17:00", source_a, 2,
        ),
        AttendanceRevisionRecord(
            "123", date(2026, 6, 1), "09:00", date(2026, 6, 1),
            "18:00", source_a, 3,
        ),
        AttendanceRevisionRecord(
            "123", date(2026, 6, 2), "08:00", date(2026, 6, 2),
            "17:00", source_b, 2,
        ),
        AttendanceRevisionRecord(
            "456", date(2026, 6, 3), "08:00", date(2026, 6, 3),
            "17:00", source_a, 4,
        ),
    ]
    anomalies = [
        OutlookDataAnomaly(
            source_a, 5, "123", "06/02/2026", "bad", "06/02/2026",
            "17:00", "TIME_FORMAT", "Invalid TIME IN.",
        ),
        OutlookDataAnomaly(
            source_b, 6, "789", "not-a-date", "08:00", "06/03/2026",
            "17:00", "DATE_FORMAT", "Invalid DATE IN.",
        ),
        OutlookDataAnomaly(
            source_b, 7, "", "06/04/2026", "08:00", "06/04/2026",
            "17:00", "NIK_REQUIRED", "NIK is required.",
        ),
    ]
    result = SimpleNamespace(
        message_results=[
            SimpleNamespace(
                status="SUCCESS",
                valid_records=valid_records,
                anomalies=anomalies,
            )
        ]
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary_Per_Karyawan"

    OutlookProcessReportWriter()._write_employee_summary(sheet, result)

    assert list(sheet.values) == [
        (
            "NIK", "Total Hari Revisi", "Valid For TXT", "Anomaly",
            "Total File Sumber", "Status",
        ),
        ("(EMPTY)", 1, 0, 1, 1, "ANOMALY"),
        ("123", 2, 3, 1, 2, "PARTIAL"),
        ("456", 1, 1, 0, 1, "VALID"),
        ("789", 0, 0, 1, 1, "ANOMALY"),
    ]
    assert sheet["A2"].number_format == "@"
    assert sheet["B2"].number_format == "#,##0"
    assert sheet["F2"].fill == OutlookProcessReportWriter.FAILED_FILL
    assert sheet["F3"].fill == OutlookProcessReportWriter.WARNING_FILL
    assert sheet["F4"].fill == OutlookProcessReportWriter.SUCCESS_FILL
    workbook.close()


def test_engine_writes_split_txt_in_dry_run(tmp_path: Path) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    attachment = tmp_path / "attendance.xlsx"
    _configuration(config_path, output_root)
    _ho_attachment(attachment)
    message = OutlookMessage(
        entry_id="entry-1",
        store_id="store",
        subject="ATT_REV 06-2026",
        sender_name="HO User",
        sender_email="ho@example.com",
        cc="cc@example.com",
        received_time=None,
        attachments=[
            OutlookAttachment(
                file_name=attachment.name,
                path=attachment,
            )
        ],
    )
    client = FakeClient([message])

    result = OutlookRevisiEngine(
        configuration_file=config_path,
        workflow="HO",
        dry_run=True,
        client=client,
    ).run()

    assert result.success
    assert result.success_email == 1
    assert result.output_txt_count == 2
    assert all(path.exists() for path in result.message_results[0].output_files)
    month_label = f"{result.job_id[:4]}-{result.job_id[4:6]}"
    assert result.output_folder.parent == output_root / "HO" / month_label
    assert re.fullmatch(r"\d{8}_\d{6}", result.output_folder.name)
    assert {path.name for path in result.output_folder.iterdir()} == {
        "TXT", "Report", "Process.log", "summary.json"
    }
    assert not (result.output_folder / "Attachments").exists()
    assert result.report_status == "CREATED"
    assert result.report_file is not None and result.report_file.exists()
    report = load_workbook(result.report_file, data_only=True)
    assert report.sheetnames == [
        "Dashboard", "Email_Result", "Attachment_Result", "Valid_Data",
        "Summary_Per_Karyawan", "Data_Anomaly", "Output_Summary",
    ]
    assert report["Valid_Data"].max_row == 3
    assert report["Valid_Data"]["E2"].value == "123456789"
    assert list(report["Summary_Per_Karyawan"].values) == [
        (
            "NIK", "Total Hari Revisi", "Valid For TXT", "Anomaly",
            "Total File Sumber", "Status",
        ),
        ("123456789", 1, 1, 0, 1, "VALID"),
        ("987654321", 1, 1, 0, 1, "VALID"),
    ]
    assert report["Email_Result"]["Q2"].value == "SUCCESS"
    assert report["Email_Result"]["J2"].value == "ATT_REV 06-2026"
    for sheet_name in (
        "Email_Result", "Attachment_Result", "Valid_Data", "Data_Anomaly"
    ):
        headers = [cell.value for cell in report[sheet_name][1]]
        assert "Email_ID" not in headers
    assert report["Dashboard"]["B8"].value == "06-2026"
    assert report["Dashboard"]["B9"].value == "karina.hr.1@oto.co.id"
    output_types = {
        row[2].value for row in report["Output_Summary"].iter_rows(min_row=2)
    }
    assert {"HRIS_TXT", "EXCEL_REPORT", "PROCESS_LOG", "SUMMARY_JSON"} <= output_types
    report.close()
    output_names = [
        path.name for path in result.message_results[0].output_files
    ]
    assert re.fullmatch(
        rf"Outlook_Revisi_HO_001_{result.output_folder.name[:8]}_\d{{3}}\.txt",
        output_names[0],
    )
    assert re.fullmatch(
        rf"Outlook_Revisi_HO_002_{result.output_folder.name[:8]}_\d{{3}}\.txt",
        output_names[1],
    )
    assert output_names[0].rsplit("_", 1)[-1] != output_names[1].rsplit("_", 1)[-1]
    assert not (output_root / "Branch").exists()
    assert client.replies == []
    history = OutlookRevisiEngine._history_path(output_root, "HO")
    assert "entry-1" not in history.read_text(encoding="utf-8")


def test_job_folder_reservation_uses_timestamp_and_safe_collision_suffix(
    tmp_path: Path,
    monkeypatch,
) -> None:
    class _FixedDateTime:
        @staticmethod
        def now():
            class _FixedNow:
                @staticmethod
                def strftime(_format: str) -> str:
                    return "20260706_092245"

            return _FixedNow()

    monkeypatch.setattr(outlook_engine, "datetime", _FixedDateTime)
    _, first_folder = OutlookRevisiEngine._reserve_job_folder(tmp_path, "Branch")
    _, second_folder = OutlookRevisiEngine._reserve_job_folder(tmp_path, "Branch")

    assert first_folder != second_folder
    assert first_folder.name == "20260706_092245"
    assert second_folder.name == "20260706_092245_02"
    assert first_folder.parent == tmp_path / "Branch" / "2026-07"
    assert second_folder.parent == first_folder.parent
    assert first_folder.is_dir()
    assert second_folder.is_dir()
    assert not (tmp_path / "HO").exists()


def test_branch_run_creates_only_branch_job_folder(tmp_path: Path) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    _configuration(config_path, output_root)
    client = FakeClient([])

    result = OutlookRevisiEngine(
        configuration_file=config_path,
        workflow="Branch",
        dry_run=True,
        client=client,
    ).run()

    month_label = f"{result.job_id[:4]}-{result.job_id[4:6]}"
    assert result.output_folder.parent == output_root / "Branch" / month_label
    assert {path.name for path in result.output_folder.iterdir()} == {
        "TXT", "Report", "Process.log", "summary.json"
    }
    assert not (result.output_folder / "Attachments").exists()
    assert result.report_file is not None
    assert result.report_file.name == (
        f"Outlook_Process_Report_Branch_{result.job_id}.xlsx"
    )
    assert result.report_file.exists()
    assert not (output_root / "HO").exists()


def test_same_sender_can_process_two_different_branch_configurations(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    _configuration(config_path, output_root)
    workbook = load_workbook(config_path)
    workbook["Branch_Sender_Master"].append(
        [
            "Y",
            "OTO",
            "BKS",
            "Branch User",
            "branch@example.com",
            "bkscc@example.com",
        ]
    )
    workbook.save(config_path)
    workbook.close()

    messages: list[OutlookMessage] = []
    for entry_id, branch_code, required_cc in (
        ("branch-amu", "AMU", "branchcc@example.com"),
        ("branch-bks", "BKS", "bkscc@example.com"),
    ):
        attachment = tmp_path / f"{branch_code}.txt"
        attachment.write_text(
            "06/01/2026,123456789,06/01/2026,08:00,"
            "06/01/2026,17:00\n",
            encoding="utf-8",
        )
        messages.append(
            OutlookMessage(
                entry_id=entry_id,
                store_id="store",
                subject=f"[OTO-{branch_code}] Attendance 06-2026",
                sender_name="Branch User",
                sender_email="branch@example.com",
                cc=required_cc,
                received_time=None,
                attachments=[OutlookAttachment(attachment.name, attachment)],
            )
        )

    result = OutlookRevisiEngine(
        configuration_file=config_path,
        workflow="Branch",
        dry_run=True,
        client=FakeClient(messages),
    ).run()

    assert result.success_email == 2
    assert result.failed_email == 0
    assert result.valid_row_count == 2
    assert [item.required_cc for item in result.message_results] == [
        "branchcc@example.com",
        "bkscc@example.com",
    ]
    assert [item.expected_subject for item in result.message_results] == [
        "[OTO-AMU] Attendance 06-2026",
        "[OTO-BKS] Attendance 06-2026",
    ]
    assert all(
        item.validation_subject == "PASS"
        and item.validation_cc == "PASS"
        for item in result.message_results
    )


def test_other_workflow_is_skipped_without_saving_attachment(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    attachment = tmp_path / "branch.txt"
    attachment.write_text(
        '06/01/2026,123456789,06/01/2026,08:00,06/01/2026,17:00\n',
        encoding="utf-8",
    )
    _configuration(config_path, output_root)
    message = OutlookMessage(
        entry_id="branch-entry",
        store_id="store",
        subject="[OTO-AMU] Attendance 06-2026",
        sender_name="Branch User",
        sender_email="branch@example.com",
        cc="branchcc@example.com",
        received_time=None,
        attachments=[OutlookAttachment(attachment.name, attachment)],
    )
    client = FakeClient([message])

    result = OutlookRevisiEngine(
        configuration_file=config_path,
        workflow="HO",
        dry_run=True,
        client=client,
    ).run()

    assert result.total_email == 1
    assert result.target_email == 0
    assert result.skipped_other_workflow == 1
    assert result.failed_email == 0
    assert result.message_results[0].status == "SKIPPED_OTHER_WORKFLOW"
    assert not (result.output_folder / "Attachments").exists()
    assert list((result.output_folder / "TXT").iterdir()) == []
    assert client.replies == []
    assert result.report_file is not None
    report = load_workbook(result.report_file, data_only=True)
    assert report["Email_Result"].max_row == 1
    dashboard_values = {
        cell.value
        for row in report["Dashboard"].iter_rows()
        for cell in row
    }
    assert "Email_Skipped_Other_Workflow" not in dashboard_values
    assert "SKIPPED" not in dashboard_values
    report.close()


def test_message_limit_counts_workflow_candidates_not_unrelated_email(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    attachment = tmp_path / "attendance.xlsx"
    _configuration(config_path, output_root)
    _ho_attachment(attachment)
    unrelated = OutlookMessage(
        entry_id="unrelated",
        store_id="store",
        subject="Shokz support notification",
        sender_name="Shokz",
        sender_email="hello@shokz.com",
        cc="",
        received_time=None,
        attachments=[],
    )
    target = OutlookMessage(
        entry_id="target",
        store_id="store",
        subject="ATT_REV 06-2026",
        sender_name="HO User",
        sender_email="ho@example.com",
        cc="cc@example.com",
        received_time=None,
        attachments=[OutlookAttachment(attachment.name, attachment)],
    )

    result = OutlookRevisiEngine(
        configuration_file=config_path,
        workflow="HO",
        dry_run=True,
        message_limit=1,
        client=FakeClient([unrelated, target]),
    ).run()

    assert result.total_email == 1
    assert result.success_email == 1
    assert result.message_results[0].entry_id == "target"


def test_subject_rule_does_not_classify_reply_as_new_submission(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    _configuration(config_path, output_root)
    configuration = OutlookRevisiConfigurationReader(config_path).read()
    engine = OutlookRevisiEngine(config_path, workflow="HO", dry_run=True)
    reply = OutlookMessage(
        entry_id="reply",
        store_id="store",
        subject="Re: ATT_REV 06-2026",
        sender_name="Mailbox",
        sender_email="mailbox@example.com",
        cc="",
        received_time=None,
        attachments=[],
    )

    assert engine._detect_message_workflow(configuration, reply) == ""


def test_invalid_workflow_does_not_create_output_or_scan_inbox(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    _configuration(config_path, output_root)
    client = FakeClient([])

    with pytest.raises(ValueError, match="workflow must be"):
        OutlookRevisiEngine(
            configuration_file=config_path,
            workflow="",
            dry_run=True,
            client=client,
        )

    assert client.fetch_count == 0
    assert not output_root.exists()


def test_missing_cc_error_reports_required_and_resolved_values(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    attachment = tmp_path / "attendance.xlsx"
    _configuration(config_path, output_root)
    _ho_attachment(attachment)
    message = OutlookMessage(
        entry_id="missing-cc",
        store_id="store",
        subject="ATT_REV 06-2026",
        sender_name="HO User",
        sender_email="ho@example.com",
        cc="",
        received_time=None,
        attachments=[OutlookAttachment(attachment.name, attachment)],
    )

    result = OutlookRevisiEngine(
        configuration_file=config_path,
        workflow="HO",
        dry_run=True,
        client=FakeClient([message]),
    ).run()

    assert result.message_results[0].errors == [
        "Required CC email is missing: cc@example.com. "
        "Actual CC resolved: (none)"
    ]
    report = load_workbook(result.report_file, data_only=True)
    assert report["Email_Result"]["M2"].value == "FAIL"
    assert report["Email_Result"]["Q2"].value == "FAILED"
    assert report["Email_Result"]["R2"].value == "CC_INVALID"
    assert report["Attachment_Result"].max_row == 1
    report.close()


def test_draft_reply_does_not_move_or_mark_message_processed(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    attachment = tmp_path / "attendance.xlsx"
    _configuration(config_path, output_root)
    _ho_attachment(attachment)
    message = OutlookMessage(
        entry_id="draft-entry",
        store_id="store",
        subject="ATT_REV 06-2026",
        sender_name="HO User",
        sender_email="ho@example.com",
        cc="cc@example.com",
        received_time=None,
        attachments=[OutlookAttachment(attachment.name, attachment)],
    )
    client = FakeClient([message])

    result = OutlookRevisiEngine(
        configuration_file=config_path,
        workflow="HO",
        dry_run=False,
        client=client,
    ).run()

    assert result.success_email == 1
    assert result.message_results[0].reply_sent is False
    assert len(client.replies) == 1
    assert client.replies[0]["send_mode"] == "DRAFT"
    assert client.moves == []
    history = OutlookRevisiEngine._history_path(output_root, "HO")
    assert "draft-entry" not in history.read_text(encoding="utf-8")


def test_failed_subject_reply_can_render_expected_subject(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    attachment = tmp_path / "attendance.xlsx"
    _configuration(config_path, output_root)
    _ho_attachment(attachment)
    workbook = load_workbook(config_path)
    workbook["General"]["B12"] = "SEND"
    workbook["Reply_Templates"]["F7"] = (
        "Received: {ORIGINAL_SUBJECT}\nExpected: {EXPECTED_SUBJECT}"
    )
    workbook.save(config_path)
    workbook.close()
    message = OutlookMessage(
        entry_id="wrong-subject",
        store_id="store",
        subject="WRONG SUBJECT",
        sender_name="HO User",
        sender_email="ho@example.com",
        cc="cc@example.com",
        received_time=None,
        attachments=[OutlookAttachment(attachment.name, attachment)],
    )
    client = FakeClient([message])

    result = OutlookRevisiEngine(
        configuration_file=config_path,
        workflow="HO",
        dry_run=False,
        client=client,
    ).run()

    message_result = result.message_results[0]
    assert message_result.failure_code == "SUBJECT_INVALID"
    assert message_result.expected_subject == "ATT_REV 06-2026"
    assert len(client.replies) == 1
    assert client.replies[0]["body"] == (
        "Received: WRONG SUBJECT\nExpected: ATT_REV 06-2026"
    )
    assert client.moves == [(message, "Deleted Items")]
    assert message_result.move_result == "MOVED"
    history = OutlookRevisiEngine._history_path(output_root, "HO")
    assert "wrong-subject" in history.read_text(encoding="utf-8")


def test_invalid_data_is_written_as_structured_anomaly(tmp_path: Path) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    attachment = tmp_path / "invalid.xlsx"
    _configuration(config_path, output_root)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["NIK", "DATE IN", "TIME IN", "DATE OUT", "TIMEOUT"])
    sheet.append([None, "06/01/2026", "08:00", "06/01/2026", "17:00"])
    workbook.save(attachment)
    message = OutlookMessage(
        entry_id="invalid-data",
        store_id="store",
        subject="ATT_REV 06-2026",
        sender_name="HO User",
        sender_email="ho@example.com",
        cc="cc@example.com",
        received_time=None,
        attachments=[OutlookAttachment(attachment.name, attachment)],
    )

    result = OutlookRevisiEngine(
        config_path, "HO", dry_run=True, client=FakeClient([message])
    ).run()

    assert result.failed_email == 1
    assert result.valid_row_count == 0
    assert result.anomaly_row_count == 1
    assert result.message_results[0].failure_code == "NIK_REQUIRED"
    retained_attachments = result.output_folder / "Attachments"
    assert retained_attachments.is_dir()
    assert [path.name for path in retained_attachments.iterdir()] == [
        "invalid.xlsx"
    ]
    report = load_workbook(result.report_file, data_only=True)
    assert report["Data_Anomaly"]["J2"].value == "NIK_REQUIRED"
    assert report["Valid_Data"].max_row == 1
    report.close()


def test_massive_empty_rows_are_aggregated_not_expanded(tmp_path: Path) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    attachment = tmp_path / "branch.txt"
    _configuration(config_path, output_root)
    attachment.write_text(
        ('"","","","","",""\n' * 2000)
        + '"06/01/2026","001234567","06/01/2026","08:00","06/01/2026","17:00"\n',
        encoding="utf-8",
    )
    message = OutlookMessage(
        entry_id="empty-rows",
        store_id="store",
        subject="[OTO-AMU] Attendance 06-2026",
        sender_name="Branch User",
        sender_email="branch@example.com",
        cc="branchcc@example.com",
        received_time=None,
        attachments=[OutlookAttachment(attachment.name, attachment)],
    )

    result = OutlookRevisiEngine(
        config_path, "Branch", dry_run=True, client=FakeClient([message])
    ).run()

    assert result.success_email == 1
    assert result.anomaly_row_count == 2000
    report = load_workbook(result.report_file, data_only=True)
    assert report["Attachment_Result"]["L2"].value == 2000
    assert report["Data_Anomaly"].max_row == 1
    assert report["Valid_Data"].max_row == 2
    report.close()


def test_branch_exports_valid_rows_when_other_rows_are_invalid(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    attachment = tmp_path / "Cabang Sampit.txt"
    _configuration(config_path, output_root)
    attachment.write_text(
        "\n".join(
            [
                '"06/01/2026","001234567","06/01/2026","08:00","06/01/2026","17:00"',
                '"06/02/2026","001234568","06/02/2026","","06/02/2026","17:00"',
                '"06/03/2026","001234569","06/03/2026","09:00","06/03/2026","18:00"',
            ]
        ),
        encoding="utf-8",
    )
    message = OutlookMessage(
        entry_id="branch-partial",
        store_id="store",
        subject="[OTO-AMU] Attendance 06-2026",
        sender_name="Branch User",
        sender_email="branch@example.com",
        cc="branchcc@example.com",
        received_time=None,
        attachments=[OutlookAttachment(attachment.name, attachment)],
    )

    result = OutlookRevisiEngine(
        config_path,
        "Branch",
        dry_run=True,
        client=FakeClient([message]),
    ).run()

    message_result = result.message_results[0]
    assert result.success_email == 1
    assert result.failed_email == 0
    assert result.valid_row_count == 2
    assert result.anomaly_row_count == 1
    assert message_result.status == "SUCCESS"
    assert message_result.validation_data == "WARNING"
    assert message_result.failure_code == "DATA_ANOMALY"
    assert len(message_result.output_files) == 2
    assert len(message_result.valid_records) == 2
    assert message_result.attachment_results[0].file_status == "WARNING"
    assert "row 2: TIME IN is invalid" in message_result.errors[0]
    assert sum(
        len(path.read_text(encoding="utf-8").splitlines())
        for path in message_result.output_files
    ) == 2

    report = load_workbook(result.report_file, data_only=True)
    assert report["Email_Result"]["P2"].value == "WARNING"
    assert report["Email_Result"]["Q2"].value == "SUCCESS"
    assert report["Attachment_Result"]["G2"].value == "WARNING"
    assert report["Valid_Data"].max_row == 3
    assert report["Data_Anomaly"].max_row == 2
    report.close()


def test_ho_exports_valid_rows_when_other_excel_rows_are_invalid(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    attachment = tmp_path / "attendance-partial.xlsx"
    _configuration(config_path, output_root)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["NIK", "DATE IN", "TIME IN", "DATE OUT", "TIMEOUT"])
    sheet.append(["123456789", "06/01/2026", "08:00", "06/01/2026", "17:00"])
    sheet.append(["123456790", "06/02/2026", None, "06/02/2026", "17:00"])
    sheet.append(["123456791", "06/03/2026", "09:00", "06/03/2026", "18:00"])
    workbook.save(attachment)
    message = OutlookMessage(
        entry_id="ho-partial",
        store_id="store",
        subject="ATT_REV 06-2026",
        sender_name="HO User",
        sender_email="ho@example.com",
        cc="cc@example.com",
        received_time=None,
        attachments=[OutlookAttachment(attachment.name, attachment)],
    )

    result = OutlookRevisiEngine(
        config_path,
        "HO",
        dry_run=True,
        client=FakeClient([message]),
    ).run()

    message_result = result.message_results[0]
    assert result.success_email == 1
    assert result.failed_email == 0
    assert result.valid_row_count == 2
    assert result.anomaly_row_count == 1
    assert message_result.status == "SUCCESS"
    assert message_result.validation_data == "WARNING"
    assert message_result.failure_code == "DATA_ANOMALY"
    assert len(message_result.valid_records) == 2
    assert message_result.attachment_results[0].file_status == "WARNING"
    assert "row 3: TIME IN is invalid" in message_result.errors[0]
    assert sum(
        len(path.read_text(encoding="utf-8").splitlines())
        for path in message_result.output_files
    ) == 2

    report = load_workbook(result.report_file, data_only=True)
    assert report["Email_Result"]["P2"].value == "WARNING"
    assert report["Email_Result"]["Q2"].value == "SUCCESS"
    assert report["Attachment_Result"]["G2"].value == "WARNING"
    assert report["Valid_Data"].max_row == 3
    assert report["Data_Anomaly"].max_row == 2
    report.close()


def test_multiple_branch_emails_share_txt_until_split_limit(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    _configuration(config_path, output_root)
    workbook = load_workbook(config_path)
    workbook["General"]["B13"] = 2
    workbook.save(config_path)
    workbook.close()

    messages = []
    for number in range(1, 4):
        attachment = tmp_path / f"branch-{number}.txt"
        attachment.write_text(
            (
                f'"06/0{number}/2026","00123456{number}",'
                f'"06/0{number}/2026","08:00",'
                f'"06/0{number}/2026","17:00"\n'
            ),
            encoding="utf-8",
        )
        messages.append(
            OutlookMessage(
                entry_id=f"branch-{number}",
                store_id="store",
                subject="[OTO-AMU] Attendance 06-2026",
                sender_name="Branch User",
                sender_email="branch@example.com",
                cc="branchcc@example.com",
                received_time=None,
                attachments=[OutlookAttachment(attachment.name, attachment)],
            )
        )

    result = OutlookRevisiEngine(
        config_path,
        "Branch",
        dry_run=True,
        client=FakeClient(messages),
    ).run()

    output_files = [
        path
        for message_result in result.message_results
        for path in message_result.output_files
    ]
    assert result.success_email == 3
    assert result.valid_row_count == 3
    assert len(output_files) == 3
    output_names = [path.name for path in output_files]
    assert output_names[0] == output_names[1]
    assert re.fullmatch(
        rf"Outlook_Revisi_Branch_001_{result.job_id[:8]}_\d{{3}}\.txt",
        output_names[0],
    )
    assert re.fullmatch(
        rf"Outlook_Revisi_Branch_002_{result.job_id[:8]}_\d{{3}}\.txt",
        output_names[2],
    )
    assert output_names[0].rsplit("_", 1)[-1] != output_names[2].rsplit("_", 1)[-1]
    unique_output_files = sorted(
        {path.resolve() for path in output_files},
        key=str,
    )
    assert result.output_txt_count == 2
    assert len(unique_output_files) == 2
    assert all(path.exists() for path in output_files)
    assert [
        len(path.read_text(encoding="utf-8").splitlines())
        for path in unique_output_files
    ] == [2, 1]

    report = load_workbook(result.report_file, data_only=True)
    output_rows = [
        row
        for row in report["Output_Summary"].iter_rows(
            min_row=2,
            values_only=True,
        )
        if row[2] == "HRIS_TXT"
    ]
    assert len(output_rows) == 2
    assert sum(row[5] for row in output_rows) == 3
    report.close()


def test_report_failure_keeps_txt_and_updates_summary(
    tmp_path: Path,
    monkeypatch,
) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    attachment = tmp_path / "attendance.xlsx"
    _configuration(config_path, output_root)
    _ho_attachment(attachment)
    message = OutlookMessage(
        entry_id="report-failure",
        store_id="store",
        subject="ATT_REV 06-2026",
        sender_name="HO User",
        sender_email="ho@example.com",
        cc="cc@example.com",
        received_time=None,
        attachments=[OutlookAttachment(attachment.name, attachment)],
    )
    monkeypatch.setattr(
        outlook_engine.OutlookProcessReportWriter,
        "write",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("write denied")),
    )

    result = OutlookRevisiEngine(
        config_path, "HO", dry_run=True, client=FakeClient([message])
    ).run()

    assert result.report_status == "FAILED"
    assert result.final_status == "COMPLETED WITH WARNING"
    assert all(path.exists() for path in result.message_results[0].output_files)
    assert (
        result.output_folder / "Attachments" / "attendance.xlsx"
    ).exists()
    summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
    assert summary["report_status"] == "FAILED"
    assert "Report generation failed" in result.process_log.read_text(encoding="utf-8")


def test_reconciliation_mismatch_marks_warning(tmp_path: Path) -> None:
    config_path = tmp_path / "config.xlsx"
    output_root = tmp_path / "output" / "Outlook-Revisi"
    attachment = tmp_path / "attendance.xlsx"
    _configuration(config_path, output_root)
    _ho_attachment(attachment)
    message = OutlookMessage(
        entry_id="reconcile",
        store_id="store",
        subject="ATT_REV 06-2026",
        sender_name="HO User",
        sender_email="ho@example.com",
        cc="cc@example.com",
        received_time=None,
        attachments=[OutlookAttachment(attachment.name, attachment)],
    )
    engine = OutlookRevisiEngine(
        config_path, "HO", dry_run=True, client=FakeClient([message])
    )
    result = engine.run()
    result.message_results[0].valid_records[0].output_file = None

    engine._reconcile(result)

    assert result.reconciliation_status == "WARNING"
    assert result.final_status == "COMPLETED WITH WARNING"
    assert result.reconciliation_issues == [
        "Valid row count does not match exported record count."
    ]
