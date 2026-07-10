from pathlib import Path

from openpyxl import Workbook

from shared.config_manager import HRISConfigurationReader


def _base_workbook(path: Path, include_assisted: bool = True) -> None:
    workbook = Workbook()
    workbook.active.title = "General"
    for name in ("Run_Control", "Browser", "Upload", "Reference"):
        workbook.create_sheet(name)
    for name in ("General", "Browser", "Upload", "Reference"):
        workbook[name].append(["Parameter", "Value", "Description"])
    workbook["Run_Control"].append(
        ["Active", "Sequence", "Workflow", "Run_Control_ID", "Description"]
    )
    workbook["Run_Control"].append(["Y", 1, "HO", "001", "Test"])
    if include_assisted:
        sheet = workbook.create_sheet("Assisted_Steps")
        sheet.append([
            "Active", "Sequence", "Step_Name", "Action", "Input_Source",
            "Method", "Required", "Wait_After_Seconds", "Description",
        ])
        sheet.append(["Y", 2, "second", "click", "NONE", "coordinate", True, 1, ""])
        sheet.append(["N", 1, "inactive", "click", "NONE", "manual", True, 1, ""])
        sheet.append(["Y", 1, "first", "type", "START_DATE", "manual", True, .5, ""])
    workbook.save(path)


def test_reads_active_assisted_steps_sorted(tmp_path: Path) -> None:
    path = tmp_path / "config.xlsx"
    _base_workbook(path)
    configuration = HRISConfigurationReader(path).read()
    assert [step.step_name for step in configuration.assisted_steps] == [
        "first", "second"
    ]


def test_missing_sheet_uses_defaults(tmp_path: Path) -> None:
    path = tmp_path / "config.xlsx"
    _base_workbook(path, include_assisted=False)
    configuration = HRISConfigurationReader(path).read()
    assert len(configuration.assisted_steps) == 9
