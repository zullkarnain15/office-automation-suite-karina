import json
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

import attendance.extractor as extractor_module
from attendance.engine import AttendanceProcessEngine
from attendance.engine import AttendanceTXTRecordDeduplicator
from attendance.engine import AttendanceValidationEngine


def _valid_record(
    source_mdb: str,
    check_in: datetime,
    check_out: datetime,
) -> dict[str, object]:
    return {
        "nik": "123456789",
        "name": "Karina",
        "attendance_date": check_in.date(),
        "check_in": check_in,
        "check_out": check_out,
        "tap_count": 2,
        "pair_status": "PAIRED",
        "is_valid_for_txt": True,
        "validation_status": "VALID_FOR_TXT",
        "validation_remarks": "",
        "source_mdb": source_mdb,
        "source_mdb_path": str(Path("C:/attendance") / source_mdb),
    }


def test_deduplicator_uses_exact_final_txt_fields() -> None:
    retained = _valid_record(
        "machine-a.mdb",
        datetime(2026, 7, 10, 8, 0, 5),
        datetime(2026, 7, 10, 17, 0, 5),
    )
    same_txt_minutes = _valid_record(
        "machine-b.mdb",
        datetime(2026, 7, 10, 8, 0, 45),
        datetime(2026, 7, 10, 17, 0, 45),
    )
    different_time_out = _valid_record(
        "machine-c.mdb",
        datetime(2026, 7, 10, 8, 0, 5),
        datetime(2026, 7, 10, 17, 1, 5),
    )
    records = [retained, same_txt_minutes, different_time_out]

    result = AttendanceTXTRecordDeduplicator().deduplicate(records)

    assert result["unique_records"] == [retained, different_time_out]
    assert result["duplicate_records"] == [same_txt_minutes]
    assert result["summary"] == {
        "input_valid_records": 3,
        "unique_txt_records": 2,
        "duplicate_removed_records": 1,
    }
    assert same_txt_minutes["is_valid_for_txt"] is False
    assert same_txt_minutes["validation_status"] == "DUPLICATE_REMOVED"
    assert "machine-a.mdb" in str(
        same_txt_minutes["validation_remarks"]
    )

    summary = AttendanceValidationEngine().summarize_validation(records)
    assert summary["valid_for_txt_records"] == 2
    assert summary["anomaly_records"] == 1
    assert summary["duplicate_removed_records"] == 1


def test_process_removes_duplicate_txt_row_and_keeps_report_audit(
    tmp_path: Path,
    monkeypatch,
) -> None:
    source_times = {
        "machine-a.mdb": ((8, 0, 5), (17, 0, 5)),
        "machine-b.mdb": ((8, 0, 45), (17, 0, 45)),
        "machine-c.mdb": ((8, 0, 5), (17, 1, 5)),
    }

    class FakeAttendanceMDBExtractor:
        def __init__(self, mdb_path: str | Path) -> None:
            self.mdb_path = Path(mdb_path)

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc_value, traceback) -> None:
            return None

        def fetch_raw_logs(
            self,
            date_from: datetime,
            date_to: datetime,
        ) -> list[dict[str, object]]:
            del date_from, date_to
            check_in, check_out = source_times[self.mdb_path.name]
            return [
                {
                    "nik": "123456789",
                    "name": "Karina",
                    "checktime": datetime(2026, 7, 10, *check_in),
                    "source_mdb": self.mdb_path.name,
                    "source_mdb_path": str(self.mdb_path),
                },
                {
                    "nik": "123456789",
                    "name": "Karina",
                    "checktime": datetime(2026, 7, 10, *check_out),
                    "source_mdb": self.mdb_path.name,
                    "source_mdb_path": str(self.mdb_path),
                },
            ]

    monkeypatch.setattr(
        extractor_module,
        "AttendanceMDBExtractor",
        FakeAttendanceMDBExtractor,
    )

    mdb_list = [
        SimpleNamespace(
            code=f"MDB-{index}",
            description=path.name,
            mdb_path=path,
        )
        for index, path in enumerate(
            (
                tmp_path / "machine-a.mdb",
                tmp_path / "machine-b.mdb",
                tmp_path / "machine-c.mdb",
            ),
            start=1,
        )
    ]
    configuration = SimpleNamespace(
        general={"Split_TXT_Rows": 10000},
        ho_mdb_list=mdb_list,
        branch_mdb_list=[],
    )

    result = AttendanceProcessEngine().run(
        configuration=configuration,
        output_root=tmp_path,
        workflow="HO",
        date_from=datetime(2026, 7, 10),
        date_to=datetime(2026, 7, 10),
    )

    assert result["paired_record_count"] == 3
    assert result["valid_record_count"] == 2
    assert result["duplicate_removed_count"] == 1
    assert result["anomaly_record_count"] == 1
    assert result["validation_summary"]["duplicate_removed_records"] == 1

    txt_path = Path(result["txt_result"]["generated_files"][0]["file_path"])
    assert txt_path.read_text(encoding="utf-8").splitlines() == [
        '"07/10/2026","123456789","07/10/2026","08:00",'
        '"07/10/2026","17:00"',
        '"07/10/2026","123456789","07/10/2026","08:00",'
        '"07/10/2026","17:01"',
    ]

    report_path = Path(result["report_result"]["report_file"])
    workbook = load_workbook(report_path, data_only=True)
    try:
        summary_values = {
            row[0]: row[1]
            for row in workbook["Summary"].iter_rows(
                min_row=2,
                values_only=True,
            )
        }
        anomaly_rows = list(
            workbook["Anomaly"].iter_rows(min_row=2, values_only=True)
        )
        assert summary_values["Total Valid TXT Records"] == 2
        assert summary_values["Total Anomaly Records"] == 1
        assert summary_values["Duplicate TXT Records Removed"] == 1
        assert len(anomaly_rows) == 1
        assert anomaly_rows[0][7] == "DUPLICATE_REMOVED"
        assert anomaly_rows[0][9] == "machine-b.mdb"
        assert "machine-a.mdb" in anomaly_rows[0][8]
    finally:
        workbook.close()

    process_log = Path(result["artifact_result"]["process_log"])
    assert "Duplicate removed   : 1" in process_log.read_text(encoding="utf-8")

    summary_path = Path(result["artifact_result"]["summary_json"])
    summary_json = json.loads(summary_path.read_text(encoding="utf-8"))
    assert summary_json["duplicate_removed_count"] == 1
    assert summary_json["deduplication_summary"] == {
        "input_valid_records": 3,
        "unique_txt_records": 2,
        "duplicate_removed_records": 1,
    }
