from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook

from hris.assisted_calibrator import HRISAssistedCalibrator
from hris.click_profile import HRISClickProfileManager


class FakeAutomation:
    def size(self):
        return SimpleNamespace(width=1920, height=1080)

    def position(self):
        return SimpleNamespace(x=100, y=200)


class FakeBrowserManager:
    def __init__(self) -> None:
        self.opened = False
        self.closed = False

    def open_login_page(self):
        self.opened = True
        return SimpleNamespace(page=object())

    def close(self):
        self.closed = True


def _configuration(path: Path, profile_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "General"
    for name in ("Run_Control", "Browser", "Upload", "Reference"):
        workbook.create_sheet(name)
    for name in ("General", "Browser", "Upload", "Reference"):
        workbook[name].append(["Parameter", "Value", "Description"])
    workbook["General"].append(["HRIS_URL", "https://example.invalid", ""])
    workbook["Upload"].append(["Click_Profile_Path", str(profile_path), ""])
    workbook["Upload"].append(["Browser_X", 0, ""])
    workbook["Upload"].append(["Browser_Y", 0, ""])
    workbook["Upload"].append(["Browser_Width", 1200, ""])
    workbook["Upload"].append(["Browser_Height", 800, ""])
    workbook["Upload"].append(["Browser_Zoom", 100, ""])
    workbook["Run_Control"].append(
        ["Active", "Sequence", "Workflow", "Run_Control_ID", "Description"]
    )
    sheet = workbook.create_sheet("Assisted_Steps")
    sheet.append([
        "Active", "Sequence", "Step_Name", "Action", "Input_Source",
        "Method", "Required", "Wait_After_Seconds", "Description",
    ])
    sheet.append([
        "Y", 1, "run_control_id", "click_type", "RUN_CONTROL_ID",
        "coordinate", True, 0, "Run control",
    ])
    workbook.save(path)


def test_calibrator_opens_browser_then_captures_and_closes(
    tmp_path: Path,
) -> None:
    config_path = tmp_path / "config.xlsx"
    profile_path = tmp_path / "profile.json"
    _configuration(config_path, profile_path)
    browser = FakeBrowserManager()
    prompts = []

    calibrator = HRISAssistedCalibrator(
        config_path,
        instruction_callback=prompts.append,
        coordinate_callback=lambda message: (321, 456),
        browser_manager_factory=lambda configuration: browser,
        automation=FakeAutomation(),
    )
    result = calibrator.run()

    assert result == profile_path
    assert browser.opened and browser.closed
    assert "Overtime Upload Attendance" in prompts[0]
    profile = HRISClickProfileManager.load_profile(profile_path)
    assert profile["profile_version"] == "1.1"
    assert profile["steps"]["run_control_id"]["x"] == 321
