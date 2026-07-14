from datetime import date
from pathlib import Path

from openpyxl import Workbook

from utilities.attendance_reconciliation.attendance_reader import (
    AttendanceReportReader,
)
from utilities.attendance_reconciliation.outlook_reader import OutlookReportReader


def _attendance_report(path: Path, workflow: str = "HO") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    detail = workbook.active
    detail.title = "Attendance_Detail"
    detail.append([
        "NIK", "Nama", "Tanggal", "Jam Masuk", "Jam Keluar", "Tap Count",
        "Validation Status", "Validation Remarks", "Source MDB",
    ])
    detail.append([
        "001234567", "Synthetic", "07/01/2026", "08:00", "17:00", 2,
        "VALID_FOR_TXT", "", "synthetic.mdb",
    ])
    detail.append([
        "001234568", "Synthetic", "07/31/2026", "08:00", "17:00", 2,
        "VALID_FOR_TXT", "", "synthetic.mdb",
    ])
    anomaly = workbook.create_sheet("Anomaly")
    anomaly.append([
        "NIK", "Nama", "Tanggal", "Jam Masuk", "Jam Keluar", "Tap Count",
        "Pair Status", "Validation Status", "Validation Remarks", "Source MDB",
    ])
    anomaly.append([
        "001234569", "Synthetic", "07/15/2026", "18:30", None, 1,
        "SINGLE_TAP", "INVALID", "Only one tap", "synthetic.mdb",
    ])
    workbook.save(path)


def _outlook_report(path: Path, workflow: str = "HO") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    valid = workbook.active
    valid.title = "Valid_Data"
    valid.append([
        "No", "Attachment_Name", "Source_Row", "Workflow", "NIK",
        "Date_In", "Time_In", "Date_Out", "Time_Out", "Output_TXT",
    ])
    valid.append([
        1, "synthetic.xlsx", 2, workflow, "001234567", "07/01/2026",
        "08:00", "07/01/2026", "17:00", "out.txt",
    ])
    valid.append([
        2, "cross.xlsx", 3, workflow, "001234570", "07/10/2026",
        "08:00", "07/11/2026", "17:00", "out.txt",
    ])
    anomaly = workbook.create_sheet("Data_Anomaly")
    anomaly.append([
        "No", "Attachment_Name", "Source_Row", "Workflow", "NIK",
        "Date_In", "Time_In", "Date_Out", "Time_Out", "Anomaly_Code",
        "Anomaly_Reason", "Raw_Value",
    ])
    anomaly.append([
        1, "bad.xlsx", 4, workflow, "001234571", "07/05/2026", "bad",
        "07/05/2026", "17:00", "TIME_FORMAT", "Invalid time", "",
    ])
    workbook.save(path)


def test_recursive_attendance_scan_filters_period_and_interprets_single_tap(
    tmp_path: Path,
) -> None:
    report = (
        tmp_path / "Attendance" / "HO" / "job" / "Report"
        / "Export_Attendance_HO_job.xlsx"
    )
    _attendance_report(report)
    result = AttendanceReportReader().read_folder(
        tmp_path, "HO", date(2026, 7, 1), date(2026, 7, 15)
    )

    assert result.reports_used == 1
    assert len(result.records) == 2
    anomaly = next(item for item in result.records if item.is_anomaly)
    assert anomaly.internal_status == "SINGLE_TAP_OUT"
    assert anomaly.interpretation_rule == "TIME_GT_12_AS_OUT"
    assert anomaly.machine_in is None
    assert anomaly.machine_out.strftime("%H:%M") == "18:30"


def test_outlook_reader_marks_cross_date_invalid_and_keeps_anomaly_audit(
    tmp_path: Path,
) -> None:
    report = (
        tmp_path / "Outlook-Revisi" / "HO" / "job" / "Report"
        / "Outlook_Process_Report_HO_job.xlsx"
    )
    _outlook_report(report)
    result = OutlookReportReader().read_folder(
        tmp_path, "HO", date(2026, 7, 1), date(2026, 7, 15)
    )

    assert len(result.records) == 1
    assert len(result.invalid_records) == 1
    assert "CROSS_DATE_REVISION" in result.invalid_records[0].reason
    assert len(result.audit_anomalies) == 1


def test_wrong_workflow_temporary_attachment_and_corrupt_files_are_skipped(
    tmp_path: Path,
) -> None:
    wrong = (
        tmp_path / "Attendance" / "Branch" / "Report"
        / "Export_Attendance_Branch_job.xlsx"
    )
    _attendance_report(wrong, "Branch")
    attachment = (
        tmp_path / "Attachments" / "Report"
        / "Export_Attendance_HO_attachment.xlsx"
    )
    _attendance_report(attachment)
    corrupt = (
        tmp_path / "Attendance" / "HO" / "bad" / "Report"
        / "Export_Attendance_HO_bad.xlsx"
    )
    corrupt.parent.mkdir(parents=True)
    corrupt.write_text("not an xlsx", encoding="utf-8")
    temporary = corrupt.parent / "~$Export_Attendance_HO_lock.xlsx"
    temporary.write_text("locked", encoding="utf-8")

    result = AttendanceReportReader().read_folder(
        tmp_path, "HO", date(2026, 7, 1), date(2026, 7, 31)
    )
    statuses = {item.status for item in result.log_entries}

    assert "WRONG_WORKFLOW" in statuses
    assert "CORRUPT_WORKBOOK" in statuses
    assert "TEMPORARY_FILE" in statuses
    assert all("Attachments" not in str(item.file_path) for item in result.log_entries)
