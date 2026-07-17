import json
import re
from datetime import date
from pathlib import Path
from threading import Event

import pytest
from openpyxl import Workbook, load_workbook

from utilities.attendance_reconciliation.engine import ReconciliationEngine
from utilities.attendance_reconciliation.excel_writer import SHEET_ORDER
from utilities.attendance_reconciliation.excel_writer import status_color
from utilities.attendance_reconciliation.models import ALL_COMPARISON_STATUSES
from utilities.attendance_reconciliation.models import ReconciliationCancelled
from utilities.attendance_reconciliation.models import ReconciliationRequest
from utilities.attendance_reconciliation.models import SOURCE_MODE_SCAN


def _attendance(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    detail = workbook.active
    detail.title = "Attendance_Detail"
    detail.append([
        "NIK", "Nama", "Tanggal", "Jam Masuk", "Jam Keluar", "Tap Count",
        "Validation Status", "Validation Remarks", "Source MDB",
    ])
    detail.append([
        "001", "Synthetic", "07/01/2026", "08:00", "17:00", 2,
        "VALID_FOR_TXT", "", "synthetic.mdb",
    ])
    anomaly = workbook.create_sheet("Anomaly")
    anomaly.append([
        "NIK", "Nama", "Tanggal", "Jam Masuk", "Jam Keluar", "Tap Count",
        "Pair Status", "Validation Status", "Validation Remarks", "Source MDB",
    ])
    anomaly.append([
        "002", "Synthetic", "07/02/2026", "18:00", None, 1,
        "SINGLE_TAP", "INVALID", "single", "synthetic.mdb",
    ])
    workbook.save(path)


def _outlook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    valid = workbook.active
    valid.title = "Valid_Data"
    valid.append([
        "No", "Attachment_Name", "Source_Row", "Workflow", "NIK",
        "Date_In", "Time_In", "Date_Out", "Time_Out", "Output_TXT",
    ])
    valid.append([
        1, "same.xlsx", 2, "HO", "001", "07/01/2026", "08:00",
        "07/01/2026", "17:00", "out.txt",
    ])
    valid.append([
        2, "revision-only.xlsx", 3, "HO", "003", "07/03/2026", "09:00",
        "07/03/2026", "18:00", "out.txt",
    ])
    anomaly = workbook.create_sheet("Data_Anomaly")
    anomaly.append([
        "No", "Attachment_Name", "Source_Row", "Workflow", "NIK",
        "Date_In", "Time_In", "Date_Out", "Time_Out", "Anomaly_Code",
        "Anomaly_Reason", "Raw_Value",
    ])
    workbook.save(path)


def request(tmp_path: Path) -> ReconciliationRequest:
    attendance = (
        tmp_path / "attendance" / "HO" / "2026-07" / "job" / "Report"
        / "Export_Attendance_HO_job.xlsx"
    )
    outlook = (
        tmp_path / "outlook" / "HO" / "2026-07" / "job" / "Report"
        / "Outlook_Process_Report_HO_job.xlsx"
    )
    _attendance(attendance)
    _outlook(outlook)
    return ReconciliationRequest(
        SOURCE_MODE_SCAN, "HO", tmp_path / "attendance", tmp_path / "outlook",
        date(2026, 7, 1), date(2026, 7, 31), tmp_path / "output",
    )


def test_end_to_end_generates_workbook_log_and_summary(
    tmp_path: Path,
) -> None:
    result = ReconciliationEngine().run(request(tmp_path))

    assert result.success
    assert result.report_file and result.report_file.exists()
    month_label = f"{result.job_id[:4]}-{result.job_id[4:6]}"
    assert result.output_folder.parent == (
        tmp_path
        / "output"
        / "Utilities"
        / "Attendance-Reconciliation"
        / "HO"
        / month_label
    )
    workbook = load_workbook(result.report_file, data_only=True)
    assert tuple(workbook.sheetnames) == SHEET_ORDER
    guide_statuses = {
        row[0]
        for row in workbook["Guide_Status"].iter_rows(
            min_row=18, values_only=True
        )
        if row[0]
    }
    assert set(ALL_COMPARISON_STATUSES) <= guide_statuses
    assert list(workbook["Summary_Per_Karyawan"].values) == [
        (
            "NIK",
            "Nama",
            "Total Hari Attendance",
            "Total Hari Revisi",
            "Machine Complete",
            "Machine Anomaly",
            "Revision Match",
            "Revision Different",
            "Revision Only",
            "Conflict",
            "Total Perlu Review",
            "Status Akhir",
        ),
        ("001", "Synthetic", 1, 1, 1, 0, 1, 0, 0, 0, 0, "NORMAL"),
        (
            "002",
            "Synthetic",
            1,
            0,
            0,
            1,
            0,
            0,
            0,
            0,
            1,
            "REVIEW REQUIRED",
        ),
        ("003", None, 0, 1, 0, 0, 0, 0, 1, 0, 1, "REVIEW REQUIRED"),
    ]
    assert workbook["Summary_Per_Karyawan"]["L2"].fill.fgColor.rgb == (
        "00" + status_color("MACHINE_COMPLETE_NO_REVISION")
    )
    dashboard = dict(
        workbook["Dashboard"].iter_rows(
            min_row=2,
            values_only=True,
        )
    )
    assert dashboard["Process Started At"]
    assert dashboard["Process Completed At"]
    assert re.fullmatch(r"\d{2,}:\d{2}:\d{2}", dashboard["Process Duration"])
    assert dashboard["Process Duration Seconds"] >= 0
    comparison_sheet = workbook["Comparison_Detail"]
    assert comparison_sheet.freeze_panes == "A2"
    assert comparison_sheet.auto_filter.ref == "A1:AE4"
    assert comparison_sheet["A1"].fill.fgColor.rgb == "001F4E78"
    assert comparison_sheet["AA2"].fill.fgColor.rgb == (
        "00" + status_color(comparison_sheet["AA2"].value)
    )
    workbook.close()
    summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
    assert summary["attendance_records"] == 2
    assert summary["revision_only"] == 1
    assert summary["total_comparison_records"] == 3
    assert summary["process_started_at"]
    assert summary["process_completed_at"]
    assert re.fullmatch(r"\d{2,}:\d{2}:\d{2}", summary["process_duration"])
    assert summary["process_duration_seconds"] >= 0
    log = result.process_log.read_text(encoding="utf-8")
    assert "Job started" in log
    assert "Streaming Excel report completed" in log
    assert "Process duration:" in log
    assert "Job completed successfully" in log


def test_employee_summary_marks_mixed_normal_and_review_as_partial(
    tmp_path: Path,
) -> None:
    selected = request(tmp_path)
    attendance_report = next(
        selected.attendance_path.rglob("Export_Attendance_*.xlsx")
    )
    workbook = load_workbook(attendance_report)
    workbook["Anomaly"].append([
        "001",
        "Synthetic",
        "07/04/2026",
        "18:00",
        None,
        1,
        "SINGLE_TAP",
        "INVALID",
        "single",
        "synthetic.mdb",
    ])
    workbook.save(attendance_report)
    workbook.close()

    result = ReconciliationEngine().run(selected)

    workbook = load_workbook(result.report_file, data_only=True)
    rows = {
        row[0]: row
        for row in workbook["Summary_Per_Karyawan"].iter_rows(
            min_row=2,
            values_only=True,
        )
    }
    assert rows["001"] == (
        "001",
        "Synthetic",
        2,
        1,
        1,
        1,
        1,
        0,
        0,
        0,
        1,
        "PARTIAL",
    )
    assert (
        workbook["Summary_Per_Karyawan"]["L2"].fill.fgColor.rgb
        == "00FFF2CC"
    )
    workbook.close()


def test_status_color_mapping_uses_required_palette() -> None:
    assert status_color("MACHINE_COMPLETE_NO_REVISION") == "E2F0D9"
    assert status_color("MACHINE_COMPLETE_REVISION_DIFFERENT") == "FFF2CC"
    assert status_color("MACHINE_ANOMALY_NO_REVISION") == "FCE4D6"
    assert status_color("REVISION_ONLY") == "DDEBF7"
    assert status_color("INVALID_SOURCE_DATA") == "F4CCCC"


def test_stale_scan_is_rejected(tmp_path: Path) -> None:
    engine = ReconciliationEngine()
    original = request(tmp_path)
    scan = engine.scan(original)
    changed = ReconciliationRequest(
        original.source_mode, original.workflow, original.attendance_path,
        original.outlook_path, original.start_date, date(2026, 7, 30),
        original.output_folder,
    )
    with pytest.raises(ValueError, match="stale"):
        engine.run(changed, scan)


def test_cancelled_scan_stops_cooperatively(tmp_path: Path) -> None:
    engine = ReconciliationEngine()
    selected = request(tmp_path)
    cancelled = Event()
    cancelled.set()
    with pytest.raises(ReconciliationCancelled):
        engine.scan(selected, cancelled)


def test_cancelled_comparison_writes_cancel_log_without_report(
    tmp_path: Path,
) -> None:
    engine = ReconciliationEngine()
    selected = request(tmp_path)
    scan = engine.scan(selected)
    cancelled = Event()
    cancelled.set()

    with pytest.raises(ReconciliationCancelled):
        engine.run(selected, scan, cancelled)

    process_logs = list(
        (
            selected.output_folder
            / "Utilities"
            / "Attendance-Reconciliation"
            / "HO"
        ).rglob("Process.log")
    )
    assert len(process_logs) == 1
    job_folder = process_logs[0].parent
    assert job_folder.parent.name == (
        f"{job_folder.name[:4]}-{job_folder.name[4:6]}"
    )
    assert not list(job_folder.glob("Report/*.xlsx"))
    assert "CANCELLED" in process_logs[0].read_text(encoding="utf-8")


def test_specific_job_folder_mode_uses_only_selected_jobs(tmp_path: Path) -> None:
    selected = request(tmp_path)
    selected.source_mode = "Select Specific Job Folder"
    selected.attendance_path = (
        selected.attendance_path / "HO" / "2026-07" / "job"
    )
    selected.outlook_path = (
        selected.outlook_path / "HO" / "2026-07" / "job"
    )

    scan = ReconciliationEngine().scan(selected)

    assert scan.attendance.reports_found == 1
    assert scan.outlook.reports_found == 1


def test_scan_mode_supports_legacy_and_month_hierarchy(tmp_path: Path) -> None:
    selected = request(tmp_path)
    _attendance(
        selected.attendance_path
        / "HO"
        / "legacy-job"
        / "Report"
        / "Export_Attendance_HO_legacy.xlsx"
    )
    _outlook(
        selected.outlook_path
        / "HO"
        / "legacy-job"
        / "Report"
        / "Outlook_Process_Report_HO_legacy.xlsx"
    )

    scan = ReconciliationEngine().scan(selected)

    assert scan.attendance.reports_found == 2
    assert scan.outlook.reports_found == 2
