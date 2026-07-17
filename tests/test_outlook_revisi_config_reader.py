from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

from shared.config_manager import OutlookRevisiConfigurationReader


def _append_title_and_header(sheet, title: str, headers: list[str]) -> None:
    sheet.append([title])
    sheet.append(["Instruction row"])
    sheet.append([])
    sheet.append(headers)


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "General"

    for name in (
        "HO_Sender_Master",
        "Branch_Sender_Master",
        "Subject_Rules",
        "Attachment_Rules",
        "Validation_Rules",
        "Reply_Templates",
    ):
        workbook.create_sheet(name)

    _append_title_and_header(
        workbook["General"],
        "Outlook - Revisi Configuration",
        ["Parameter", "Value", "Description"],
    )
    workbook["General"].append(
        ["Mailbox_SMTP", "karina.hr.1@oto.co.id", ""]
    )
    workbook["General"].append(
        ["Output_Root", r"output\Outlook-Revisi", ""]
    )
    workbook["General"].append(
        ["PIC_HR_Emails", "hr.one@example.com; hr.two@example.com", ""]
    )

    _append_title_and_header(
        workbook["HO_Sender_Master"],
        "HO Sender Master",
        ["Active", "Sender_Name", "Sender_Email", "Required_CC_Email"],
    )
    workbook["HO_Sender_Master"].append(
        ["Y", "HO User", "ho@example.com", "cc@example.com"]
    )
    workbook["HO_Sender_Master"].append(["N", "Inactive", "", ""])
    workbook["HO_Sender_Master"].append(["Y", "", "", ""])

    _append_title_and_header(
        workbook["Branch_Sender_Master"],
        "Branch Sender Master",
        [
            "Active",
            "Company",
            "Branch_Code",
            "Sender_Name",
            "Sender_Email",
            "Required_CC_Email",
        ],
    )
    workbook["Branch_Sender_Master"].append(
        ["TRUE", "OTO", "AMU", "Branch User", "amu@example.com", "bm@example.com"]
    )

    _append_title_and_header(
        workbook["Subject_Rules"],
        "Subject Rules",
        ["Active", "Workflow", "Subject_Pattern"],
    )
    workbook["Subject_Rules"].append(["Y", "HO", "ATT_REV {PERIOD}"])
    workbook["Subject_Rules"].append(
        ["Y", "Branch", "[{COMPANY}-{BRANCH_CODE}] Attendance {PERIOD}"]
    )

    _append_title_and_header(
        workbook["Attachment_Rules"],
        "Attachment Rules",
        ["Active", "Workflow", "Allowed_Extensions"],
    )
    workbook["Attachment_Rules"].append(["Y", "HO", ".xlsx;.xls"])
    workbook["Attachment_Rules"].append(["Y", "Branch", ".txt"])

    _append_title_and_header(
        workbook["Validation_Rules"],
        "Validation Rules",
        ["Active", "Rule_Code", "Workflow", "Rule_Value"],
    )
    workbook["Validation_Rules"].append(["Y", "NIK_REQUIRED", "All", "TRUE"])

    _append_title_and_header(
        workbook["Reply_Templates"],
        "Reply Templates",
        [
            "Active",
            "Reply_Code",
            "Recipient_Type",
            "Trigger",
            "Subject_Template",
            "Body_Template",
        ],
    )
    workbook["Reply_Templates"].append(
        ["Y", "SUCCESS_SENDER", "SENDER", "PROCESS_SUCCESS", "Re: {ORIGINAL_SUBJECT}", "OK"]
    )

    workbook.save(path)


def test_reads_outlook_revisi_configuration_with_title_rows(tmp_path: Path) -> None:
    path = tmp_path / "config" / "outlook" / "outlook.xlsx"
    path.parent.mkdir(parents=True)
    _create_workbook(path)

    configuration = OutlookRevisiConfigurationReader(path).read()

    assert configuration.general["Mailbox_SMTP"] == "karina.hr.1@oto.co.id"
    assert configuration.get_output_root() == tmp_path / "output" / "Outlook-Revisi"
    assert configuration.get_pic_hr_emails() == [
        "hr.one@example.com",
        "hr.two@example.com",
    ]
    assert len(configuration.ho_senders) == 1
    assert configuration.ho_senders[0].sender_email == "ho@example.com"
    assert configuration.branch_senders[0].branch_code == "AMU"
    assert configuration.attachment_rules[0].allowed_extensions == [
        ".xlsx",
        ".xls",
    ]
    assert configuration.validation_rules[0].workflow == "All"
    assert configuration.reply_templates[0].trigger == "PROCESS_SUCCESS"


def test_output_root_named_output_adds_outlook_revisi_folder(
    tmp_path: Path,
) -> None:
    path = tmp_path / "config" / "outlook" / "outlook.xlsx"
    path.parent.mkdir(parents=True)
    _create_workbook(path)

    workbook = load_workbook(path)
    workbook["General"]["B6"] = "output"
    workbook.save(path)
    workbook.close()

    configuration = OutlookRevisiConfigurationReader(path).read()

    assert configuration.get_output_root() == (
        tmp_path / "output" / "Outlook-Revisi"
    )


def test_rejects_missing_outlook_sheet(tmp_path: Path) -> None:
    path = tmp_path / "config.xlsx"
    workbook = Workbook()
    workbook.active.title = "General"
    workbook.save(path)

    with pytest.raises(ValueError, match="Missing required Outlook"):
        OutlookRevisiConfigurationReader(path).read()


def test_reads_real_outlook_revisi_workbook_when_available() -> None:
    candidates = [
        Path("config/outlook/OAS-K_Outlook-Revisi_Configuration.xlsx"),
        Path("dist/config/outlook/OAS-K_Outlook-Revisi_Configuration.xlsx"),
        Path(r"C:/Users/User/Downloads/OAS-K_Outlook-Revisi_Configuration.xlsx"),
    ]
    path = next((item for item in candidates if item.exists()), None)

    if path is None:
        pytest.skip("Outlook - Revisi workbook is not available.")

    configuration = OutlookRevisiConfigurationReader(path).read()

    assert configuration.general["Integration_Method"] == "OOM_COM"
    assert configuration.general["Mailbox_SMTP"] == "karina.hr.1@oto.co.id"
    assert len(configuration.subject_rules) == 2
    assert len(configuration.attachment_rules) == 2
    assert len(configuration.validation_rules) >= 1
    assert len(configuration.reply_templates) >= 1
