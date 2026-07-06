"""
Temporary test file for Sprint 6.21 HRIS Full Upload Engine.

Run from project root:
py test_hris_full_upload_engine_local.py
"""

from __future__ import annotations

from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook

from hris.engine import HRISFullUploadEngine


CONFIG_FILE = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\config\hris\OAS-K_HRIS_Configuration.xlsx"
)

TEMP_CONFIG_FILE = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\temp\OAS-K_HRIS_Configuration_Full_Engine_Test.xlsx"
)

SAMPLE_TXT_FOLDER = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\temp\hris_txt_sample_full_engine"
)

OUTPUT_ROOT = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\output"
)

MOCK_SITE_FOLDER = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\temp\hris_mock_site"
)

MOCK_UPLOAD_FILE = MOCK_SITE_FOLDER / "full_engine_upload.html"

WORKFLOW = "HO"
START_DATE = "03/30/2026"
END_DATE = "03/30/2026"


def prepare_sample_txt_files() -> None:
    SAMPLE_TXT_FOLDER.mkdir(
        parents=True,
        exist_ok=True,
    )

    sample_files = [
        "Attendance_HO_001.txt",
        "Attendance_HO_002.txt",
        "Attendance_HO_003.txt",
    ]

    for file_name in sample_files:
        file_path = SAMPLE_TXT_FOLDER / file_name
        file_path.write_text(
            '"03/30/2026","100808219","03/30/2026","13:20","03/30/2026","16:39"\n',
            encoding="utf-8",
        )


def prepare_mock_upload_page() -> None:
    MOCK_SITE_FOLDER.mkdir(
        parents=True,
        exist_ok=True,
    )

    html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Mock HRIS Full Engine - OAS-K</title>
</head>
<body>
    <h1>Mock HRIS Home</h1>

    <button onclick="document.getElementById('tao').style.display='block';">
        Time Attendance & Overtime
    </button>

    <div id="tao" style="display:none;">
        <button onclick="document.getElementById('upload').style.display='block';">
            Overtime Upload Attendance
        </button>
    </div>

    <div id="upload" style="display:none;">
        <h2>Upload Attendance Page</h2>

        <input id="run_control_id" type="text">
        <input id="start_date" type="text">
        <input id="end_date" type="text">
        <input id="attachment_file" type="file">

        <button id="upload_button" onclick="document.getElementById('upload_ok_button').style.display='block';">
            Upload
        </button>

        <button id="upload_ok_button" style="display:none;" onclick="document.getElementById('run_button').style.display='block';">
            OK
        </button>

        <button id="run_button" style="display:none;" onclick="document.getElementById('run_ok_button').style.display='block';">
            Run
        </button>

        <button id="run_ok_button" style="display:none;" onclick="document.getElementById('success_message').style.display='block';">
            OK
        </button>

        <p id="success_message" style="display:none;">
            Upload simulated successfully.
        </p>
    </div>
</body>
</html>
"""

    MOCK_UPLOAD_FILE.write_text(
        html,
        encoding="utf-8",
    )


def prepare_temp_config() -> Path:
    TEMP_CONFIG_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    copy2(
        CONFIG_FILE,
        TEMP_CONFIG_FILE,
    )

    local_url = MOCK_UPLOAD_FILE.resolve().as_uri()

    workbook = load_workbook(TEMP_CONFIG_FILE)
    sheet = workbook["General"]

    for row in sheet.iter_rows(min_row=2):
        parameter = row[0].value

        if str(parameter).strip() == "HRIS_URL":
            row[1].value = local_url
            break

    workbook.save(TEMP_CONFIG_FILE)
    workbook.close()

    print(f"Temporary HRIS_URL set to: {local_url}")
    print(f"Temporary config file     : {TEMP_CONFIG_FILE}")

    return TEMP_CONFIG_FILE


def main() -> None:
    print("=" * 80)
    print("OAS-K Sprint 6.21 HRIS Full Upload Engine Local Test")
    print("=" * 80)

    prepare_sample_txt_files()
    prepare_mock_upload_page()
    temp_config_file = prepare_temp_config()

    engine = HRISFullUploadEngine(
        configuration_file=temp_config_file,
        txt_folder=SAMPLE_TXT_FOLDER,
        output_root=OUTPUT_ROOT,
        workflow=WORKFLOW,
        start_date=START_DATE,
        end_date=END_DATE,
        wait_for_manual_login=False,
    )

    result = engine.run()

    print()
    print("FULL UPLOAD ENGINE RESULT")
    print("-" * 80)
    print(f"Success             : {result.success}")
    print(f"Message             : {result.message}")
    print(f"Job ID              : {result.job_id}")
    print(f"Workflow            : {result.workflow}")
    print(f"Success Count       : {result.success_count}")
    print(f"Failed Count        : {result.failed_count}")
    print(f"Report Folder       : {result.report_folder}")
    print(f"Report File         : {result.report_file}")
    print(f"Summary JSON File   : {result.summary_json_file}")
    print(f"Process Log File    : {result.process_log_file}")

    print()
    print("=" * 80)
    print("HRIS Full Upload Engine local test finished.")
    print("=" * 80)


if __name__ == "__main__":
    main()  