"""
Temporary test file for Sprint 6.20 HRIS Full Local Upload Simulation.

This test uses local mock HRIS upload page.
It does not access real HRIS URL.

Run from project root:
py test_hris_full_local_simulation.py
"""

from __future__ import annotations

from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook

from hris.artifact_writer import HRISJobArtifactWriter
from hris.batch_uploader import HRISBatchUploader
from hris.browser import HRISBrowserManager
from hris.file_manager import HRISFileManager
from hris.job_manager import HRISUploadJobManager
from hris.navigator import HRISNavigator
from hris.report_writer import HRISUploadReportWriter
from shared.config_manager import HRISConfigurationReader


CONFIG_FILE = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\config\hris\OAS-K_HRIS_Configuration.xlsx"
)

TEMP_CONFIG_FILE = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\temp\OAS-K_HRIS_Configuration_Full_Local_Test.xlsx"
)

SAMPLE_TXT_FOLDER = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\temp\hris_txt_sample_full_local"
)

OUTPUT_ROOT = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\output"
)

MOCK_SITE_FOLDER = Path(
    r"D:\Python Project\OfficeAutomationSuite-Karina\temp\hris_mock_site"
)

MOCK_UPLOAD_FILE = MOCK_SITE_FOLDER / "full_local_upload.html"

WORKFLOW = "HO"
START_DATE = "03/30/2026"
END_DATE = "03/30/2026"


def prepare_sample_txt_files() -> None:
    """
    Create sample TXT files for full local simulation.
    """
    SAMPLE_TXT_FOLDER.mkdir(
        parents=True,
        exist_ok=True,
    )

    sample_files = [
        "Attendance_HO_001.txt",
        "Attendance_HO_002.txt",
        "Attendance_HO_003.txt",
    ]

    for file_name in sample_files:
        file_path = SAMPLE_TXT_FOLDER / file_name
        file_path.write_text(
            '"03/30/2026","100808219","03/30/2026","13:20","03/30/2026","16:39"\n',
            encoding="utf-8",
        )


def prepare_mock_upload_page() -> None:
    """
    Create local mock HRIS upload page.
    """
    MOCK_SITE_FOLDER.mkdir(
        parents=True,
        exist_ok=True,
    )

    html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Mock HRIS Full Local Upload - OAS-K</title>
    <style>
        body { font-family: Arial, sans-serif; background: #f5f5f5; padding: 40px; }
        .card { width: 680px; margin: auto; background: white; border: 1px solid #ddd; border-radius: 8px; padding: 24px; }
        h1 { color: #1F4E78; font-size: 22px; }
        button { display: block; width: 100%; margin-top: 12px; padding: 10px; background: #1F4E78; color: white; border: none; border-radius: 4px; font-weight: bold; cursor: pointer; }
        input { width: 100%; padding: 8px; margin-top: 6px; box-sizing: border-box; }
        label { display: block; margin-top: 12px; font-weight: bold; }
        .section { margin-top: 18px; padding: 14px; border: 1px solid #ddd; display: none; }
        .success { color: #2E8B57; font-weight: bold; display: none; }
    </style>
</head>
<body>
    <div class="card">
        <h1>Mock HRIS Home</h1>
        <p>Local full simulation page for OAS-K Sprint 6.20.</p>

        <button onclick="document.getElementById('tao').style.display='block';">
            Time Attendance & Overtime
        </button>

        <div id="tao" class="section">
            <button onclick="document.getElementById('upload').style.display='block';">
                Overtime Upload Attendance
            </button>
        </div>

        <div id="upload" class="section">
            <h2>Upload Attendance Page</h2>

            <label>Run Control ID</label>
            <input id="run_control_id" type="text">

            <label>Start Date</label>
            <input id="start_date" type="text">

            <label>End Date</label>
            <input id="end_date" type="text">

            <label>Attachment</label>
            <input id="attachment_file" type="file">

            <button id="upload_button" onclick="document.getElementById('upload_ok_button').style.display='block';">
                Upload
            </button>

            <button id="upload_ok_button" style="display:none;" onclick="document.getElementById('run_button').style.display='block';">
                OK
            </button>

            <button id="run_button" style="display:none;" onclick="document.getElementById('run_ok_button').style.display='block';">
                Run
            </button>

            <button id="run_ok_button" style="display:none;" onclick="document.getElementById('success_message').style.display='block';">
                OK
            </button>

            <p id="success_message" class="success">
                Upload simulated successfully.
            </p>
        </div>
    </div>
</body>
</html>
"""

    MOCK_UPLOAD_FILE.write_text(
        html,
        encoding="utf-8",
    )


def prepare_temp_config() -> Path:
    """
    Create temporary HRIS config copy with local mock URL.
    """
    TEMP_CONFIG_FILE.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    copy2(
        CONFIG_FILE,
        TEMP_CONFIG_FILE,
    )

    local_url = MOCK_UPLOAD_FILE.resolve().as_uri()

    workbook = load_workbook(TEMP_CONFIG_FILE)
    sheet = workbook["General"]

    for row in sheet.iter_rows(min_row=2):
        parameter = row[0].value

        if str(parameter).strip() == "HRIS_URL":
            row[1].value = local_url
            break

    workbook.save(TEMP_CONFIG_FILE)
    workbook.close()

    print(f"Temporary HRIS_URL set to: {local_url}")
    print(f"Temporary config file     : {TEMP_CONFIG_FILE}")

    return TEMP_CONFIG_FILE


def main() -> None:
    print("=" * 80)
    print("OAS-K Sprint 6.20 HRIS Full Local Upload Simulation")
    print("=" * 80)

    prepare_sample_txt_files()
    prepare_mock_upload_page()
    temp_config_file = prepare_temp_config()

    reader = HRISConfigurationReader(temp_config_file)
    configuration = reader.read()

    job_manager = HRISUploadJobManager()
    artifact_writer = HRISJobArtifactWriter()
    report_writer = HRISUploadReportWriter()
    file_manager = HRISFileManager()

    upload_plan = job_manager.create_upload_plan(
        configuration=configuration,
        txt_folder=SAMPLE_TXT_FOLDER,
        workflow=WORKFLOW,
    )

    if not upload_plan.is_valid:
        print("Upload plan is invalid.")
        print(upload_plan.message)
        return

    job_id = job_manager.create_job_id(WORKFLOW)

    artifacts = artifact_writer.prepare_job_artifacts(
        output_root=OUTPUT_ROOT,
        workflow=WORKFLOW,
        job_id=job_id,
        upload_plan=upload_plan,
    )

    report_file = report_writer.write_upload_report(
        artifacts=artifacts,
        upload_plan=upload_plan,
    )

    artifact_writer.write_process_log(
        artifacts=artifacts,
        message="Full local simulation started.",
    )

    browser_manager = HRISBrowserManager(
        configuration=configuration,
    )

    try:
        session = browser_manager.open_login_page()

        artifact_writer.write_process_log(
            artifacts=artifacts,
            message="Browser opened.",
        )

        navigator = HRISNavigator(
            page=session.page,
        )

        navigator.open_overtime_upload_attendance()

        artifact_writer.write_process_log(
            artifacts=artifacts,
            message="Upload attendance page opened.",
        )

        batch_uploader = HRISBatchUploader(
            page=session.page,
        )

        batch_result = batch_uploader.upload_batch(
            upload_plan=upload_plan,
            start_date=START_DATE,
            end_date=END_DATE,
            stop_on_first_failure=True,
        )

        artifact_writer.update_summary_after_batch(
            artifacts=artifacts,
            batch_result=batch_result,
        )

        move_results = file_manager.move_uploaded_files(
            artifacts=artifacts,
            plan_items=batch_result.results,
        )

        artifact_writer.update_summary_after_file_move(
            artifacts=artifacts,
            move_results=move_results,
        )

        summary = artifact_writer.read_summary(
            artifacts.summary_json_file,
        )

        report_writer.update_report_after_upload(
            report_file=report_file,
            summary=summary,
        )

        artifact_writer.write_process_log(
            artifacts=artifacts,
            message="Full local simulation finished.",
        )

        print()
        print("FULL LOCAL SIMULATION RESULT")
        print("-" * 80)
        print(f"Job ID              : {job_id}")
        print(f"Status              : {summary.get('status')}")
        print(f"Progress Percent    : {summary.get('progress_percent')}")
        print(f"Processed Items     : {summary.get('processed_items')}")
        print(f"Success             : {summary.get('success')}")
        print(f"Failed              : {summary.get('failed')}")
        print(f"Skipped             : {summary.get('skipped')}")
        print(f"Upload Folder       : {artifacts.upload_folder}")
        print(f"Failed Folder       : {artifacts.failed_folder}")
        print(f"Report File         : {report_file}")
        print(f"Report Exists       : {report_file.exists()}")
        print(f"Summary JSON Exists : {artifacts.summary_json_file.exists()}")
        print(f"Process Log Exists  : {artifacts.process_log_file.exists()}")

        print()
        print("FINAL PLAN ITEMS")
        print("-" * 80)

        for item in summary.get("plan_items", []):
            print(
                f"{item.get('sequence')}. "
                f"{item.get('txt_file_name')} -> "
                f"Run Control {item.get('run_control_id')} | "
                f"{item.get('status')} | "
                f"Moved To: {item.get('moved_to')}"
            )

    finally:
        print()
        print("Closing browser session...")
        browser_manager.close()

    print()
    print("=" * 80)
    print("HRIS Full Local Upload Simulation finished.")
    print("=" * 80)


if __name__ == "__main__":
    main()