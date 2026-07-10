"""Idempotently add HRIS assisted automation settings to the existing workbook."""

from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


WORKBOOK = Path("config/hris/OAS-K_HRIS_Configuration.xlsx")
BACKUP = Path("config/hris/OAS-K_HRIS_Configuration.pre_assisted_backup.xlsx")

UPLOAD_PARAMETERS = (
    ("Click_Profile_Path", r"config\hris\HRIS_Click_Profile.json", "Path hasil kalibrasi assisted action"),
    ("Assisted_Mode_Enabled", "TRUE", "Use recorder only from OK after Upload; earlier steps use Playwright"),
    ("Manual_Recovery_Enabled", "TRUE", "Pause and allow operator recovery if step fails"),
    ("Require_Profile_Match", "TRUE", "Validate screen/browser profile before replay"),
    ("Use_Date_From_Config", "FALSE", "If TRUE, GUI uses Start_Date and End_Date from config"),
    ("Start_Date", "03/30/2026", "Optional default start date for HRIS upload"),
    ("End_Date", "03/30/2026", "Optional default end date for HRIS upload"),
    ("Browser_X", 0, "Fixed Edge window X"),
    ("Browser_Y", 0, "Fixed Edge window Y"),
    ("Browser_Width", 1200, "Fixed Edge window width"),
    ("Browser_Height", 800, "Fixed Edge window height"),
    ("Browser_Zoom", 100, "Expected browser zoom percentage"),
    ("Assisted_Verification_Enabled", "TRUE", "Verify HRIS submission result after Run and OK"),
    ("Verification_Wait_Seconds", 1, "Initial wait before reading HRIS result text"),
    ("Verification_Timeout_Seconds", 10, "Maximum automatic verification time per item"),
    ("Verification_Poll_Seconds", 1, "Interval between HRIS result checks"),
    ("Verification_Success_Texts", "Process Instance|Submitted|Queued", "Pipe-separated HRIS submission success phrases"),
    ("Verification_Failure_Texts", "Error|Invalid|Failed", "Pipe-separated HRIS failure phrases"),
    ("Manual_Verification_On_Unknown", "TRUE", "Ask operator when automatic verification is inconclusive"),
    ("Manual_Verification_On_Error", "TRUE", "Ask operator when an HRIS error phrase is detected"),
)

HEADERS = (
    "Active", "Sequence", "Step_Name", "Action", "Input_Source",
    "Method", "Required", "Wait_After_Seconds", "Description",
)

STEPS = (
    ("Y", 1, "ok_after_upload", "click", "NONE", "coordinate", True, 1, "Click OK after Upload"),
    ("Y", 2, "run", "click", "NONE", "coordinate", True, 2, "Click Run"),
    ("Y", 3, "ok_after_run", "click", "NONE", "coordinate", True, 1, "Click OK after Run"),
)


def copy_row_style(sheet, source_row: int, target_row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)


def update_upload_sheet(workbook) -> None:
    sheet = workbook["Upload"]
    existing = {
        str(sheet.cell(row, 1).value).strip(): row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, 1).value
    }
    style_row = 2 if sheet.max_row >= 2 else 1
    for parameter, value, description in UPLOAD_PARAMETERS:
        row = existing.get(parameter)
        if row is None:
            row = sheet.max_row + 1
            copy_row_style(sheet, style_row, row, 3)
        sheet.cell(row, 1).value = parameter
        sheet.cell(row, 2).value = value
        sheet.cell(row, 3).value = description
    sheet.auto_filter.ref = f"A1:C{sheet.max_row}"
    sheet.freeze_panes = "A2"


def update_assisted_steps_sheet(workbook) -> None:
    if "Assisted_Steps" in workbook.sheetnames:
        sheet = workbook["Assisted_Steps"]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet("Assisted_Steps")

    sheet.append(HEADERS)
    for row in STEPS:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = (10, 10, 22, 16, 20, 14, 11, 20, 64)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    sheet.row_dimensions[1].height = 24
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:I{sheet.max_row}"

    action_validation = DataValidation(
        type="list",
        formula1='"click,click_type,type,press,attach_file,wait,manual_continue"',
    )
    method_validation = DataValidation(
        type="list",
        formula1='"coordinate,playwright,manual,assisted"',
    )
    input_validation = DataValidation(
        type="list",
        formula1='"NONE,RUN_CONTROL_ID,START_DATE,END_DATE,TXT_FILE_PATH"',
    )
    required_validation = DataValidation(type="list", formula1='"TRUE,FALSE"')
    active_validation = DataValidation(type="list", formula1='"Y,N"')
    for validation, target in (
        (active_validation, "A2:A500"),
        (action_validation, "D2:D500"),
        (input_validation, "E2:E500"),
        (method_validation, "F2:F500"),
        (required_validation, "G2:G500"),
    ):
        sheet.add_data_validation(validation)
        validation.add(target)

    inactive_fill = PatternFill("solid", fgColor="E7E6E6")
    sheet.conditional_formatting.add(
        f"A2:I{sheet.max_row}",
        FormulaRule(formula=['$A2="N"'], fill=inactive_fill),
    )


def main() -> None:
    if not WORKBOOK.exists():
        raise FileNotFoundError(WORKBOOK)
    if not BACKUP.exists():
        shutil.copy2(WORKBOOK, BACKUP)

    workbook = load_workbook(WORKBOOK)
    update_upload_sheet(workbook)
    update_assisted_steps_sheet(workbook)
    workbook.save(WORKBOOK)
    workbook.close()


if __name__ == "__main__":
    main()
